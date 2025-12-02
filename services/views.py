from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponseNotFound, HttpResponseServerError, HttpResponse, JsonResponse
from django.db.models import Q
from django.db import models
from decimal import Decimal
from .models import Service, UserProfile, Order, Sponsor, PaymentMethod, CustomService, Chat, Message, ProfileChange, ServiceRequestModal
from datetime import timedelta


# Chat views are now in chat_views.py and imported directly in urls.py

# Import health check
from .health import health_check

@login_required
def ai_dashboard(request):
    """AI Dashboard for continuous website improvement"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Import AI modules
    from .ml_analytics import WebsiteOptimizer
    from .personalization import PersonalizationEngine
    from .content_generator import ContentGenerator
    
    # Initialize AI modules
    optimizer = WebsiteOptimizer()
    personalization = PersonalizationEngine()
    content_generator = ContentGenerator()
    
    # Get analytics data
    analytics_data = optimizer.analyze_user_behavior()
    
    # Get AI suggestions
    suggestions = optimizer.suggest_improvements()
    
    # Get user preferences for sample user
    user_preferences = personalization.get_user_preferences("sample_user")
    
    # Generate sample content
    sample_content = content_generator.generate_service_description("Elétrica")
    
    # Example of how to record user interaction data
    # In a real application, this would be done automatically by the middleware
    # optimizer.collect_user_data(
    #     user_id=str(request.user.id),
    #     page_visited=request.path,
    #     time_spent=120,  # This would be calculated automatically
    #     actions_taken=["view", "click"],
    #     converted=False
    # )
    
    return render(request, 'services/ai_dashboard.html', {
        'analytics_data': analytics_data,
        'suggestions': suggestions,
        'user_preferences': user_preferences,
        'sample_content': sample_content
    })


@login_required
def profile_view(request):
    """User profile page with complete profile information, order history, and settings"""
    try:
        # Get or create user profile
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Get user's orders
        user_orders = Order.objects.filter(customer=request.user).order_by('-created_at')
        
        # Get user's payment methods
        payment_methods = PaymentMethod.objects.filter(user=request.user)
        
        # Handle POST requests for profile updates
        if request.method == 'POST':
            # Update profile information
            user_profile.phone = request.POST.get('phone', user_profile.phone)
            user_profile.address = request.POST.get('address', user_profile.address)
            user_profile.number = request.POST.get('number', user_profile.number)
            user_profile.complement = request.POST.get('complement', user_profile.complement)
            user_profile.city = request.POST.get('city', user_profile.city)
            user_profile.state = request.POST.get('state', user_profile.state)
            user_profile.zip_code = request.POST.get('zip_code', user_profile.zip_code)
            user_profile.birth_date = request.POST.get('birth_date', user_profile.birth_date)
            
            # Handle avatar upload
            if 'avatar' in request.FILES:
                user_profile.avatar = request.FILES['avatar']
            
            user_profile.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('profile')
        
        context = {
            'user_profile': user_profile,
            'user_orders': user_orders,
            'payment_methods': payment_methods,
        }
        return render(request, 'services/profile.html', context)
    except Exception as e:
        messages.error(request, f'Erro ao carregar o perfil: {str(e)}')
        return redirect('home')


@login_required
def my_orders(request):
    """Redireciona para a página de Meus Pedidos (meus_pedidos)"""
    return redirect('meus_pedidos')


@login_required
def order_details(request, order_id):
    """Detailed view of a specific order"""
    try:
        order = Order.objects.select_related(
            'professional', 'professional__userprofile', 'service', 'payment_method'
        ).prefetch_related('review').get(id=order_id, customer=request.user)
        
        context = {
            'order': order,
        }
        
        return render(request, 'services/order_details.html', context)
        
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado')
        return redirect('my_orders')
    except Exception as e:
        messages.error(request, f'Erro ao carregar detalhes do pedido: {str(e)}')
        return redirect('my_orders')


@login_required
def cancel_order(request, order_id):
    """Cancel an order"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        order = Order.objects.get(id=order_id, customer=request.user)
        
        # Only allow cancellation of pending or confirmed orders
        if order.status not in ['pending', 'confirmed']:
            return JsonResponse({'error': 'Este pedido não pode ser cancelado'}, status=400)
        
        order.status = 'cancelled'
        order.save()
        
        return JsonResponse({'success': True, 'message': 'Pedido cancelado com sucesso'})
        
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Pedido não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Erro ao cancelar pedido'}, status=500)


# Example of how to use AI modules in other views
def home(request):
    """Home page for regular users - shows available services"""
    # If user is authenticated and is a provider, redirect to provider dashboard
    if request.user.is_authenticated:
        try:
            user_profile = request.user.userprofile
            if user_profile.user_type == 'professional':
                return redirect('provider_dashboard')
        except Exception as e:
            # Handle the case where the database columns don't exist yet
            if "no such column" in str(e):
                # Create a default user profile if columns are missing
                try:
                    user_profile = UserProfile.objects.create(user=request.user)
                except:
                    pass
            pass
    
    services = Service.objects.filter(is_active=True)
    custom_services = CustomService.objects.filter(is_active=True)
    
    # Example of personalization in action
    # In a real application, this would be done by the context processor
    personalized_content = None
    if request.user.is_authenticated:
        try:
            from .personalization import PersonalizationEngine
            personalization = PersonalizationEngine()
            user_id = str(request.user.id)
            
            # Record user interaction (this would normally be done by middleware)
            personalization.record_user_interaction(
                user_id=user_id,
                page_visited="/",
                interaction_type="view"
            )
            
            # Get personalized recommendations
            recommendations = personalization.recommend_content(user_id, current_page="/")
            if recommendations:
                # For demonstration, we'll just use the first recommendation
                personalized_content = recommendations[0] if recommendations else None
        except Exception:
            # Silently fail to avoid breaking the application
            pass

    
    return render(request, 'services/home.html', {
        'services': services,
        'custom_services': custom_services,
        'personalized_content': personalized_content
    })


def providers_by_service(request, service_category):
    """Show providers by service category"""
    # Get all providers who have offered custom services in this category
    providers_queryset = User.objects.filter(
        custom_services__category=service_category,
        custom_services__is_active=True
    ).distinct()
    
    # Apply rating filter if provided
    rating_filter = request.GET.get('rating')
    if rating_filter:
        try:
            min_rating = float(rating_filter)
            # Filter providers with rating >= min_rating
            providers_queryset = providers_queryset.filter(userprofile__rating__gte=min_rating)
        except (ValueError, TypeError):
            pass
    
    # Get the category display name
    category_name = dict(Service.CATEGORY_CHOICES).get(service_category, service_category)
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(providers_queryset, 12)  # Show 12 providers per page
    page_number = request.GET.get('page')
    providers = paginator.get_page(page_number)
    
    return render(request, 'services/providers_by_service.html', {
        'providers': providers,
        'category': service_category,
        'category_name': category_name,
        'rating_filter': rating_filter
    })


@login_required
def all_professionals(request):
    """Show all professionals across all service categories with enhanced filtering and search"""
    from django.db.models import Avg, Count, Min, Max
    
    # Base queryset with optimized database queries
    providers_queryset = User.objects.filter(
        userprofile__user_type='professional',
        custom_services__is_active=True
    ).select_related('userprofile').prefetch_related('custom_services').distinct()
    
    # Enhanced search functionality
    search_filter = request.GET.get('search', '').strip()
    if search_filter:
        providers_queryset = providers_queryset.filter(
            Q(first_name__icontains=search_filter) | 
            Q(last_name__icontains=search_filter) | 
            Q(username__icontains=search_filter) |
            Q(userprofile__business_name__icontains=search_filter) |
            Q(userprofile__bio__icontains=search_filter) |
            Q(userprofile__specialties__icontains=search_filter) |
            Q(custom_services__name__icontains=search_filter) |
            Q(custom_services__description__icontains=search_filter)
        )
    
    # Category filter
    category_filter = request.GET.get('category')
    if category_filter:
        providers_queryset = providers_queryset.filter(custom_services__category=category_filter)
    
    # Rating filter
    rating_filter = request.GET.get('rating')
    if rating_filter:
        try:
            min_rating = float(rating_filter)
            providers_queryset = providers_queryset.filter(userprofile__rating__gte=min_rating)
        except (ValueError, TypeError):
            pass
    
    # Price filters with better handling
    price_min = request.GET.get('price_min', '').strip()
    price_max = request.GET.get('price_max', '').strip()
    
    if price_min:
        try:
            min_price = float(price_min)
            providers_queryset = providers_queryset.filter(custom_services__estimated_price__gte=min_price)
        except (ValueError, TypeError):
            pass
    
    if price_max:
        try:
            max_price = float(price_max)
            providers_queryset = providers_queryset.filter(custom_services__estimated_price__lte=max_price)
        except (ValueError, TypeError):
            pass
    
    # Enhanced location filter
    location_filter = request.GET.get('location', '').strip()
    if location_filter:
        providers_queryset = providers_queryset.filter(
            Q(userprofile__city__icontains=location_filter) |
            Q(userprofile__state__icontains=location_filter) |
            Q(userprofile__zip_code__icontains=location_filter)
        )
    
    # Availability filter
    availability_filter = request.GET.get('availability')
    if availability_filter == 'available':
        providers_queryset = providers_queryset.filter(userprofile__is_available=True)
    
    # Verification filter
    verification_filter = request.GET.get('verification')
    if verification_filter == 'verified':
        providers_queryset = providers_queryset.filter(userprofile__is_verified=True)
    
    # Experience filter
    experience_filter = request.GET.get('experience')
    if experience_filter:
        try:
            min_experience = int(experience_filter)
            providers_queryset = providers_queryset.filter(userprofile__experience_years__gte=min_experience)
        except (ValueError, TypeError):
            pass
    
    # Service radius filter
    service_radius = request.GET.get('service_radius')
    if service_radius:
        try:
            max_radius = int(service_radius)
            providers_queryset = providers_queryset.filter(userprofile__service_radius__lte=max_radius)
        except (ValueError, TypeError):
            pass
    
    # Sorting options
    sort_by = request.GET.get('sort', 'rating')
    if sort_by == 'rating':
        providers_queryset = providers_queryset.order_by('-userprofile__rating', '-userprofile__review_count')
    elif sort_by == 'price_low':
        providers_queryset = providers_queryset.annotate(
            min_price=Min('custom_services__estimated_price')
        ).order_by('min_price')
    elif sort_by == 'price_high':
        providers_queryset = providers_queryset.annotate(
            max_price=Max('custom_services__estimated_price')
        ).order_by('-max_price')
    elif sort_by == 'experience':
        providers_queryset = providers_queryset.order_by('-userprofile__experience_years')
    elif sort_by == 'reviews':
        providers_queryset = providers_queryset.order_by('-userprofile__review_count')
    elif sort_by == 'newest':
        providers_queryset = providers_queryset.order_by('-userprofile__created_at')
    elif sort_by == 'response_time':
        providers_queryset = providers_queryset.order_by('userprofile__response_time_hours')
    else:
        # Default sorting
        providers_queryset = providers_queryset.order_by('-userprofile__rating', '-userprofile__review_count')
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(providers_queryset, 12)  # Show 12 providers per page
    page_number = request.GET.get('page')
    providers = paginator.get_page(page_number)
    
    # Get all categories for filter dropdown
    categories = Service.CATEGORY_CHOICES
    
    # Get filter statistics for better UX
    total_professionals = User.objects.filter(
        userprofile__user_type='professional',
        custom_services__is_active=True
    ).distinct().count()
    
    verified_count = User.objects.filter(
        userprofile__user_type='professional',
        userprofile__is_verified=True,
        custom_services__is_active=True
    ).distinct().count()
    
    available_count = User.objects.filter(
        userprofile__user_type='professional',
        userprofile__is_available=True,
        custom_services__is_active=True
    ).distinct().count()
    
    # Price range statistics
    price_stats = CustomService.objects.filter(is_active=True).aggregate(
        min_price=Min('estimated_price'),
        max_price=Max('estimated_price'),
        avg_price=Avg('estimated_price')
    )
    
    context = {
        'providers': providers,
        'categories': categories,
        'category_filter': category_filter,
        'rating_filter': rating_filter,
        'search_filter': search_filter,
        'price_min': price_min,
        'price_max': price_max,
        'location_filter': location_filter,
        'availability_filter': availability_filter,
        'verification_filter': verification_filter,
        'experience_filter': experience_filter,
        'service_radius': service_radius,
        'sort_by': sort_by,
        'total_professionals': total_professionals,
        'verified_count': verified_count,
        'available_count': available_count,
        'price_stats': price_stats,
    }
    
    return render(request, 'services/all_professionals.html', context)


def provider_profile(request, provider_id):
    """Provider profile page"""
    try:
        provider = User.objects.get(id=provider_id)
        user_profile = provider.userprofile
        custom_services = CustomService.objects.filter(provider=provider, is_active=True)
        
        # Calculate provider statistics
        try:
            provider_orders = Order.objects.filter(professional=provider)
            total_orders = provider_orders.count()
            completed_orders = provider_orders.filter(status='completed').count()
            
            # Calculate average rating (simulated)
            average_rating = float(user_profile.rating) if user_profile.rating else 4.7
            if total_orders > 0:
                average_rating = float(user_profile.rating) if user_profile.rating else 4.7
            else:
                average_rating = 0
            
            # Calculate total services
            total_services = custom_services.count()
            
            # Simulate other statistics
            response_time = "2 horas"
            cancellation_rate = "2%"
            rehire_rate = "85%"
        except Exception:
            # Fallback values if there are errors
            total_orders = 0
            completed_orders = 0
            average_rating = float(user_profile.rating) if user_profile.rating else 0

            total_services = custom_services.count()
            response_time = "2 horas"
            cancellation_rate = "2%"
            rehire_rate = "85%"
            
        return render(request, 'services/provider_profile.html', {
            'provider': provider,
            'user_profile': user_profile,
            'custom_services': custom_services,
            'total_orders': total_orders,
            'completed_orders': completed_orders,
            'average_rating': average_rating,
            'total_services': total_services,
            'response_time': response_time,
            'cancellation_rate': cancellation_rate,
            'rehire_rate': rehire_rate
        })
    except User.DoesNotExist:
        messages.error(request, 'Prestador não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, 'Perfil do prestador não encontrado')
        return redirect('home')


@login_required
def profile_changes(request):
    """Display profile changes for the current user"""
    # Get profile changes for the current user
    changes = request.user.profilechange_set.all().order_by('-created_at')
    
    return render(request, 'services/profile_changes.html', {
        'changes': changes
    })


@login_required
def admin_profile_changes(request):
    """Display all profile changes (admin only)"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Get all profile changes
    changes = ProfileChange.objects.all().order_by('-created_at')
    
    return render(request, 'services/admin_profile_changes.html', {
        'changes': changes
    })




def login_view(request):
    """Login page"""
    # If user is already authenticated, redirect to home
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username_or_email = request.POST['username']
        password = request.POST['password']
        remember_me = request.POST.get('remember', False)
        
        # Try to authenticate with username first
        user = authenticate(request, username=username_or_email, password=password)
        
        # If authentication failed, try to find user by email and authenticate
        if user is None:
            try:
                # Check if input is an email
                if '@' in username_or_email:
                    user_obj = User.objects.get(email=username_or_email)
                    user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
        
        if user is not None:
            login(request, user)
            
            # Handle "remember me" functionality
            if remember_me:
                # Keep session for 2 weeks
                request.session.set_expiry(1209600)
            else:
                # Expire session when browser closes
                request.session.set_expiry(0)
            
            return redirect('home')
        else:
            messages.error(request, 'Nome de usuário/e-mail ou senha inválidos')
    
    return render(request, 'services/login.html')

def register_view(request):
    """Registration page - now accessible to everyone"""
    # Remove the admin-only restriction - allow public registration
    # Only redirect to login if there's a technical issue
    
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password_confirm = request.POST['password_confirm']
        user_type = request.POST.get('user_type', 'customer')
        
        if password != password_confirm:
            messages.error(request, 'As senhas não coincidem')
            return render(request, 'registration/clean_register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Nome de usuário já existe')
            return render(request, 'registration/clean_register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'E-mail já cadastrado')
            return render(request, 'registration/clean_register.html')
        
        # Create user
        user = User.objects.create_user(username=username, email=email, password=password)
        
        # Create user profile
        try:
            UserProfile.objects.create(user=user, user_type=user_type)
        except Exception as e:
            # Handle the case where the database columns don't exist yet
            if "no such column" in str(e):
                # Try to create profile without the new fields
                try:
                    profile = UserProfile(user=user, user_type=user_type)
                    profile.save(update_fields=['user', 'user_type', 'created_at', 'updated_at'])
                except:
                    pass
            else:
                raise e
        
        messages.success(request, f'Conta criada com sucesso para {username}')
        return redirect('login')
    
    return render(request, 'registration/clean_register.html')

def logout_view(request):
    """Logout view"""
    logout(request)
    messages.success(request, 'Você saiu da sua conta com sucesso.')
    return redirect('home')

@login_required
def profile_new(request):
    """Enhanced profile page with all settings merged"""
    try:
        user_profile = request.user.userprofile
    except Exception as e:
        # Create a default user profile if it doesn't exist
        user_profile = UserProfile()
        user_profile.user = request.user
        user_profile.save()
    
    # Get user's payment methods
    payment_methods = PaymentMethod.objects.filter(user=request.user)
    
    # Get professional-specific data if user is a professional
    portfolio_items = []
    availability = []
    professional_stats = {}
    custom_services = []
    solicitacoes_recentes = []
    stats_solicitacoes = {}
    provider_orders = []
    total_earnings = 0
    
    if user_profile.user_type == 'professional':
        from .models import PortfolioItem, ProfessionalAvailability, Order, Review, CustomService, ServiceRequest
        from django.db.models import Avg, Sum, Count, Q
        from django.utils import timezone
        from datetime import timedelta, datetime
        
        # Get portfolio items
        portfolio_items = PortfolioItem.objects.filter(professional=request.user).order_by('-is_featured', '-completion_date')[:6]
        
        # Get availability
        availability = ProfessionalAvailability.objects.filter(professional=request.user).order_by('weekday')
        
        # Get custom services
        custom_services = CustomService.objects.filter(professional=request.user).order_by('-created_at')
        
        # Get service requests (solicitações)
        solicitacoes_recentes = ServiceRequest.objects.filter(
            professional=request.user
        ).order_by('-created_at')[:5]
        
        # Calculate statistics for service requests
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        
        stats_solicitacoes = {
            'pending': ServiceRequest.objects.filter(professional=request.user, status='pending').count(),
            'contacted': ServiceRequest.objects.filter(professional=request.user, status='contacted').count(),
            'scheduled': ServiceRequest.objects.filter(professional=request.user, status='scheduled').count(),
            'completed': ServiceRequest.objects.filter(professional=request.user, status='completed').count(),
            'esta_semana': ServiceRequest.objects.filter(
                professional=request.user,
                status='completed',
                created_at__gte=week_start
            ).count(),
        }
        
        # Calculate professional statistics
        total_orders = Order.objects.filter(professional=request.user).count()
        completed_orders = Order.objects.filter(professional=request.user, status='completed').count()
        in_progress_orders = Order.objects.filter(professional=request.user, status='in_progress').count()
        cancelled_orders = Order.objects.filter(professional=request.user, status='cancelled').count()
        pending_orders = Order.objects.filter(professional=request.user, status='pending').count()
        
        # Get recent orders
        provider_orders = Order.objects.filter(professional=request.user).order_by('-created_at')[:10]
        
        # Get recent reviews
        recent_reviews = Review.objects.filter(professional=request.user).order_by('-created_at')[:5]
        
        # Calculate average rating
        avg_rating = Review.objects.filter(professional=request.user).aggregate(Avg('rating'))['rating__avg'] or 0
        
        # Calculate earnings
        total_earnings = Order.objects.filter(
            professional=request.user,
            status='completed'
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Calculate earnings by period
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        week_start_datetime = timezone.make_aware(datetime.combine(week_start, datetime.min.time()))
        month_start = today_start.replace(day=1)
        
        earnings_today = Order.objects.filter(
            professional=request.user,
            status='completed',
            completed_at__gte=today_start
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        earnings_week = Order.objects.filter(
            professional=request.user,
            status='completed',
            completed_at__gte=week_start_datetime
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        earnings_month = Order.objects.filter(
            professional=request.user,
            status='completed',
            completed_at__gte=month_start
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Count completed today and scheduled today
        completed_today = Order.objects.filter(
            professional=request.user,
            status='completed',
            completed_at__gte=today_start
        ).count()
        
        scheduled_today = Order.objects.filter(
            professional=request.user,
            status='scheduled',
            scheduled_date=today
        ).count()
        
        # Calculate weekly stats
        weekly_completed = Order.objects.filter(
            professional=request.user,
            status='completed',
            completed_at__gte=week_start_datetime
        ).count()
        
        weekly_rating = Review.objects.filter(
            professional=request.user,
            created_at__gte=week_start_datetime
        ).aggregate(Avg('rating'))['rating__avg'] or 0
        
        # Calculate conversion rate
        total_requests = ServiceRequest.objects.filter(professional=request.user).count()
        completed_requests = ServiceRequest.objects.filter(professional=request.user, status='completed').count()
        conversion_rate = round((completed_requests / total_requests * 100), 1) if total_requests > 0 else 0
        
        professional_stats = {
            'total_orders': total_orders,
            'completed_orders': completed_orders,
            'in_progress_orders': in_progress_orders,
            'cancelled_orders': cancelled_orders,
            'pending_orders': pending_orders,
            'recent_reviews': recent_reviews,
            'avg_rating': round(avg_rating, 1) if avg_rating else 0,
            'completion_rate': round((completed_orders / total_orders * 100), 1) if total_orders > 0 else 0,
            'total_earnings': total_earnings,
            'monthly_earnings': earnings_month,
            'weekly_earnings': earnings_week,
            'daily_earnings': earnings_today,
            'earnings_today': earnings_today,
            'completed_today': completed_today,
            'scheduled_today': scheduled_today,
            'pending_requests': stats_solicitacoes['pending'],
            'weekly_completed': weekly_completed,
            'weekly_rating': round(weekly_rating, 1) if weekly_rating else 0,
            'conversion_rate': conversion_rate,
            'completed_services': completed_orders,
            'avg_response_time': user_profile.response_time_hours or 2,
            'satisfaction_rate': round(avg_rating / 5 * 100, 0) if avg_rating else 0,
        }
        
        # Update profile stats
        user_profile.total_jobs = total_orders
        user_profile.completed_jobs = completed_orders
        user_profile.rating = avg_rating
        user_profile.review_count = Review.objects.filter(professional=request.user).count()
        user_profile.save()
    
    if request.method == 'POST':
        # Handle different form submissions based on hidden fields
        if 'change_password' in request.POST:
            # Change password
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')
            
            # Verify current password
            if not request.user.check_password(current_password):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Senha atual incorreta!'})
                else:
                    messages.error(request, 'Senha atual incorreta!')
                    return redirect('profile_new')
            
            # Check if new passwords match
            if new_password != confirm_password:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'As novas senhas não coincidem!'})
                else:
                    messages.error(request, 'As novas senhas não coincidem!')
                    return redirect('profile_new')
            
            # Check password strength
            if len(new_password) < 8:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'A nova senha deve ter pelo menos 8 caracteres!'})
                else:
                    messages.error(request, 'A nova senha deve ter pelo menos 8 caracteres!')
                    return redirect('profile_new')
            
            # Update password
            request.user.set_password(new_password)
            request.user.save()
            
            # Log the user in with the new password
            from django.contrib.auth import login
            login(request, request.user)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Senha alterada com sucesso!'})
            else:
                messages.success(request, 'Senha alterada com sucesso!')
                return redirect('profile_new')
        elif 'update_notifications' in request.POST:
            # Update notification preferences
            user_profile.email_notifications = request.POST.get('email_notifications') == 'on'
            user_profile.promotional_emails = request.POST.get('promotional_emails') == 'on'
            user_profile.security_alerts = request.POST.get('security_alerts') == 'on'
            user_profile.push_notifications = request.POST.get('push_notifications') == 'on'
            user_profile.push_messages = request.POST.get('push_messages') == 'on'
            
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Preferências de notificação atualizadas com sucesso!'})
            else:
                messages.success(request, 'Preferências de notificação atualizadas com sucesso!')
                return redirect('profile_new')
        elif 'update_preferences' in request.POST:
            # Update general preferences
            user_profile.dark_mode = request.POST.get('dark_mode') == 'on'
            
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Preferências atualizadas com sucesso!'})
            else:
                messages.success(request, 'Preferências atualizadas com sucesso!')
                return redirect('profile_new')
        elif 'avatar' in request.FILES:
            # Handle avatar upload
            try:
                # Get the uploaded file
                avatar = request.FILES['avatar']
                
                # Update the user's profile with the new avatar
                user_profile.avatar = avatar
                user_profile.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True, 
                        'message': 'Foto de perfil atualizada com sucesso!',
                        'avatar_url': user_profile.avatar.url if user_profile.avatar else None
                    })
                else:
                    messages.success(request, 'Foto de perfil atualizada com sucesso!')
                    return redirect('profile_new')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False, 
                        'message': f'Erro ao atualizar a foto de perfil: {str(e)}'
                    })
                else:
                    messages.error(request, f'Erro ao atualizar a foto de perfil: {str(e)}')
                    return redirect('profile_new')
        elif 'update_profile' in request.POST:
            # Update basic profile information
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name = request.POST.get('last_name', '')
            request.user.save()
            
            user_profile.phone = request.POST.get('phone', '')
            birth_date = request.POST.get('birth_date', '')
            if birth_date:
                user_profile.birth_date = birth_date
                
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Informações pessoais atualizadas com sucesso!'})
            else:
                messages.success(request, 'Informações pessoais atualizadas com sucesso!')
                return redirect('profile_new')
                
        elif 'update_address' in request.POST:
            # Update address information
            user_profile.zip_code = request.POST.get('zip_code', '')
            user_profile.address = request.POST.get('address', '')
            user_profile.number = request.POST.get('number', '')
            user_profile.complement = request.POST.get('complement', '')
            user_profile.city = request.POST.get('city', '')
            user_profile.state = request.POST.get('state', '')
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Endereço atualizado com sucesso!'})
            else:
                messages.success(request, 'Endereço atualizado com sucesso!')
                return redirect('profile_new')
                
        elif 'update_professional_info' in request.POST:
            # Update professional-specific information
            user_profile.bio = request.POST.get('bio', '')
            user_profile.experience_years = int(request.POST.get('experience_years', 0) or 0)
            user_profile.specialties = request.POST.get('specialties', '')
            user_profile.certifications = request.POST.get('certifications', '')
            user_profile.portfolio_url = request.POST.get('portfolio_url', '')
            user_profile.linkedin_url = request.POST.get('linkedin_url', '')
            user_profile.website_url = request.POST.get('website_url', '')
            user_profile.business_name = request.POST.get('business_name', '')
            user_profile.cnpj = request.POST.get('cnpj', '')
            user_profile.business_hours = request.POST.get('business_hours', '')
            user_profile.service_radius = int(request.POST.get('service_radius', 10) or 10)
            user_profile.service_type = request.POST.get('service_type', '')
            user_profile.hourly_rate = request.POST.get('hourly_rate', '') or None
            user_profile.response_time_hours = int(request.POST.get('response_time_hours', 24) or 24)
            user_profile.is_available = request.POST.get('is_available') == 'on'
            user_profile.is_verified = request.POST.get('is_verified') == 'on'
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Informações profissionais atualizadas com sucesso!'})
            else:
                messages.success(request, 'Informações profissionais atualizadas com sucesso!')
                return redirect('profile_new')
                
        elif 'update_pricing' in request.POST:
            # Update service pricing
            # This would typically be stored in a separate ServicePricing model
            # For now, we'll store it in the user profile as JSON
            import json
            pricing_data = {}
            for key, value in request.POST.items():
                if key.startswith('price_') and value:
                    service_type = key.replace('price_', '')
                    try:
                        pricing_data[service_type] = float(value)
                    except ValueError:
                        pass
            
            user_profile.service_pricing = json.dumps(pricing_data) if pricing_data else ''
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Preços dos serviços atualizados com sucesso!'})
            else:
                messages.success(request, 'Preços dos serviços atualizados com sucesso!')
                return redirect('profile_new')
                
        elif 'update_service_areas' in request.POST:
            # Update service areas and travel policy
            user_profile.service_areas = request.POST.get('service_areas', '')
            user_profile.travel_policy = request.POST.get('travel_policy', 'free')
            travel_cost = request.POST.get('travel_cost_per_km', '')
            user_profile.travel_cost_per_km = float(travel_cost) if travel_cost else 0.0
            user_profile.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Áreas de atendimento atualizadas com sucesso!'})
            else:
                messages.success(request, 'Áreas de atendimento atualizadas com sucesso!')
                return redirect('profile_new')
                
        elif 'add_portfolio_item' in request.POST:
            # Add new portfolio item
            from .models import PortfolioItem
            
            portfolio_item = PortfolioItem(
                professional=request.user,
                title=request.POST.get('portfolio_title', ''),
                description=request.POST.get('portfolio_description', ''),
                category=request.POST.get('portfolio_category', ''),
                completion_date=request.POST.get('portfolio_date', ''),
                client_name=request.POST.get('portfolio_client', ''),
                is_featured=request.POST.get('portfolio_featured') == 'on'
            )
            
            if 'portfolio_image' in request.FILES:
                portfolio_item.image = request.FILES['portfolio_image']
                
            portfolio_item.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Item adicionado ao portfólio com sucesso!'})
            else:
                messages.success(request, 'Item adicionado ao portfólio com sucesso!')
                return redirect('profile_new')
        else:
            # Fallback for other updates
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Perfil atualizado com sucesso!'})
            else:
                messages.success(request, 'Perfil atualizado com sucesso!')
                return redirect('profile_new')
    
    # Calculate customer statistics if not a professional
    total_orders = 0
    if user_profile.user_type != 'professional':
        from .models import Order
        total_orders = Order.objects.filter(customer=request.user).count()
    
    # For GET requests, just render the page
    context = {
        'user_profile': user_profile,
        'payment_methods': payment_methods,
        'portfolio_items': portfolio_items,
        'availability': availability,
        'professional_stats': professional_stats,
        'service_categories': Service.CATEGORY_CHOICES,
        'custom_services': custom_services,
        'solicitacoes_recentes': solicitacoes_recentes,
        'stats_solicitacoes': stats_solicitacoes,
        'provider_orders': provider_orders,
        'total_orders': total_orders,
        'total_earnings': total_earnings,
        'average_rating': professional_stats.get('avg_rating', 0) if professional_stats else 0,
        'completed_orders': professional_stats.get('completed_orders', 0) if professional_stats else 0,
        'pending_orders': professional_stats.get('pending_orders', 0) if professional_stats else 0,
    }
    
    return render(request, 'services/profile_new.html', context)


@login_required
def upload_avatar(request):
    """Handle avatar upload via AJAX"""
    if request.method == 'POST' and request.FILES.get('avatar'):
        try:
            user_profile = request.user.userprofile
            avatar = request.FILES['avatar']
            
            # Update the user's profile with the new avatar
            user_profile.avatar = avatar
            user_profile.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Foto de perfil atualizada com sucesso!',
                'avatar_url': user_profile.avatar.url if user_profile.avatar else None
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Erro ao atualizar a foto de perfil: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Nenhuma imagem foi enviada.'
    })


@login_required
def add_payment_method(request):
    """Add a payment method"""
    if request.method == 'POST':
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                payment_type = request.POST.get('payment_type')
                card_number = request.POST.get('card_number', '')  # Full card number
                cardholder_name = request.POST.get('cardholder_name', '')
                expiry_date = request.POST.get('expiry_date', '')
                
                # Validate required fields based on payment type
                if payment_type in ['credit_card', 'debit_card']:
                    if not card_number:
                        return JsonResponse({
                            'success': False, 
                            'message': 'Número do cartão é obrigatório'
                        })
                    if not cardholder_name:
                        return JsonResponse({
                            'success': False, 
                            'message': 'Nome do titular é obrigatório'
                        })
                
                # Extract last 4 digits of card number
                card_number_last4 = ''
                if card_number:
                    # Remove spaces and extract last 4 digits
                    card_number_clean = card_number.replace(' ', '')
                    if len(card_number_clean) >= 4:
                        card_number_last4 = card_number_clean[-4:]
                
                # Create payment method
                payment_method = PaymentMethod.objects.create(
                    user=request.user,
                    payment_type=payment_type,
                    card_number_last4=card_number_last4,
                    cardholder_name=cardholder_name
                )
                
                # Parse and save expiry date if provided
                if expiry_date:
                    try:
                        # Parse expiry date (MM/YY format)
                        month, year = map(int, expiry_date.split('/'))
                        # Assume year 20XX if year is less than 50, otherwise 19XX
                        if year < 50:
                            year += 2000
                        else:
                            year += 1900
                        # Set to first day of the month
                        from datetime import date
                        payment_method.expiry_date = date(year, month, 1)
                        payment_method.save()
                    except Exception:
                        pass  # Invalid date format, continue without saving date
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Método de pagamento adicionado com sucesso',
                    'payment_method_id': payment_method.id
                })
            except Exception as e:
                return JsonResponse({
                    'success': False, 
                    'message': 'Erro ao adicionar método de pagamento: ' + str(e)
                })
        else:
            # Regular form submission
            payment_type = request.POST.get('payment_type')
            card_number = request.POST.get('card_number', '')  # Full card number
            cardholder_name = request.POST.get('cardholder_name', '')
            expiry_date = request.POST.get('expiry_date', '')
            
            # Validate required fields
            if not payment_type:
                messages.error(request, 'Tipo de pagamento é obrigatório')
                return redirect('profile_new')
            
            # Extract last 4 digits of card number
            card_number_last4 = ''
            if card_number:
                # Remove spaces and extract last 4 digits
                card_number_clean = card_number.replace(' ', '')
                if len(card_number_clean) >= 4:
                    card_number_last4 = card_number_clean[-4:]
            
            if payment_type in ['credit_card', 'debit_card']:
                if not card_number_last4:
                    messages.error(request, 'Número do cartão é obrigatório')
                    return redirect('profile_new')
                if not cardholder_name:
                    messages.error(request, 'Nome do titular é obrigatório')
                    return redirect('profile_new')
            
            try:
                # Create payment method
                payment_method = PaymentMethod.objects.create(
                    user=request.user,
                    payment_type=payment_type,
                    card_number_last4=card_number_last4,
                    cardholder_name=cardholder_name
                )
                
                # Parse and save expiry date if provided
                if expiry_date:
                    try:
                        # Parse expiry date (MM/YY format)
                        month, year = map(int, expiry_date.split('/'))
                        # Assume year 20XX if year is less than 50, otherwise 19XX
                        if year < 50:
                            year += 2000
                        else:
                            year += 1900
                        # Set to first day of the month
                        from datetime import date
                        payment_method.expiry_date = date(year, month, 1)
                        payment_method.save()
                    except Exception:
                        pass  # Invalid date format, continue without saving date
                
                messages.success(request, 'Método de pagamento adicionado com sucesso')
                return redirect('profile_new')
            except Exception as e:
                messages.error(request, 'Erro ao adicionar método de pagamento: ' + str(e))
                return redirect('profile_new')
    
    return render(request, 'services/add_payment_method.html')

@login_required
def save_zip_code(request):
    """Save ZIP code to user profile"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        zip_code = request.POST.get('zip_code', '').replace('-', '')
        
        if len(zip_code) != 8:
            return JsonResponse({'success': False, 'message': 'CEP inválido'})
        
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            user_profile.zip_code = zip_code
            user_profile.save()
            
            return JsonResponse({'success': True, 'message': 'CEP salvo com sucesso no seu perfil'})
        except UserProfile.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Perfil não encontrado'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': 'Erro ao salvar CEP'})
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'})

@login_required
def sponsors_view(request):
    """Sponsors page"""
    sponsors = Sponsor.objects.filter(is_active=True)
    return render(request, 'services/sponsors.html', {'sponsors': sponsors})

@login_required
def provider_dashboard(request):
    """Provider dashboard"""
    # Check if user is a provider
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    except Exception as e:
        # Handle the case where the database columns don't exist yet
        if "no such column" in str(e):
            messages.error(request, 'Acesso negado')
            return redirect('home')
        else:
            raise e
    
    # Get provider's services (for now, we'll show all services)
    services = Service.objects.filter(is_active=True)
    
    # Get provider's custom services with error handling for decimal issues
    custom_services = []
    try:
        custom_services = CustomService.objects.filter(provider=request.user, is_active=True)
        # Force evaluation to catch decimal errors
        list(custom_services)
    except Exception as e:
        # Handle decimal conversion errors
        custom_services = []
        messages.error(request, 'Erro ao carregar serviços personalizados. Alguns dados podem estar corrompidos.')
    
    # Get provider's orders
    try:
        provider_orders = Order.objects.filter(professional=request.user).order_by('-created_at')[:10]
    except Exception as e:
        # Handle decimal conversion errors
        provider_orders = []
        messages.error(request, 'Erro ao carregar pedidos. Alguns dados podem estar corrompidos.')
    
    # Calculate statistics
    total_orders = 0
    completed_orders = 0
    pending_orders = 0
    total_earnings = 0
    
    try:
        all_provider_orders = Order.objects.filter(professional=request.user)
        total_orders = all_provider_orders.count()
        completed_orders = all_provider_orders.filter(status='completed').count()
        pending_orders = all_provider_orders.exclude(status='completed').count()
        
        # Calculate total earnings (handling decimal errors)
        for order in all_provider_orders:
            try:
                total_earnings += float(order.total_price)
            except:
                # Skip orders with decimal conversion issues
                pass
    except Exception as e:
        # Handle decimal conversion errors
        messages.error(request, 'Erro ao calcular estatísticas. Alguns dados podem estar corrompidos.')
    
    # Calculate rating statistics
    total_ratings = 0
    average_rating = 0
    try:
        # In a real application, you would have a ratings model
        # For now, we'll simulate some rating data
        total_ratings = min(total_orders, 50)  # Simulate up to 50 ratings
        average_rating = 4.7 if total_ratings > 0 else 0  # Simulate 4.7 average rating
    except Exception as e:
        pass
    
    # Adicionar dados das solicitações de serviço
    stats_solicitacoes = {}
    solicitacoes_recentes = []
    
    try:
        from .models import ServiceRequestModal
        from django.utils import timezone
        from datetime import timedelta
        
        # Buscar serviços do prestador
        servicos_prestador = CustomService.objects.filter(provider=request.user)
        
        if servicos_prestador.exists():
            # Estatísticas das solicitações
            hoje = timezone.now().date()
            semana_passada = hoje - timedelta(days=7)
            
            stats_solicitacoes = {
                'total': ServiceRequestModal.objects.filter(provider=request.user).count(),
                'pending': ServiceRequestModal.objects.filter(provider=request.user, status='pending').count(),
                'contacted': ServiceRequestModal.objects.filter(provider=request.user, status='contacted').count(),
                'scheduled': ServiceRequestModal.objects.filter(provider=request.user, status='scheduled').count(),
                'completed': ServiceRequestModal.objects.filter(provider=request.user, status='completed').count(),
                'esta_semana': ServiceRequestModal.objects.filter(
                    provider=request.user, 
                    created_at__date__gte=semana_passada
                ).count(),
            }
            
            # Solicitações recentes (últimas 3)
            solicitacoes_recentes = ServiceRequestModal.objects.filter(
                provider=request.user
            ).select_related('user', 'service').order_by('-created_at')[:3]
            
    except Exception as e:
        print(f"Erro ao buscar dados das solicitações: {e}")
        stats_solicitacoes = {
            'total': 0, 'pending': 0, 'contacted': 0, 
            'scheduled': 0, 'completed': 0, 'esta_semana': 0
        }
    
    return render(request, 'services/provider_dashboard.html', {
        'services': services,
        'custom_services': custom_services,
        'provider_orders': provider_orders,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'total_earnings': total_earnings,
        'total_ratings': total_ratings,
        'average_rating': average_rating,
        'stats_solicitacoes': stats_solicitacoes,
        'solicitacoes_recentes': solicitacoes_recentes,
    })

@login_required
def provider_ai_insights(request):
    """AI-powered insights for providers"""
    # Check if user is a provider
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            return JsonResponse({'error': 'Acesso negado'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Acesso negado'}, status=403)
    except Exception as e:
        return JsonResponse({'error': 'Erro desconhecido'}, status=500)
    
    # Get provider's data
    try:
        custom_services = CustomService.objects.filter(provider=request.user, is_active=True)
        all_provider_orders = Order.objects.filter(professional=request.user)
        total_orders = all_provider_orders.count()
        completed_orders = all_provider_orders.filter(status='completed').count()
        average_rating = 4.7 if total_orders > 0 else 0  # Simulated rating
    except Exception as e:
        return JsonResponse({'error': 'Erro ao carregar dados'}, status=500)
    
    # AI-Powered Features Data
    # 1. Performance Insights Data
    performance_data = {
        'revenue_growth': 18,  # Simulated data
        'on_time_completion': 92,
        'average_rating': average_rating,
        'completion_rate': (completed_orders / total_orders * 100) if total_orders > 0 else 0
    }
    
    # 2. Revenue Prediction Data
    historical_revenue = [1200, 1900, 1500, 2200, 1800, 2500]  # Simulated data
    
    # 3. Provider Profile Data for AI Analysis
    provider_profile = {
        'services': list(custom_services),
        'competitiveness': 0.65,  # Simulated competitiveness score
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'average_rating': average_rating
    }
    
    # Generate AI insights
    insights = generate_performance_insights(performance_data)
    revenue_prediction = predict_revenue(historical_revenue, provider_profile)
    recommendations = generate_personalized_recommendations(provider_profile, {})
    
    return JsonResponse({
        'insights': insights,
        'revenue_prediction': revenue_prediction,
        'recommendations': recommendations
    })

def generate_performance_insights(user_data):
    """Generate performance insights for the provider"""
    insights = []
    
    # Revenue trend analysis
    if user_data['revenue_growth'] > 15:
        insights.append({
            'type': 'success',
            'title': 'Excelente Crescimento',
            'message': f"Sua receita aumentou {user_data['revenue_growth']}% este mês. Continue oferecendo serviços de qualidade!",
            'action': 'Ver detalhes'
        })
    
    # Efficiency analysis
    if user_data['on_time_completion'] > 90:
        insights.append({
            'type': 'info',
            'title': 'Alta Eficiência',
            'message': f"Você concluiu {user_data['on_time_completion']}% dos pedidos no prazo. Isso demonstra profissionalismo.",
            'action': 'Manter desempenho'
        })
    
    # Rating analysis
    if user_data['average_rating'] > 4.5:
        insights.append({
            'type': 'warning',
            'title': 'Excelente Reputação',
            'message': f"Sua avaliação média é {user_data['average_rating']}/5.0. Continue mantendo esse padrão de qualidade!",
            'action': 'Ver avaliações'
        })
    
    # Completion rate analysis
    if user_data['completion_rate'] < 80:
        insights.append({
            'type': 'danger',
            'title': 'Oportunidade de Melhoria',
            'message': f"Sua taxa de conclusão está em {user_data['completion_rate']:.1f}%. Considere revisar seus processos para melhorar.",
            'action': 'Ver pedidos'
        })
    
    return insights

def predict_revenue(historical_data, user_profile):
    """Predict future revenue based on historical data"""
    # Simple moving average prediction
    def calculate_average_growth(data):
        if len(data) < 2:
            return 0
        total_growth = 0
        for i in range(1, len(data)):
            total_growth += (data[i] - data[i-1]) / data[i-1]
        return total_growth / (len(data) - 1)
    
    def calculate_confidence_score(data):
        # Simple confidence calculation based on data points
        if len(data) < 3:
            return 0.3  # Low confidence
        if len(data) < 6:
            return 0.6  # Medium confidence
        return 0.9  # High confidence
    
    def generate_recommendations(avg_growth, profile):
        recommendations = []
        if avg_growth > 0.1:
            recommendations.append({
                'type': 'success',
                'title': 'Crescimento Sustentável',
                'description': 'Seu crescimento mensal é consistente. Considere expandir sua área de atendimento.',
                'action': 'Expandir serviços'
            })
        elif avg_growth < 0:
            recommendations.append({
                'type': 'danger',
                'title': 'Oportunidade de Recuperação',
                'description': 'Seu crescimento está negativo. Foque em melhorar a satisfação do cliente.',
                'action': 'Ver feedbacks'
            })
        if len(profile.get('services', [])) < 3:
            recommendations.append({
                'type': 'info',
                'title': 'Diversificação de Serviços',
                'description': 'Você oferece poucos serviços. Considere adicionar novas categorias para aumentar sua receita.',
                'action': 'Adicionar serviço'
            })
        return recommendations
    
    avg_monthly_growth = calculate_average_growth(historical_data)
    current_revenue = historical_data[-1] if historical_data else 0
    
    predictions = {
        'next_month': current_revenue * (1 + avg_monthly_growth),
        'three_months': current_revenue * pow(1 + avg_monthly_growth, 3),
        'six_months': current_revenue * pow(1 + avg_monthly_growth, 6)
    }
    
    return {
        'predictions': predictions,
        'confidence': calculate_confidence_score(historical_data),
        'recommendations': generate_recommendations(avg_monthly_growth, user_profile)
    }

def generate_personalized_recommendations(provider_profile, market_data):
    """Generate personalized recommendations for the provider"""
    recommendations = []
    
    # Recommend new service categories based on market demand
    def analyze_market_trends(data):
        # Simulated market trend analysis
        return [
            {'id': 'cleaning', 'name': 'Limpeza', 'growth': 15},
            {'id': 'electrical', 'name': 'Elétrica', 'growth': 12},
            {'id': 'plumbing', 'name': 'Encanamento', 'growth': 10},
            {'id': 'painting', 'name': 'Pintura', 'growth': 8}
        ]
    
    trending_categories = analyze_market_trends(market_data)
    provider_categories = [s.category for s in provider_profile.get('services', [])]
    
    for category in trending_categories:
        if category['id'] not in provider_categories:
            recommendations.append({
                'type': 'service_opportunity',
                'title': f"Oportunidade: {category['name']}",
                'description': f"Alta demanda detectada para serviços de {category['name']} (+{category['growth']}% em 30 dias)",
                'action': 'Adicionar serviço'
            })
    
    # Recommend pricing adjustments
    if provider_profile.get('competitiveness', 1.0) < 0.7:
        recommendations.append({
            'type': 'pricing',
            'title': 'Oportunidade de Repricing',
            'description': 'Seus preços estão acima da média do mercado. Considere ajustar para aumentar sua competitividade.',
            'action': 'Analisar preços'
        })
    
    return recommendations

@login_required
def add_custom_service(request):
    """Add a custom service"""
    # Check if user is admin or provider
    is_admin = False
    provider_id = None
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type == 'admin':
            is_admin = True
            # If admin, check if a provider ID was specified
            provider_id = request.GET.get('provider')
        elif user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    except Exception as e:
        # Handle the case where the database columns don't exist yet
        if "no such column" in str(e):
            messages.error(request, 'Acesso negado')
            return redirect('home')
        else:
            raise e
    
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        category = request.POST['category']
        estimated_price = request.POST['estimated_price']
        estimated_duration_hours = request.POST['estimated_duration']
        is_active = request.POST.get('is_active', 'true') == 'true'
        
        # Determine the provider (current user or specified provider for admin)
        provider = request.user
        if is_admin and provider_id:
            try:
                provider = User.objects.get(id=provider_id)
            except User.DoesNotExist:
                messages.error(request, 'Prestador não encontrado')
                return redirect('admin_providers_list')
        
        # Create custom service
        CustomService.objects.create(
            name=name,
            description=description,
            category=category,
            estimated_price=estimated_price,
            estimated_duration=timedelta(hours=int(estimated_duration_hours)),
            provider=provider,
            is_active=is_active
        )
        
        messages.success(request, 'Serviço personalizado adicionado com sucesso')
        if is_admin and provider_id:
            return redirect('admin_provider_services', provider_id=provider_id)
        else:
            return redirect('provider_dashboard')
    
    return render(request, 'services/add_custom_service.html')


@login_required
def edit_custom_service(request, service_id):
    """Edit a custom service"""
    # Check if user is admin or provider
    is_admin = False
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type == 'admin':
            is_admin = True
        elif user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Get the custom service - if admin, no need to check ownership
    try:
        if is_admin:
            custom_service = CustomService.objects.get(id=service_id)
        else:
            custom_service = CustomService.objects.get(id=service_id, provider=request.user)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado')
        if is_admin:
            return redirect('admin_providers_list')
        else:
            return redirect('provider_dashboard')
    
    if request.method == 'POST':
        custom_service.name = request.POST['name']
        custom_service.description = request.POST['description']
        custom_service.category = request.POST['category']
        custom_service.estimated_price = request.POST['estimated_price']
        custom_service.estimated_duration = timedelta(hours=int(request.POST['estimated_duration']))
        custom_service.is_active = 'is_active' in request.POST
        custom_service.save()
        
        messages.success(request, 'Serviço personalizado atualizado com sucesso')
        if is_admin:
            return redirect('admin_provider_services', provider_id=custom_service.provider.id)
        else:
            return redirect('provider_dashboard')
    
    return render(request, 'services/edit_custom_service.html', {
        'custom_service': custom_service,
        'is_admin': is_admin
    })


@login_required
def remove_custom_service(request, service_id):
    """Remove a custom service"""
    # Check if user is admin or provider
    is_admin = False
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type == 'admin':
            is_admin = True
        elif user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Get the custom service - if admin, no need to check ownership
    try:
        if is_admin:
            custom_service = CustomService.objects.get(id=service_id)
        else:
            custom_service = CustomService.objects.get(id=service_id, provider=request.user)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado')
        if is_admin:
            return redirect('admin_providers_list')
        else:
            return redirect('provider_dashboard')
    
    if request.method == 'POST':
        custom_service.delete()
        messages.success(request, 'Serviço personalizado removido com sucesso')
        if is_admin:
            return redirect('admin_provider_services', provider_id=custom_service.provider.id)
        else:
            return redirect('provider_dashboard')
    
    return render(request, 'services/remove_custom_service.html', {
        'custom_service': custom_service,
        'is_admin': is_admin
    })


@login_required
def admin_dashboard(request):
    """Admin dashboard"""
    # Check if user is admin
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    except Exception as e:
        # Handle the case where the database columns don't exist yet
        if "no such column" in str(e):
            messages.error(request, 'Acesso negado')
            return redirect('home')
        else:
            raise e
    
    # Get dashboard data
    total_users = User.objects.count()
    total_services = Service.objects.count()
    total_orders = Order.objects.count()
    total_sponsors = Sponsor.objects.filter(is_active=True).count()
    
    # Get recent orders
    recent_orders = Order.objects.select_related('customer', 'service').order_by('-created_at')[:10]
    
    
    return render(request, 'services/admin_dashboard.html', {
        'total_users': total_users,
        'total_services': total_services,
        'total_orders': total_orders,
        'total_sponsors': total_sponsors,
        'recent_orders': recent_orders
    })


@login_required
def requested_services(request):
    """Show requested services"""
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'customer':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    requested_services = Order.objects.select_related('service').filter(customer=request.user)
    return render(request, 'services/requested_services.html', {'requested_services': requested_services})





@login_required
def order_confirmation(request, order_id):
    """Order confirmation"""
    order = Order.objects.get(id=order_id)
    
    return render(request, 'services/order_confirmation.html', {
        'order': order
    })


@login_required
def edit_payment_method(request, payment_method_id):
    """Edit a payment method"""
    try:
        payment_method = PaymentMethod.objects.get(id=payment_method_id, user=request.user)
    except PaymentMethod.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Método de pagamento não encontrado'})
        else:
            messages.error(request, 'Método de pagamento não encontrado')
            return redirect('profile_new')
    
    if request.method == 'POST':
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                payment_type = request.POST.get('payment_type')
                card_number = request.POST.get('card_number', '')  # Full card number
                cardholder_name = request.POST.get('cardholder_name', '')
                expiry_date = request.POST.get('expiry_date', '')
                
                # Update payment method fields
                payment_method.payment_type = payment_type
                payment_method.cardholder_name = cardholder_name
                
                # Extract last 4 digits of card number if provided
                if card_number:
                    card_number_clean = card_number.replace(' ', '')
                    if len(card_number_clean) >= 4:
                        payment_method.card_number_last4 = card_number_clean[-4:]
                
                # Parse and save expiry date if provided
                if expiry_date:
                    try:
                        month, year = map(int, expiry_date.split('/'))
                        if year < 50:
                            year += 2000
                        else:
                            year += 1900
                        from datetime import date
                        payment_method.expiry_date = date(year, month, 1)
                    except Exception:
                        payment_method.expiry_date = None
                else:
                    payment_method.expiry_date = None
                
                payment_method.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Método de pagamento atualizado com sucesso'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': 'Erro ao atualizar método de pagamento: ' + str(e)
                })
        else:
            # Handle regular form submission
            try:
                payment_type = request.POST.get('payment_type')
                card_number = request.POST.get('card_number', '')  # Full card number
                cardholder_name = request.POST.get('cardholder_name', '')
                expiry_date = request.POST.get('expiry_date', '')
                
                # Update payment method fields
                payment_method.payment_type = payment_type
                payment_method.cardholder_name = cardholder_name
                
                # Extract last 4 digits of card number if provided
                if card_number:
                    card_number_clean = card_number.replace(' ', '')
                    if len(card_number_clean) >= 4:
                        payment_method.card_number_last4 = card_number_clean[-4:]
                
                # Parse and save expiry date if provided
                if expiry_date:
                    try:
                        month, year = map(int, expiry_date.split('/'))
                        if year < 50:
                            year += 2000
                        else:
                            year += 1900
                        from datetime import date
                        payment_method.expiry_date = date(year, month, 1)
                    except Exception:
                        payment_method.expiry_date = None
                else:
                    payment_method.expiry_date = None
                
                payment_method.save()
                
                messages.success(request, 'Método de pagamento atualizado com sucesso')
                return redirect('profile_new')
            except Exception as e:
                messages.error(request, 'Erro ao atualizar método de pagamento: ' + str(e))
                return redirect('profile_new')
    
    # Handle GET request - return payment method data for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            'id': payment_method.id,
            'payment_type': payment_method.payment_type,
            'card_number_last4': payment_method.card_number_last4,
            'cardholder_name': payment_method.cardholder_name,
            'expiry_date': payment_method.expiry_date.strftime('%m/%y') if payment_method.expiry_date else ''
        }
        return JsonResponse({'success': True, 'data': data})
    
    # For regular requests, redirect to profile
    return redirect('profile_new')


@login_required
def delete_payment_method(request, payment_method_id):
    """Delete a payment method"""
    try:
        payment_method = PaymentMethod.objects.get(id=payment_method_id, user=request.user)
    except PaymentMethod.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Método de pagamento não encontrado'})
        else:
            messages.error(request, 'Método de pagamento não encontrado')
            return redirect('profile_new')
    
    if request.method == 'POST':
        try:
            payment_method.delete()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Método de pagamento removido com sucesso'
                })
            else:
                messages.success(request, 'Método de pagamento removido com sucesso')
                return redirect('profile_new')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Erro ao remover método de pagamento: ' + str(e)
                })
            else:
                messages.error(request, 'Erro ao remover método de pagamento: ' + str(e))
                return redirect('profile_new')
    
    # For non-POST requests, redirect to profile
    return redirect('profile_new')


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse

from .models import CustomService, Order, PaymentMethod, UserProfile





@login_required
def request_service_from_search(request, custom_service_id):
    """Request a custom service from search results with multi-step form"""
    # Check if user is a customer
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'customer':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Acesso negado. Apenas clientes podem solicitar serviços.'})
            messages.error(request, 'Acesso negado. Apenas clientes podem solicitar serviços.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Acesso negado. Perfil de usuário não encontrado.'})
        messages.error(request, 'Acesso negado. Perfil de usuário não encontrado.')
        return redirect('home')

    try:
        custom_service = CustomService.objects.select_related('provider').get(id=custom_service_id)
    except CustomService.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Serviço não encontrado. O serviço pode ter sido removido.'})
        messages.error(request, 'Serviço não encontrado. O serviço pode ter sido removido.')
        return redirect('search_new')

    if request.method == 'POST':
        # Get form data
        scheduled_datetime = request.POST.get('scheduled_datetime')
        time_preference = request.POST.get('time_preference')
        urgency = request.POST.get('urgency')
        address = request.POST.get('address')
        number = request.POST.get('number')
        complement = request.POST.get('complement', '')
        city = request.POST.get('city', '')
        state = request.POST.get('state', '')
        cep = request.POST.get('cep', '')
        contact_name = request.POST.get('contact_name', '')
        contact_phone = request.POST.get('contact_phone', '')
        contact_email = request.POST.get('contact_email', '')
        notes = request.POST.get('notes', '')
        accessibility_needs = request.POST.get('accessibility_needs')
        insurance_required = request.POST.get('insurance_required')
        terms_accepted = request.POST.get('terms_accepted')

        # Validate required fields
        missing_fields = []
        if not scheduled_datetime:
            missing_fields.append('data e hora preferidas')
        if not address:
            missing_fields.append('endereço')
        if not number:
            missing_fields.append('número')
        if not city:
            missing_fields.append('cidade')
        if not state:
            missing_fields.append('estado')
        if not contact_name:
            missing_fields.append('nome completo')
        if not contact_phone:
            missing_fields.append('telefone')
        if not contact_email:
            missing_fields.append('e-mail')
        if not terms_accepted:
            missing_fields.append('aceite dos termos e condições')

        if missing_fields:
            error_message = f'Por favor, preencha os seguintes campos obrigatórios: {", ".join(missing_fields)}.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': error_message,
                    'errors': {
                        'missing_fields': missing_fields
                    }
                })
            messages.error(request, error_message)
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Validate email format
        import re
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_regex, contact_email):
            error_message = 'Por favor, informe um e-mail válido.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': error_message,
                    'errors': {
                        'contact_email': 'Formato de e-mail inválido'
                    }
                })
            messages.error(request, error_message)
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Validate phone format (Brazilian format)
        phone_regex = r'^\(?[0-9]{2}\)? [0-9]{4,5}-[0-9]{4}$'
        if not re.match(phone_regex, contact_phone):
            error_message = 'Por favor, informe um telefone válido no formato (61) 98196-1144.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': error_message,
                    'errors': {
                        'contact_phone': 'Formato de telefone inválido'
                    }
                })
            messages.error(request, error_message)
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Combine date and time
        from datetime import datetime
        try:
            scheduled_datetime_obj = datetime.strptime(scheduled_datetime, "%Y-%m-%dT%H:%M")
        except ValueError:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': 'Data ou hora inválida. Por favor, selecione uma data e hora válidas.',
                    'errors': {
                        'scheduled_datetime': 'Formato de data/hora inválido'
                    }
                })
            messages.error(request, 'Data ou hora inválida. Por favor, selecione uma data e hora válidas.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Check if scheduled date is in the future
        from datetime import datetime
        if scheduled_datetime_obj < datetime.now():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': 'A data e hora devem ser futuras. Por favor, selecione uma data futura.',
                    'errors': {
                        'scheduled_datetime': 'A data deve ser futura'
                    }
                })
            messages.error(request, 'A data e hora devem ser futuras. Por favor, selecione uma data futura.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Combine address components
        address_components = [address, number]
        if complement:
            address_components.append(complement)
        full_address = ', '.join(address_components)

        if city:
            full_address += f", {city}"
        if state:
            full_address += f", {state}"
        if cep:
            full_address += f", {cep}"

        # Add special requirements to notes
        special_requirements = []
        if accessibility_needs:
            special_requirements.append("Precisa de profissional com experiência em acessibilidade")
        if insurance_required:
            special_requirements.append("Precisa de profissional com seguro de responsabilidade civil")

        if special_requirements:
            special_notes = "Requisitos especiais: " + "; ".join(special_requirements)
            if notes:
                notes += f"\n\n{special_notes}"
            else:
                notes = special_notes

        # Create order
        try:
            order = Order.objects.create(
                customer=request.user,
                service=None,  # No standard service for custom services
                professional=custom_service.provider,
                scheduled_date=scheduled_datetime_obj,
                address=full_address,
                notes=notes,
                total_price=custom_service.estimated_price,
                # Service details
                service_name=custom_service.name,
                service_description=custom_service.description,
                service_category=custom_service.category,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email
            )

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Solicitação de serviço enviada com sucesso!',
                    'redirect_url': reverse('order_confirmation', args=[order.id])
                })
            messages.success(request, 'Solicitação de serviço enviada com sucesso!')
            return redirect('order_confirmation', order_id=order.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating order: {e}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Erro ao criar solicitação. Por favor, tente novamente mais tarde.'})
            messages.error(request, 'Erro ao criar solicitação. Por favor, tente novamente mais tarde.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

    # For GET requests, show the form
    return render(request, 'services/request_service_new.html', {
        'custom_service': custom_service
    })


@login_required
def request_custom_service(request, custom_service_id):
    """Request a custom service with a dedicated page"""
    # Check if user is a customer
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'customer':
            messages.error(request, 'Acesso negado. Apenas clientes podem solicitar serviços.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado. Perfil de usuário não encontrado.')
        return redirect('home')

    try:
        custom_service = CustomService.objects.select_related('provider').get(id=custom_service_id)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado. O serviço pode ter sido removido.')
        return redirect('search_new')

    if request.method == 'POST':
        # Get form data
        scheduled_datetime = request.POST.get('scheduled_datetime')
        time_preference = request.POST.get('time_preference')
        urgency = request.POST.get('urgency')
        address = request.POST.get('address')
        number = request.POST.get('number')
        complement = request.POST.get('complement', '')
        city = request.POST.get('city', '')
        state = request.POST.get('state', '')
        cep = request.POST.get('cep', '')
        contact_name = request.POST.get('contact_name', '')
        contact_phone = request.POST.get('contact_phone', '')
        contact_email = request.POST.get('contact_email', '')
        notes = request.POST.get('notes', '')
        accessibility_needs = request.POST.get('accessibility_needs')
        insurance_required = request.POST.get('insurance_required')
        terms_accepted = request.POST.get('terms_accepted')

        # Validate required fields
        missing_fields = []
        if not scheduled_datetime:
            missing_fields.append('data e hora preferidas')
        if not address:
            missing_fields.append('endereço')
        if not number:
            missing_fields.append('número')
        if not city:
            missing_fields.append('cidade')
        if not state:
            missing_fields.append('estado')
        if not contact_name:
            missing_fields.append('nome completo')
        if not contact_phone:
            missing_fields.append('telefone')
        if not contact_email:
            missing_fields.append('e-mail')
        if not terms_accepted:
            missing_fields.append('aceite dos termos e condições')

        if missing_fields:
            error_message = f'Por favor, preencha os seguintes campos obrigatórios: {", ".join(missing_fields)}.'
            messages.error(request, error_message)
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Validate email format
        import re
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_regex, contact_email):
            messages.error(request, 'Por favor, informe um e-mail válido.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Validate phone format (Brazilian format)
        phone_regex = r'^\(?[0-9]{2}\)? [0-9]{4,5}-[0-9]{4}$'
        if not re.match(phone_regex, contact_phone):
            messages.error(request, 'Por favor, informe um telefone válido no formato (61) 98196-1144.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Combine date and time
        from datetime import datetime
        try:
            scheduled_datetime_obj = datetime.strptime(scheduled_datetime, "%Y-%m-%dT%H:%M")
        except ValueError:
            messages.error(request, 'Data ou hora inválida. Por favor, selecione uma data e hora válidas.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Check if scheduled date is in the future
        from datetime import datetime
        if scheduled_datetime_obj < datetime.now():
            messages.error(request, 'A data e hora devem ser futuras. Por favor, selecione uma data futura.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

        # Combine address components
        address_components = [address, number]
        if complement:
            address_components.append(complement)
        full_address = ', '.join(address_components)

        if city:
            full_address += f", {city}"
        if state:
            full_address += f", {state}"
        if cep:
            full_address += f", {cep}"

        # Add special requirements to notes
        special_requirements = []
        if accessibility_needs:
            special_requirements.append("Precisa de profissional com experiência em acessibilidade")
        if insurance_required:
            special_requirements.append("Precisa de profissional com seguro de responsabilidade civil")

        if special_requirements:
            special_notes = "Requisitos especiais: " + "; ".join(special_requirements)
            if notes:
                notes += f"\n\n{special_notes}"
            else:
                notes = special_notes

        # Create order
        try:
            order = Order.objects.create(
                customer=request.user,
                service=None,  # No standard service for custom services
                professional=custom_service.provider,
                scheduled_date=scheduled_datetime_obj,
                address=full_address,
                notes=notes,
                total_price=custom_service.estimated_price,
                # Service details
                service_name=custom_service.name,
                service_description=custom_service.description,
                service_category=custom_service.category,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email
            )

            messages.success(request, 'Solicitação de serviço enviada com sucesso!')
            return redirect('order_confirmation', order_id=order.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating order: {e}')
            messages.error(request, 'Erro ao criar solicitação. Por favor, tente novamente mais tarde.')
            return render(request, 'services/request_service_new.html', {
                'custom_service': custom_service
            })

    # For GET requests, show the form
    return render(request, 'services/request_service_new.html', {
        'custom_service': custom_service
    })


@login_required
def order_confirmation(request, order_id):
    """Order confirmation page"""
    order = Order.objects.select_related('service', 'customer').get(id=order_id)

    # Enviar notificação de confirmação
    try:
        from .notifications import NotificationService
        NotificationService.send_order_confirmation(request.user, order)
    except:
        pass  # Silently fail to avoid breaking the flow

    return render(request, 'services/order_confirmation.html', {'order': order})


@login_required
def order_payment(request, order_id):
    """Order payment page"""
    order = Order.objects.select_related('service', 'customer').get(id=order_id)

    # Check if the order belongs to the current user
    if order.customer != request.user:
        messages.error(request, 'Acesso negado')
        return redirect('home')

    # Get user's payment methods
    payment_methods = PaymentMethod.objects.filter(user=request.user)

    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')
        # In a real application, you would process the payment here
        # For now, we'll just update the order status
        order.status = 'confirmed'
        order.save()

        messages.success(request, 'Pagamento realizado com sucesso!')
        return redirect('order_confirmation', order_id=order.id)

    return render(request, 'services/order_payment.html', {
        'order': order,
        'payment_methods': payment_methods
    })


@login_required
def requested_services(request):
    """Requested services page - shows all services requested by the user"""
    # Get user's orders
    user_orders = Order.objects.filter(customer=request.user).order_by('-created_at')

    if request.method == 'POST':
        # Handle bulk payment
        order_ids = request.POST.getlist('order_ids')
        if order_ids:
            # Store selected order IDs in session for payment page
            request.session['bulk_payment_order_ids'] = order_ids
            return redirect('bulk_payment')
        else:
            messages.error(request, 'Por favor, selecione pelo menos um pedido para pagar.')

    return render(request, 'services/requested_services.html', {
        'user_orders': user_orders
    })


@login_required
def bulk_payment(request):
    """Bulk payment page for multiple orders"""
    # Get order IDs from session
    order_ids = request.session.get('bulk_payment_order_ids', [])

    if not order_ids:
        messages.error(request, 'Nenhum pedido selecionado para pagamento.')
        return redirect('requested_services')

    # Get selected orders
    selected_orders = Order.objects.filter(
        id__in=order_ids,
        customer=request.user,
        status='pending'
    )

    if not selected_orders:
        messages.error(request, 'Pedidos selecionados não são válidos ou já foram pagos.')
        return redirect('requested_services')

    # Calculate total amount
    total_amount = sum(order.total_price for order in selected_orders)

    # Get user's payment methods
    payment_methods = PaymentMethod.objects.filter(user=request.user)

    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')

        # In a real application, you would process the payment here
        # For now, we'll just update the order statuses
        for order in selected_orders:
            order.status = 'confirmed'
            order.save()

        # Clear session data
        if 'bulk_payment_order_ids' in request.session:
            del request.session['bulk_payment_order_ids']

        messages.success(request, f'Pagamento de R$ {total_amount:.2f} realizado com sucesso para {selected_orders.count()} pedido(s)!')
        return redirect('requested_services')

    return render(request, 'services/bulk_payment.html', {
        'selected_orders': selected_orders,
        'total_amount': total_amount,
        'payment_methods': payment_methods
    })


@login_required
def admin_dashboard(request):
    """Admin dashboard page"""
    if not request.user.is_superuser:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Total users
    total_users = User.objects.count()
    
    # Total services
    total_services = CustomService.objects.filter(is_active=True).count()
    
    # Total orders
    total_orders = Order.objects.count()
    
    # Total sponsors
    total_sponsors = Sponsor.objects.count()
    
    # Recent orders
    recent_orders = Order.objects.all().order_by('-created_at')[:10]
    
    return render(request, 'services/admin_dashboard.html', {
        'total_users': total_users,
        'total_services': total_services,
        'total_orders': total_orders,
        'total_sponsors': total_sponsors,
        'recent_orders': recent_orders
    })


# REMOVED: Service request functions have been removed as per requirements


@login_required
def requested_services(request):
    """Requested services page - shows all services requested by the user"""
    # Get user's orders
    user_orders = Order.objects.filter(customer=request.user).order_by('-created_at')

    if request.method == 'POST':
        # Handle bulk payment
        order_ids = request.POST.getlist('order_ids')
        if order_ids:
            # Store selected order IDs in session for payment page
            request.session['bulk_payment_order_ids'] = order_ids
            return redirect('bulk_payment')
        else:
            messages.error(request, 'Por favor, selecione pelo menos um pedido para pagar.')

    return render(request, 'services/requested_services.html', {
        'user_orders': user_orders
    })


import logging


def search_new(request):
    """New search page with database integration"""
    print("🔍 Search view called")
    print(f"📋 GET parameters: {dict(request.GET)}")
    
    # Get search parameters from GET request
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '').strip()
    rating_filter = request.GET.get('rating', '').strip()
    price_min = request.GET.get('price_min', '').strip()
    price_max = request.GET.get('price_max', '').strip()
    location_filter = request.GET.get('location', '').strip()
    
    print(f"🎯 Filtros extraídos:")
    print(f"   - Busca: '{search_query}'")
    print(f"   - Categoria: '{category_filter}'")
    print(f"   - Avaliação: '{rating_filter}'")
    print(f"   - Preço min: '{price_min}'")
    print(f"   - Preço max: '{price_max}'")
    print(f"   - Localização: '{location_filter}'")
    
    # Start with all active custom services
    custom_services_queryset = CustomService.objects.filter(is_active=True).select_related('provider__userprofile')
    initial_count = custom_services_queryset.count()
    print(f"📊 Total inicial de serviços: {initial_count}")
    
    # Apply search query filter
    if search_query:
        custom_services_queryset = custom_services_queryset.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(provider__first_name__icontains=search_query) |
            Q(provider__last_name__icontains=search_query) |
            Q(provider__username__icontains=search_query)
        )
        print(f"🔍 Após filtro de busca '{search_query}': {custom_services_queryset.count()} serviços")
    
    # Apply category filter
    if category_filter:
        custom_services_queryset = custom_services_queryset.filter(category=category_filter)
        print(f"🏷️ Após filtro de categoria '{category_filter}': {custom_services_queryset.count()} serviços")
    
    # Apply rating filter
    if rating_filter:
        # Convert rating filter to minimum rating
        try:
            min_rating = float(rating_filter)
            custom_services_queryset = custom_services_queryset.filter(provider__userprofile__rating__gte=min_rating)
            print(f"⭐ Após filtro de avaliação >= {min_rating}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ Erro ao converter rating '{rating_filter}' para float")
            pass
    
    # Apply price filters
    if price_min:
        try:
            min_price = float(price_min)
            custom_services_queryset = custom_services_queryset.filter(estimated_price__gte=min_price)
            print(f"💰 Após filtro preço mínimo >= R$ {min_price}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ Erro ao converter price_min '{price_min}' para float")
            pass
    
    if price_max:
        try:
            max_price = float(price_max)
            custom_services_queryset = custom_services_queryset.filter(estimated_price__lte=max_price)
            print(f"💰 Após filtro preço máximo <= R$ {max_price}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ Erro ao converter price_max '{price_max}' para float")
            pass
    
    # Apply location filter
    if location_filter:
        custom_services_queryset = custom_services_queryset.filter(
            Q(provider__userprofile__city__icontains=location_filter) |
            Q(provider__userprofile__state__icontains=location_filter)
        )
        print(f"📍 Após filtro de localização '{location_filter}': {custom_services_queryset.count()} serviços")
    
    # Get all categories for the filter dropdown
    categories = CustomService.CATEGORY_CHOICES
    
    # Order by rating (highest first) and then by price (lowest first)
    custom_services_queryset = custom_services_queryset.order_by('-provider__userprofile__rating', 'estimated_price')
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(custom_services_queryset, 12)  # Show 12 services per page
    page_number = request.GET.get('page')
    custom_services = paginator.get_page(page_number)
    
    # Debug information
    print(f"📄 Paginação: {len(custom_services)} serviços na página atual")
    print(f"📊 Total final após todos os filtros: {paginator.count} serviços")
    print(f"🎯 Resultado: {initial_count} → {paginator.count} serviços")
    
    context = {
        'custom_services': custom_services,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'rating_filter': rating_filter,
        'price_min': price_min,
        'price_max': price_max,
        'location_filter': location_filter,
    }
    
    print("Rendering search template")  # Debug print
    return render(request, 'services/search_new.html', context)


def search_ajax(request):
    """AJAX endpoint for filtering services"""
    from django.http import JsonResponse
    from django.template.loader import render_to_string
    
    print("🔥 AJAX Search called")
    print(f"📋 AJAX GET parameters: {dict(request.GET)}")
    
    # Get search parameters from GET request
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '').strip()
    rating_filter = request.GET.get('rating', '').strip()
    price_min = request.GET.get('price_min', '').strip()
    price_max = request.GET.get('price_max', '').strip()
    location_filter = request.GET.get('location', '').strip()
    
    print(f"🎯 AJAX Filtros extraídos:")
    print(f"   - Busca: '{search_query}'")
    print(f"   - Categoria: '{category_filter}'")
    print(f"   - Avaliação: '{rating_filter}'")
    print(f"   - Preço min: '{price_min}'")
    print(f"   - Preço max: '{price_max}'")
    print(f"   - Localização: '{location_filter}'")
    
    # Start with all active custom services
    custom_services_queryset = CustomService.objects.filter(is_active=True).select_related('provider__userprofile')
    initial_count = custom_services_queryset.count()
    print(f"📊 AJAX Total inicial de serviços: {initial_count}")
    
    # Apply search query filter
    if search_query:
        custom_services_queryset = custom_services_queryset.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(provider__first_name__icontains=search_query) |
            Q(provider__last_name__icontains=search_query) |
            Q(provider__username__icontains=search_query)
        )
        print(f"🔍 AJAX Após filtro de busca '{search_query}': {custom_services_queryset.count()} serviços")
    
    # Apply category filter
    if category_filter:
        custom_services_queryset = custom_services_queryset.filter(category=category_filter)
        print(f"🏷️ AJAX Após filtro de categoria '{category_filter}': {custom_services_queryset.count()} serviços")
    
    # Apply rating filter
    if rating_filter:
        try:
            min_rating = float(rating_filter)
            custom_services_queryset = custom_services_queryset.filter(provider__userprofile__rating__gte=min_rating)
            print(f"⭐ AJAX Após filtro de avaliação >= {min_rating}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ AJAX Erro ao converter rating '{rating_filter}' para float")
            pass
    
    # Apply price filters
    if price_min:
        try:
            min_price = float(price_min)
            custom_services_queryset = custom_services_queryset.filter(estimated_price__gte=min_price)
            print(f"💰 AJAX Após filtro preço mínimo >= R$ {min_price}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ AJAX Erro ao converter price_min '{price_min}' para float")
            pass
    
    if price_max:
        try:
            max_price = float(price_max)
            custom_services_queryset = custom_services_queryset.filter(estimated_price__lte=max_price)
            print(f"💰 AJAX Após filtro preço máximo <= R$ {max_price}: {custom_services_queryset.count()} serviços")
        except (ValueError, TypeError):
            print(f"❌ AJAX Erro ao converter price_max '{price_max}' para float")
            pass
    
    # Apply location filter
    if location_filter:
        custom_services_queryset = custom_services_queryset.filter(
            Q(provider__userprofile__city__icontains=location_filter) |
            Q(provider__userprofile__state__icontains=location_filter)
        )
        print(f"📍 AJAX Após filtro de localização '{location_filter}': {custom_services_queryset.count()} serviços")
    
    # Order by rating (highest first) and then by price (lowest first)
    custom_services_queryset = custom_services_queryset.order_by('-provider__userprofile__rating', 'estimated_price')
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(custom_services_queryset, 12)
    page_number = request.GET.get('page', 1)
    custom_services = paginator.get_page(page_number)
    
    print(f"📄 AJAX Paginação: {len(custom_services)} serviços na página atual")
    print(f"📊 AJAX Total final após todos os filtros: {paginator.count} serviços")
    
    # Render the services HTML
    services_html = render_to_string('services/search_results_partial.html', {
        'custom_services': custom_services
    })
    
    return JsonResponse({
        'success': True,
        'services_html': services_html,
        'total_count': paginator.count,
        'current_page': custom_services.number,
        'total_pages': paginator.num_pages,
        'has_previous': custom_services.has_previous(),
        'has_next': custom_services.has_next(),
    })


@login_required
@login_required
def request_service(request, service_id):
    """Request a service - now works with the modal"""
    # For GET requests, we'll redirect to home since we're using a modal now
    if request.method == 'GET':
        return redirect('home')
    
    # For POST requests, process the form data (this should be handled by the modal)
    # Process the form data
    cep = request.POST.get('cep', '')
    address = request.POST.get('address', '')
    number = request.POST.get('number', '')
    complement = request.POST.get('complement', '')
    city = request.POST.get('city', '')
    state = request.POST.get('state', '')
    date = request.POST.get('date', '')
    time = request.POST.get('time', '')
    urgency = request.POST.get('urgency', 'normal')
    description = request.POST.get('description', '')
    notes = request.POST.get('notes', '')
    
    # Combine address components
    full_address = f"{address}, {number}"
    if complement:
        full_address += f", {complement}"
    full_address += f", {city} - {state}, {cep}"
    
    # Combine date and time
    scheduled_datetime = f"{date} {time}"
    
    # In a real application, you would:
    # 1. Save this data to the database
    # 2. Associate it with the logged-in user
    # 3. Process payment
    # 4. Send confirmation email
    # 5. Redirect to confirmation page
    
    # For demonstration, we'll create a mock order
    # In a real implementation, you would save to the database
    order_id = 123  # This would be the actual order ID from the database
    
    messages.success(request, 'Solicitação de serviço enviada com sucesso!')
    
    # Redirect to confirmation page
    return redirect('order_confirmation', order_id=order_id)


def submit_service_review(request, order_id):
    """API endpoint to submit a service review"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the order
        order = Order.objects.get(id=order_id, customer=request.user)
        
        # Check if order is completed
        if order.status != 'completed':
            return JsonResponse({'error': 'Only completed orders can be reviewed'}, status=400)
        
        # Get review data
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '')
        
        # Validate rating
        try:
            rating = int(rating)
            if rating < 1 or rating > 5:
                raise ValueError()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid rating value'}, status=400)
        
        # In a real implementation, you would:
        # 1. Save the review to the database
        # 2. Update the provider's rating
        # 3. Update the provider's review count
        
        # For now, we'll just return a success response
        return JsonResponse({
            'success': True,
            'message': 'Review submitted successfully'
        })
        
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


def get_service_details(request, service_id):
    """API endpoint to get service details for the modal"""
    try:
        # Get the custom service with related data
        custom_service = CustomService.objects.select_related('provider__userprofile').get(id=service_id, is_active=True)
        
        # Prepare the response data
        provider_profile = getattr(custom_service.provider, 'userprofile', None)
        provider_rating = str(provider_profile.rating) if provider_profile and provider_profile.rating else '0.0'
        provider_review_count = getattr(provider_profile, 'review_count', 0) if provider_profile else 0
        
        data = {
            'id': custom_service.id,
            'name': custom_service.name,
            'description': custom_service.description,
            'category': custom_service.get_category_display(),
            'price': str(custom_service.estimated_price),
            'rating': provider_rating,
            'review_count': provider_review_count,
            'provider': {
                'id': custom_service.provider.id,
                'name': custom_service.provider.get_full_name() or custom_service.provider.username,
                'first_name': custom_service.provider.first_name,
                'last_name': custom_service.provider.last_name,
                'rating': provider_rating,
                'review_count': provider_review_count,
            }
        }
        
        return JsonResponse(data)
    except CustomService.DoesNotExist:
        return JsonResponse({'error': 'Service not found'}, status=404)
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error fetching service details for service {service_id}: {e}')
        return JsonResponse({'error': 'Internal server error'}, status=500)


def service_history_new(request):
    """New service history page"""
    try:
        # Get user's orders with related data
        user_orders_query = Order.objects.filter(customer=request.user).select_related(
            'service', 'professional'
        ).order_by('-created_at')
        
        # Try to also select professional profile if it exists
        try:
            user_orders_query = user_orders_query.select_related('professional__userprofile')
        except Exception:
            pass  # If userprofile doesn't exist, continue without it
        
        # Get all user orders for statistics using annotations for better performance
        from django.db.models import Count, Q
        order_stats = Order.objects.filter(customer=request.user).aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            in_progress=Count('id', filter=Q(status='in_progress')),
            cancelled=Count('id', filter=Q(status='cancelled'))
        )
        
        total_orders = order_stats['total']
        completed_orders = order_stats['completed']
        in_progress_orders = order_stats['in_progress']
        cancelled_orders = order_stats['cancelled']
        
        # Get filter parameters
        period = request.GET.get('period', 'all')
        status = request.GET.get('status', 'all')
        service_search = request.GET.get('service', '')
        
        # Apply filters
        if period != 'all':
            from datetime import datetime, timedelta
            now = datetime.now()
            if period == '30_days':
                start_date = now - timedelta(days=30)
                user_orders_query = user_orders_query.filter(created_at__gte=start_date)
            elif period == '3_months':
                start_date = now - timedelta(days=90)
                user_orders_query = user_orders_query.filter(created_at__gte=start_date)
            elif period == '6_months':
                start_date = now - timedelta(days=180)
                user_orders_query = user_orders_query.filter(created_at__gte=start_date)
            elif period == '1_year':
                start_date = now - timedelta(days=365)
                user_orders_query = user_orders_query.filter(created_at__gte=start_date)
        
        if status != 'all':
            user_orders_query = user_orders_query.filter(status=status)
        
        if service_search:
            user_orders_query = user_orders_query.filter(
                Q(service__name__icontains=service_search) |
                Q(professional__first_name__icontains=service_search) |
                Q(professional__last_name__icontains=service_search)
            )
        
        # Add pagination
        from django.core.paginator import Paginator
        paginator = Paginator(user_orders_query, 10)  # Show 10 orders per page
        page_number = request.GET.get('page')
        user_orders = paginator.get_page(page_number)
        
    except Exception as e:
        # Handle any database errors
        if "no such column" in str(e):
            # Return empty result set if columns don't exist
            from django.core.paginator import Paginator
            paginator = Paginator([], 10)
            user_orders = paginator.get_page(1)
            total_orders = 0
            completed_orders = 0
            in_progress_orders = 0
            cancelled_orders = 0
        else:
            # Re-raise the exception if it's not related to missing columns
            raise e
    
    return render(request, 'services/service_history_new.html', {
        'user_orders': user_orders,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'in_progress_orders': in_progress_orders,
        'cancelled_orders': cancelled_orders
    })


def sponsors_new(request):
    """New sponsors page"""
    sponsors = Sponsor.objects.filter(is_active=True)
    return render(request, 'services/sponsors_new.html')


def partnership_details(request, level):
    """Display details for a specific partnership level"""
    # Partnership level details
    partnership_levels = {
        'basic': {
            'name': 'Básico',
            'icon': 'fa-medal',
            'description': 'O nível Básico é perfeito para empresas que estão começando a expandir sua presença no mercado de serviços domésticos.',
            'benefits': [
                'Presença no diretório de parceiros',
                'Destaque em buscas relevantes',
                'Até 5% de desconto para clientes',
                'Acesso ao painel de parceiro básico',
                'Suporte por email durante horário comercial'
            ],
            'price': 'Gratuito',
            'call_to_action': 'Torne-se um parceiro Básico'
        },
        'premium': {
            'name': 'Premium',
            'icon': 'fa-award',
            'description': 'O nível Premium é ideal para empresas que desejam maximizar sua visibilidade e atrair mais clientes.',
            'benefits': [
                'Todos os benefícios do plano Básico',
                'Destaque premium em resultados',
                'Até 10% de desconto para clientes',
                'Relatórios mensais de performance',
                'Destaque em newsletters',
                'Suporte prioritário por email e chat',
                'Acesso ao painel de parceiro avançado'
            ],
            'price': 'R$ 299/mês',
            'call_to_action': 'Torne-se um parceiro Premium'
        },
        'platinum': {
            'name': 'Platina',
            'icon': 'fa-crown',
            'description': 'O nível Platina é nosso pacote mais completo, oferecendo máxima visibilidade e suporte personalizado.',
            'benefits': [
                'Todos os benefícios do plano Premium',
                'Posicionamento privilegiado',
                'Até 15% de desconto para clientes',
                'Suporte dedicado 24/7',
                'Campanhas personalizadas',
                'Destaque em todas as newsletters',
                'Acesso ao painel de parceiro premium',
                'Consultoria estratégica mensal',
                'Participação em eventos exclusivos'
            ],
            'price': 'R$ 599/mês',
            'call_to_action': 'Torne-se um parceiro Platina'
        }
    }
    
    # Get the requested level details or return 404
    if level not in partnership_levels:
        from django.http import Http404
        raise Http404("Nível de parceria não encontrado")
    
    level_data = partnership_levels[level]
    
    return render(request, 'services/partnership_details.html', {
        'level': level,
        'level_data': level_data
    })


def admin_dashboard_new(request):
    """New admin dashboard page"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado - perfil de usuário não encontrado')
        return redirect('home')
    except Exception as e:
        # Handle the case where the database columns don't exist yet
        if "no such column" in str(e):
            messages.error(request, 'Acesso negado - erro no banco de dados')
            return redirect('home')
        else:
            messages.error(request, f'Erro desconhecido: {str(e)}')
            return redirect('home')
    
    # Get dashboard data using annotations for better performance
    from django.db.models import Count, Sum, Q
    
    # Get counts in a single query
    user_counts = UserProfile.objects.values('user_type').annotate(count=Count('user_type'))
    user_type_counts = {item['user_type']: item['count'] for item in user_counts}
    
    total_users = User.objects.count()
    total_services = Service.objects.count()
    total_orders = Order.objects.count()
    total_sponsors = Sponsor.objects.filter(is_active=True).count()
    
    # Get counts by user type
    professionals_count = user_type_counts.get('professional', 0)
    customers_count = user_type_counts.get('customer', 0)
    admins_count = user_type_counts.get('admin', 0)
    
    # Calculate revenue more efficiently
    total_revenue = Order.objects.aggregate(total=Sum('total_price'))['total'] or 0
    
    # Get recent orders (real data) with optimized query
    recent_orders = Order.objects.select_related('customer', 'service').order_by('-created_at')[:5]
    
    # Get recent users (real data) with optimized query
    recent_users = User.objects.select_related('userprofile').order_by('-date_joined')[:5]
    
    # Get contact messages
    from .models import ContactMessage, ServiceRequestModal
    contact_messages = ContactMessage.objects.all().order_by('-created_at')[:10]
    new_messages_count = ContactMessage.objects.filter(status='new').count()
    
    # Get service requests statistics
    total_requests = ServiceRequestModal.objects.count()
    pending_requests = ServiceRequestModal.objects.filter(status='pending').count()
    completed_requests = ServiceRequestModal.objects.filter(status='completed').count()
    
    # Calculate percentages for user types
    total_users_for_calc = total_users if total_users > 0 else 1
    customers_percentage = round((customers_count / total_users_for_calc) * 100)
    professionals_percentage = round((professionals_count / total_users_for_calc) * 100)
    admins_percentage = round((admins_count / total_users_for_calc) * 100)
    
    # Calculate percentages for orders
    total_orders_for_calc = total_orders if total_orders > 0 else 1
    completed_percentage = round((completed_requests / total_orders_for_calc) * 100)
    pending_percentage = round((pending_requests / total_orders_for_calc) * 100)
    
    # Get recent service requests
    recent_requests = ServiceRequestModal.objects.select_related('user', 'provider').order_by('-created_at')[:5]
    
    # Get providers with their services (show all, ordered by most recent)
    providers = UserProfile.objects.filter(user_type='professional').select_related('user').prefetch_related('user__custom_services').order_by('-created_at')
    
    # Get all customers with their recent activity
    customers = UserProfile.objects.filter(user_type='customer').select_related('user').prefetch_related('user__modal_service_requests').annotate(
        request_count=Count('user__modal_service_requests')
    ).order_by('-request_count')
    
    # Get user growth data for the last 7 days
    from datetime import datetime, timedelta
    today = timezone.now().date()
    user_growth_data = []
    user_growth_labels = []
    
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        count = User.objects.filter(date_joined__date=date).count()
        user_growth_data.append(count)
        user_growth_labels.append(date.strftime('%d/%m'))
    
    # Get services by category
    from django.db.models import Count
    services_by_category = Service.objects.values('category').annotate(count=Count('id')).order_by('-count')
    category_labels = [item['category'] for item in services_by_category]
    category_data = [item['count'] for item in services_by_category]
    
    # Get revenue data for the last 6 months
    revenue_data = []
    revenue_labels = []
    
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        month_revenue = Order.objects.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        ).aggregate(total=Sum('total_price'))['total'] or 0
        
        revenue_data.append(float(month_revenue))
        revenue_labels.append(month_start.strftime('%b'))
    
    # Get orders by status
    orders_by_status = Order.objects.values('status').annotate(count=Count('id'))
    status_labels = []
    status_data = []
    status_map = {
        'pending': 'Pendente',
        'in_progress': 'Em Progresso',
        'completed': 'Concluído',
        'cancelled': 'Cancelado'
    }
    
    for item in orders_by_status:
        status_labels.append(status_map.get(item['status'], item['status']))
        status_data.append(item['count'])
    
    import json
    
    return render(request, 'services/admin_dashboard_new.html', {
        'users_count': total_users,
        'services_count': total_services,
        'orders_count': total_orders,
        'sponsors_count': total_sponsors,
        'professionals_count': professionals_count,
        'customers_count': customers_count,
        'admins_count': admins_count,
        'revenue': total_revenue,
        'revenue_formatted': f"R$ {total_revenue:,.2f}",
        'recent_orders': recent_orders,
        'recent_users': recent_users,
        'contact_messages': contact_messages,
        'new_messages_count': new_messages_count,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'completed_requests': completed_requests,
        'customers_percentage': customers_percentage,
        'professionals_percentage': professionals_percentage,
        'admins_percentage': admins_percentage,
        'completed_percentage': completed_percentage,
        'pending_percentage': pending_percentage,
        'recent_requests': recent_requests,
        'providers': providers,
        'customers': customers,
        # Chart data
        'user_growth_data': json.dumps(user_growth_data),
        'user_growth_labels': json.dumps(user_growth_labels),
        'category_labels': json.dumps(category_labels),
        'category_data': json.dumps(category_data),
        'revenue_data': json.dumps(revenue_data),
        'revenue_labels': json.dumps(revenue_labels),
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
    })


@login_required
def exportar_atividades_admin(request):
    """Exportar atividades do dashboard para CSV"""
    import csv
    from django.http import HttpResponse
    
    # Verificar se é admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=401)
    
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'error': 'Acesso negado'}, status=403)
    except:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=403)
    
    # Criar resposta HTTP com CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="atividades_admin.csv"'
    
    # Adicionar BOM para Excel reconhecer UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')  # Usar ponto e vírgula para melhor compatibilidade com Excel BR
    
    # Título do relatório
    writer.writerow(['RELATÓRIO DE ATIVIDADES - JOB FINDER'])
    writer.writerow(['Data de Geração:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')])
    writer.writerow([])  # Linha em branco
    
    # Cabeçalhos com formatação
    writer.writerow([
        'ID',
        'Nome do Usuário',
        'Email',
        'Telefone',
        'Serviço Solicitado',
        'Categoria',
        'Status',
        'Data de Criação',
        'Última Atualização',
        'Prestador Atribuído',
        'Valor Estimado',
        'Observações'
    ])
    
    # Obter dados
    from .models import ServiceRequestModal
    requests = ServiceRequestModal.objects.select_related('user', 'service', 'provider').order_by('-created_at')
    
    # Escrever dados
    for req in requests:
        writer.writerow([
            f'#{req.id}',
            req.user.get_full_name() if req.user else req.contact_name,
            req.user.email if req.user else req.contact_email,
            req.contact_phone if req.contact_phone else 'Não informado',
            req.service.name if req.service else req.service_name,
            req.service.get_category_display() if req.service else 'N/A',
            req.get_status_display(),
            req.created_at.strftime('%d/%m/%Y %H:%M'),
            req.updated_at.strftime('%d/%m/%Y %H:%M') if req.updated_at else 'N/A',
            req.provider.get_full_name() if req.provider else 'Não atribuído',
            f'R$ {req.estimated_price:.2f}' if req.estimated_price else 'Não informado',
            req.notes[:100] if req.notes else 'Sem observações'
        ])
    
    # Linha de resumo
    writer.writerow([])
    writer.writerow(['RESUMO'])
    writer.writerow(['Total de Atividades:', requests.count()])
    writer.writerow(['Pendentes:', requests.filter(status='pending').count()])
    writer.writerow(['Concluídas:', requests.filter(status='completed').count()])
    writer.writerow(['Canceladas:', requests.filter(status='cancelled').count()])
    
    return response


@login_required
def exportar_grafico_admin(request, tipo_grafico):
    """Exportar dados de gráfico específico para CSV"""
    import csv
    from django.http import HttpResponse
    from django.db.models import Count, Sum
    from datetime import timedelta
    
    # Verificar se é admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=401)
    
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'error': 'Acesso negado'}, status=403)
    except:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=403)
    
    # Criar resposta HTTP com CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    if tipo_grafico == 'usuarios':
        response['Content-Disposition'] = 'attachment; filename="crescimento_usuarios.csv"'
        
        # Cabeçalho do relatório
        writer.writerow(['RELATÓRIO DE CRESCIMENTO DE USUÁRIOS'])
        writer.writerow(['Job Finder - Painel Administrativo'])
        writer.writerow(['Gerado em:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([])
        
        # Cabeçalhos das colunas
        writer.writerow(['Data', 'Dia da Semana', 'Novos Usuários', 'Acumulado'])
        
        today = timezone.now().date()
        total_acumulado = 0
        dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            count = User.objects.filter(date_joined__date=date).count()
            total_acumulado += count
            dia_semana = dias_semana[date.weekday()]
            writer.writerow([date.strftime('%d/%m/%Y'), dia_semana, count, total_acumulado])
        
        # Resumo
        writer.writerow([])
        writer.writerow(['RESUMO'])
        writer.writerow(['Total de Novos Usuários (7 dias):', total_acumulado])
        writer.writerow(['Média Diária:', f'{total_acumulado/7:.1f}'])
    
    elif tipo_grafico == 'categorias':
        response['Content-Disposition'] = 'attachment; filename="servicos_por_categoria.csv"'
        
        # Cabeçalho do relatório
        writer.writerow(['RELATÓRIO DE SERVIÇOS POR CATEGORIA'])
        writer.writerow(['Job Finder - Painel Administrativo'])
        writer.writerow(['Gerado em:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([])
        
        # Cabeçalhos das colunas
        writer.writerow(['Categoria', 'Quantidade', 'Percentual', 'Status'])
        
        services_by_category = Service.objects.values('category').annotate(count=Count('id')).order_by('-count')
        total_services = Service.objects.count()
        
        for item in services_by_category:
            percentual = (item['count'] / total_services * 100) if total_services > 0 else 0
            writer.writerow([
                item['category'],
                item['count'],
                f'{percentual:.1f}%',
                'Ativo'
            ])
        
        # Resumo
        writer.writerow([])
        writer.writerow(['RESUMO'])
        writer.writerow(['Total de Serviços:', total_services])
        writer.writerow(['Categorias Ativas:', services_by_category.count()])
    
    elif tipo_grafico == 'receita':
        response['Content-Disposition'] = 'attachment; filename="receita_mensal.csv"'
        
        # Cabeçalho do relatório
        writer.writerow(['RELATÓRIO DE RECEITA MENSAL'])
        writer.writerow(['Job Finder - Painel Administrativo'])
        writer.writerow(['Gerado em:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([])
        
        # Cabeçalhos das colunas
        writer.writerow(['Mês/Ano', 'Receita (R$)', 'Pedidos', 'Ticket Médio (R$)', 'Variação'])
        
        today = timezone.now().date()
        receita_anterior = 0
        total_receita = 0
        total_pedidos = 0
        
        for i in range(5, -1, -1):
            month_date = today - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            if i == 0:
                month_end = today
            else:
                next_month = month_start + timedelta(days=32)
                month_end = next_month.replace(day=1) - timedelta(days=1)
            
            pedidos_mes = Order.objects.filter(
                created_at__gte=month_start,
                created_at__lte=month_end
            )
            
            month_revenue = pedidos_mes.aggregate(total=Sum('total_price'))['total'] or 0
            num_pedidos = pedidos_mes.count()
            ticket_medio = (month_revenue / num_pedidos) if num_pedidos > 0 else 0
            
            # Calcular variação
            if receita_anterior > 0:
                variacao = ((month_revenue - receita_anterior) / receita_anterior * 100)
                variacao_str = f'{variacao:+.1f}%'
            else:
                variacao_str = 'N/A'
            
            writer.writerow([
                month_start.strftime('%B/%Y'),
                f'R$ {month_revenue:.2f}',
                num_pedidos,
                f'R$ {ticket_medio:.2f}',
                variacao_str
            ])
            
            receita_anterior = month_revenue
            total_receita += month_revenue
            total_pedidos += num_pedidos
        
        # Resumo
        writer.writerow([])
        writer.writerow(['RESUMO'])
        writer.writerow(['Receita Total (6 meses):', f'R$ {total_receita:.2f}'])
        writer.writerow(['Total de Pedidos:', total_pedidos])
        writer.writerow(['Ticket Médio Geral:', f'R$ {total_receita/total_pedidos:.2f}' if total_pedidos > 0 else 'R$ 0,00'])
        writer.writerow(['Receita Média Mensal:', f'R$ {total_receita/6:.2f}'])
    
    elif tipo_grafico == 'status':
        response['Content-Disposition'] = 'attachment; filename="pedidos_por_status.csv"'
        
        # Cabeçalho do relatório
        writer.writerow(['RELATÓRIO DE PEDIDOS POR STATUS'])
        writer.writerow(['Job Finder - Painel Administrativo'])
        writer.writerow(['Gerado em:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([])
        
        # Cabeçalhos das colunas
        writer.writerow(['Status', 'Quantidade', 'Percentual', 'Valor Total (R$)'])
        
        orders_by_status = Order.objects.values('status').annotate(
            count=Count('id'),
            total_value=Sum('total_price')
        )
        
        status_map = {
            'pending': 'Pendente',
            'in_progress': 'Em Progresso',
            'completed': 'Concluído',
            'cancelled': 'Cancelado'
        }
        
        total_orders = Order.objects.count()
        
        for item in orders_by_status:
            status_label = status_map.get(item['status'], item['status'])
            percentual = (item['count'] / total_orders * 100) if total_orders > 0 else 0
            valor_total = item['total_value'] or 0
            writer.writerow([
                status_label,
                item['count'],
                f'{percentual:.1f}%',
                f'R$ {valor_total:.2f}'
            ])
        
        # Resumo
        writer.writerow([])
        writer.writerow(['RESUMO'])
        writer.writerow(['Total de Pedidos:', total_orders])
        total_value = Order.objects.aggregate(total=Sum('total_price'))['total'] or 0
        writer.writerow(['Valor Total:', f'R$ {total_value:.2f}'])
    
    else:
        return JsonResponse({'error': 'Tipo de gráfico inválido'}, status=400)
    
    return response


@login_required
def exportar_relatorio_completo_admin(request):
    """Exportar relatório completo do dashboard para CSV"""
    import csv
    from django.http import HttpResponse
    from django.db.models import Count, Sum, Avg
    from datetime import timedelta
    
    # Verificar se é admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=401)
    
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'error': 'Acesso negado'}, status=403)
    except:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=403)
    
    # Criar resposta HTTP com CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="relatorio_completo_admin.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # Cabeçalho do relatório
    writer.writerow(['═══════════════════════════════════════════════════════'])
    writer.writerow(['RELATÓRIO COMPLETO DO DASHBOARD ADMINISTRATIVO'])
    writer.writerow(['JOB FINDER - SISTEMA DE GESTÃO'])
    writer.writerow(['═══════════════════════════════════════════════════════'])
    writer.writerow([])
    writer.writerow(['Data de Geração:', timezone.now().strftime('%d/%m/%Y às %H:%M:%S')])
    writer.writerow(['Gerado por:', request.user.get_full_name() or request.user.username])
    writer.writerow([])
    writer.writerow(['═══════════════════════════════════════════════════════'])
    
    # Usuários
    writer.writerow([])
    writer.writerow(['📊 ESTATÍSTICAS DE USUÁRIOS'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Métrica', 'Valor', 'Observação'])
    
    total_users = User.objects.count()
    professionals = UserProfile.objects.filter(user_type='professional').count()
    customers = UserProfile.objects.filter(user_type='customer').count()
    admins = UserProfile.objects.filter(user_type='admin').count()
    
    # Usuários dos últimos 30 dias
    last_30_days = timezone.now() - timedelta(days=30)
    new_users_30d = User.objects.filter(date_joined__gte=last_30_days).count()
    
    writer.writerow(['Total de Usuários', total_users, '100%'])
    writer.writerow(['Prestadores', professionals, f'{(professionals/total_users*100):.1f}%' if total_users > 0 else '0%'])
    writer.writerow(['Clientes', customers, f'{(customers/total_users*100):.1f}%' if total_users > 0 else '0%'])
    writer.writerow(['Administradores', admins, f'{(admins/total_users*100):.1f}%' if total_users > 0 else '0%'])
    writer.writerow(['Novos (últimos 30 dias)', new_users_30d, f'{(new_users_30d/total_users*100):.1f}%' if total_users > 0 else '0%'])
    
    # Serviços
    writer.writerow([])
    writer.writerow(['🛠️ ESTATÍSTICAS DE SERVIÇOS'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Métrica', 'Valor', 'Observação'])
    
    total_services = Service.objects.count()
    active_services = Service.objects.filter(is_active=True).count()
    inactive_services = total_services - active_services
    
    writer.writerow(['Total de Serviços', total_services, '100%'])
    writer.writerow(['Serviços Ativos', active_services, f'{(active_services/total_services*100):.1f}%' if total_services > 0 else '0%'])
    writer.writerow(['Serviços Inativos', inactive_services, f'{(inactive_services/total_services*100):.1f}%' if total_services > 0 else '0%'])
    
    # Serviços por categoria
    writer.writerow([])
    writer.writerow(['Distribuição por Categoria:'])
    services_by_category = Service.objects.values('category').annotate(count=Count('id')).order_by('-count')
    for item in services_by_category[:5]:  # Top 5 categorias
        writer.writerow(['', item['category'], item['count']])
    
    # Pedidos
    writer.writerow([])
    writer.writerow(['📦 ESTATÍSTICAS DE PEDIDOS'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Métrica', 'Valor', 'Percentual'])
    
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    in_progress_orders = Order.objects.filter(status='in_progress').count()
    completed_orders = Order.objects.filter(status='completed').count()
    cancelled_orders = Order.objects.filter(status='cancelled').count()
    
    writer.writerow(['Total de Pedidos', total_orders, '100%'])
    writer.writerow(['Pendentes', pending_orders, f'{(pending_orders/total_orders*100):.1f}%' if total_orders > 0 else '0%'])
    writer.writerow(['Em Progresso', in_progress_orders, f'{(in_progress_orders/total_orders*100):.1f}%' if total_orders > 0 else '0%'])
    writer.writerow(['Concluídos', completed_orders, f'{(completed_orders/total_orders*100):.1f}%' if total_orders > 0 else '0%'])
    writer.writerow(['Cancelados', cancelled_orders, f'{(cancelled_orders/total_orders*100):.1f}%' if total_orders > 0 else '0%'])
    
    # Taxa de conclusão
    taxa_conclusao = (completed_orders / total_orders * 100) if total_orders > 0 else 0
    writer.writerow(['Taxa de Conclusão', f'{taxa_conclusao:.1f}%', 'Meta: 80%'])
    
    # Receita
    writer.writerow([])
    writer.writerow(['💰 ESTATÍSTICAS FINANCEIRAS'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Métrica', 'Valor (R$)', 'Observação'])
    
    total_revenue = Order.objects.aggregate(total=Sum('total_price'))['total'] or 0
    completed_revenue = Order.objects.filter(status='completed').aggregate(total=Sum('total_price'))['total'] or 0
    pending_revenue = Order.objects.filter(status='pending').aggregate(total=Sum('total_price'))['total'] or 0
    avg_order = Order.objects.aggregate(avg=Avg('total_price'))['avg'] or 0
    
    writer.writerow(['Receita Total', f'R$ {total_revenue:.2f}', 'Todos os pedidos'])
    writer.writerow(['Receita Confirmada', f'R$ {completed_revenue:.2f}', 'Pedidos concluídos'])
    writer.writerow(['Receita Pendente', f'R$ {pending_revenue:.2f}', 'Pedidos pendentes'])
    writer.writerow(['Ticket Médio', f'R$ {avg_order:.2f}', 'Por pedido'])
    
    # Receita dos últimos 30 dias
    revenue_30d = Order.objects.filter(
        created_at__gte=last_30_days
    ).aggregate(total=Sum('total_price'))['total'] or 0
    writer.writerow(['Receita (30 dias)', f'R$ {revenue_30d:.2f}', 'Último mês'])
    
    # Solicitações
    writer.writerow([])
    writer.writerow(['📋 ESTATÍSTICAS DE SOLICITAÇÕES'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Métrica', 'Valor', 'Percentual'])
    
    from .models import ServiceRequestModal
    total_requests = ServiceRequestModal.objects.count()
    pending_requests = ServiceRequestModal.objects.filter(status='pending').count()
    completed_requests = ServiceRequestModal.objects.filter(status='completed').count()
    cancelled_requests = ServiceRequestModal.objects.filter(status='cancelled').count()
    
    writer.writerow(['Total de Solicitações', total_requests, '100%'])
    writer.writerow(['Pendentes', pending_requests, f'{(pending_requests/total_requests*100):.1f}%' if total_requests > 0 else '0%'])
    writer.writerow(['Concluídas', completed_requests, f'{(completed_requests/total_requests*100):.1f}%' if total_requests > 0 else '0%'])
    writer.writerow(['Canceladas', cancelled_requests, f'{(cancelled_requests/total_requests*100):.1f}%' if total_requests > 0 else '0%'])
    
    # Taxa de conversão
    taxa_conversao = (completed_requests / total_requests * 100) if total_requests > 0 else 0
    writer.writerow(['Taxa de Conversão', f'{taxa_conversao:.1f}%', 'Solicitações → Concluídas'])
    
    # Indicadores de Performance
    writer.writerow([])
    writer.writerow(['📈 INDICADORES DE PERFORMANCE (KPIs)'])
    writer.writerow(['───────────────────────────────────────────────────────'])
    writer.writerow(['Indicador', 'Valor', 'Status'])
    
    # Taxa de crescimento de usuários
    last_60_days = timezone.now() - timedelta(days=60)
    users_30_60 = User.objects.filter(date_joined__gte=last_60_days, date_joined__lt=last_30_days).count()
    crescimento = ((new_users_30d - users_30_60) / users_30_60 * 100) if users_30_60 > 0 else 0
    writer.writerow(['Crescimento de Usuários', f'{crescimento:+.1f}%', '✓ Positivo' if crescimento > 0 else '✗ Negativo'])
    
    writer.writerow(['Taxa de Conclusão de Pedidos', f'{taxa_conclusao:.1f}%', '✓ Bom' if taxa_conclusao >= 70 else '⚠ Atenção'])
    writer.writerow(['Taxa de Conversão', f'{taxa_conversao:.1f}%', '✓ Bom' if taxa_conversao >= 60 else '⚠ Atenção'])
    writer.writerow(['Ticket Médio', f'R$ {avg_order:.2f}', '✓ Bom' if avg_order >= 100 else '⚠ Atenção'])
    
    # Rodapé
    writer.writerow([])
    writer.writerow(['═══════════════════════════════════════════════════════'])
    writer.writerow(['FIM DO RELATÓRIO'])
    writer.writerow(['═══════════════════════════════════════════════════════'])
    
    return response


@login_required
def exportar_excel_profissional(request, tipo='completo'):
    """Exportar dados para Excel com formatação profissional usando openpyxl"""
    from django.http import HttpResponse
    from django.db.models import Count, Sum, Avg
    from datetime import timedelta
    import io
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    except ImportError:
        return JsonResponse({'error': 'Biblioteca openpyxl não instalada'}, status=500)
    
    # Verificar se é admin
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=401)
    
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'error': 'Acesso negado'}, status=403)
    except:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=403)
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover sheet padrão
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2D3748")
    subtitle_font = Font(bold=True, size=12, color="4A5568")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # ABA 1: Dashboard Geral
    ws_dashboard = wb.create_sheet("Dashboard Geral")
    
    # Título
    ws_dashboard['A1'] = 'RELATÓRIO ADMINISTRATIVO - JOB FINDER'
    ws_dashboard['A1'].font = title_font
    ws_dashboard.merge_cells('A1:F1')
    ws_dashboard['A1'].alignment = center_align
    
    ws_dashboard['A2'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}'
    ws_dashboard.merge_cells('A2:F2')
    ws_dashboard['A2'].alignment = center_align
    
    # Estatísticas de Usuários
    row = 4
    ws_dashboard[f'A{row}'] = '📊 USUÁRIOS'
    ws_dashboard[f'A{row}'].font = subtitle_font
    ws_dashboard.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws_dashboard[f'A{row}'] = 'Métrica'
    ws_dashboard[f'B{row}'] = 'Valor'
    for cell in [ws_dashboard[f'A{row}'], ws_dashboard[f'B{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    stats = [
        ('Total de Usuários', User.objects.count()),
        ('Prestadores', UserProfile.objects.filter(user_type='professional').count()),
        ('Clientes', UserProfile.objects.filter(user_type='customer').count()),
        ('Administradores', UserProfile.objects.filter(user_type='admin').count()),
    ]
    
    for label, value in stats:
        row += 1
        ws_dashboard[f'A{row}'] = label
        ws_dashboard[f'B{row}'] = value
        ws_dashboard[f'A{row}'].border = border
        ws_dashboard[f'B{row}'].border = border
        ws_dashboard[f'B{row}'].alignment = center_align
    
    # Estatísticas de Pedidos
    row += 2
    ws_dashboard[f'A{row}'] = '📦 PEDIDOS'
    ws_dashboard[f'A{row}'].font = subtitle_font
    ws_dashboard.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws_dashboard[f'A{row}'] = 'Status'
    ws_dashboard[f'B{row}'] = 'Quantidade'
    for cell in [ws_dashboard[f'A{row}'], ws_dashboard[f'B{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    order_stats = [
        ('Total', Order.objects.count()),
        ('Pendentes', Order.objects.filter(status='pending').count()),
        ('Em Progresso', Order.objects.filter(status='in_progress').count()),
        ('Concluídos', Order.objects.filter(status='completed').count()),
        ('Cancelados', Order.objects.filter(status='cancelled').count()),
    ]
    
    for label, value in order_stats:
        row += 1
        ws_dashboard[f'A{row}'] = label
        ws_dashboard[f'B{row}'] = value
        ws_dashboard[f'A{row}'].border = border
        ws_dashboard[f'B{row}'].border = border
        ws_dashboard[f'B{row}'].alignment = center_align
    
    # Ajustar largura das colunas
    ws_dashboard.column_dimensions['A'].width = 30
    ws_dashboard.column_dimensions['B'].width = 15
    
    # ABA 2: Atividades Detalhadas
    ws_atividades = wb.create_sheet("Atividades")
    
    # Cabeçalhos
    headers = ['ID', 'Usuário', 'Email', 'Serviço', 'Status', 'Data', 'Prestador', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws_atividades.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados
    from .models import ServiceRequestModal
    requests = ServiceRequestModal.objects.select_related('user', 'service', 'provider').order_by('-created_at')[:100]
    
    for row_idx, req in enumerate(requests, 2):
        data = [
            f'#{req.id}',
            req.user.get_full_name() if req.user else req.contact_name,
            req.user.email if req.user else req.contact_email,
            req.service.name if req.service else req.service_name,
            req.get_status_display(),
            req.created_at.strftime('%d/%m/%Y %H:%M'),
            req.provider.get_full_name() if req.provider else 'Não atribuído',
            f'R$ {req.estimated_price:.2f}' if req.estimated_price else 'N/A'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_atividades.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 5, 8]:  # ID, Data, Valor
                cell.alignment = center_align
    
    # Ajustar larguras
    column_widths = [10, 25, 30, 25, 15, 18, 25, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws_atividades.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ABA 3: Receita Mensal
    ws_receita = wb.create_sheet("Receita Mensal")
    
    # Cabeçalhos
    ws_receita['A1'] = 'Mês'
    ws_receita['B1'] = 'Receita (R$)'
    ws_receita['C1'] = 'Pedidos'
    ws_receita['D1'] = 'Ticket Médio'
    
    for cell in [ws_receita['A1'], ws_receita['B1'], ws_receita['C1'], ws_receita['D1']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos últimos 6 meses
    today = timezone.now().date()
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        pedidos_mes = Order.objects.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        )
        
        month_revenue = pedidos_mes.aggregate(total=Sum('total_price'))['total'] or 0
        num_pedidos = pedidos_mes.count()
        ticket_medio = (month_revenue / num_pedidos) if num_pedidos > 0 else 0
        
        row = 6 - i + 1
        ws_receita[f'A{row}'] = month_start.strftime('%B/%Y')
        ws_receita[f'B{row}'] = float(month_revenue)
        ws_receita[f'C{row}'] = num_pedidos
        ws_receita[f'D{row}'] = float(ticket_medio)
        
        for col in ['A', 'B', 'C', 'D']:
            ws_receita[f'{col}{row}'].border = border
            if col in ['B', 'C', 'D']:
                ws_receita[f'{col}{row}'].alignment = center_align
    
    # Formatar como moeda
    for row in range(2, 8):
        ws_receita[f'B{row}'].number_format = 'R$ #,##0.00'
        ws_receita[f'D{row}'].number_format = 'R$ #,##0.00'
    
    # Adicionar gráfico de barras
    chart = BarChart()
    chart.title = "Receita Mensal"
    chart.y_axis.title = "Receita (R$)"
    chart.x_axis.title = "Mês"
    
    data = Reference(ws_receita, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws_receita, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws_receita.add_chart(chart, "F2")
    
    # Ajustar larguras
    ws_receita.column_dimensions['A'].width = 15
    ws_receita.column_dimensions['B'].width = 15
    ws_receita.column_dimensions['C'].width = 12
    ws_receita.column_dimensions['D'].width = 15
    
    # ABA 4: Todos os Usuários
    ws_usuarios = wb.create_sheet("Usuários")
    
    # Cabeçalhos
    user_headers = ['ID', 'Nome', 'Username', 'Email', 'Tipo', 'Telefone', 'Data de Cadastro', 'Último Login', 'Ativo']
    for col, header in enumerate(user_headers, 1):
        cell = ws_usuarios.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos usuários
    users = User.objects.select_related('userprofile').order_by('-date_joined')
    
    for row_idx, user in enumerate(users, 2):
        try:
            profile = user.userprofile
            user_type = profile.get_user_type_display() if hasattr(profile, 'get_user_type_display') else profile.user_type
            phone = profile.phone if hasattr(profile, 'phone') else 'N/A'
        except:
            user_type = 'N/A'
            phone = 'N/A'
        
        data = [
            user.id,
            user.get_full_name() or 'N/A',
            user.username,
            user.email,
            user_type,
            phone,
            user.date_joined.strftime('%d/%m/%Y %H:%M'),
            user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'Nunca',
            'Sim' if user.is_active else 'Não'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_usuarios.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 7, 8, 9]:  # ID, Datas, Ativo
                cell.alignment = center_align
    
    # Ajustar larguras
    user_widths = [8, 25, 20, 30, 15, 15, 18, 18, 10]
    for col_idx, width in enumerate(user_widths, 1):
        ws_usuarios.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ABA 5: Todos os Serviços
    ws_servicos = wb.create_sheet("Serviços")
    
    # Cabeçalhos
    service_headers = ['ID', 'Nome', 'Categoria', 'Preço', 'Ativo', 'Data de Criação']
    for col, header in enumerate(service_headers, 1):
        cell = ws_servicos.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos serviços
    services = Service.objects.all().order_by('-created_at')
    
    for row_idx, service in enumerate(services, 2):
        data = [
            service.id,
            service.name,
            service.get_category_display() if hasattr(service, 'get_category_display') else service.category,
            float(service.price) if hasattr(service, 'price') else 0,
            'Sim' if service.is_active else 'Não',
            service.created_at.strftime('%d/%m/%Y') if hasattr(service, 'created_at') else 'N/A'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_servicos.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 4, 5, 6]:  # ID, Preço, Ativo, Data
                cell.alignment = center_align
            if col_idx == 4:  # Preço
                cell.number_format = 'R$ #,##0.00'
    
    # Ajustar larguras
    service_widths = [8, 30, 20, 15, 10, 15]
    for col_idx, width in enumerate(service_widths, 1):
        ws_servicos.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ABA 6: Todos os Pedidos
    ws_pedidos = wb.create_sheet("Pedidos")
    
    # Cabeçalhos
    order_headers = ['ID', 'Cliente', 'Serviço', 'Prestador', 'Status', 'Valor Total', 'Data do Pedido', 'Data de Conclusão']
    for col, header in enumerate(order_headers, 1):
        cell = ws_pedidos.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos pedidos
    orders = Order.objects.select_related('customer', 'service', 'professional').order_by('-created_at')[:200]
    
    for row_idx, order in enumerate(orders, 2):
        data = [
            order.id,
            order.customer.get_full_name() if order.customer else 'N/A',
            order.service.name if order.service else 'N/A',
            order.professional.get_full_name() if hasattr(order, 'professional') and order.professional else 'N/A',
            order.get_status_display(),
            float(order.total_price) if hasattr(order, 'total_price') else 0,
            order.created_at.strftime('%d/%m/%Y %H:%M'),
            order.completed_at.strftime('%d/%m/%Y %H:%M') if hasattr(order, 'completed_at') and order.completed_at else 'N/A'
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_pedidos.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 5, 6, 7, 8]:  # ID, Status, Valor, Datas
                cell.alignment = center_align
            if col_idx == 6:  # Valor
                cell.number_format = 'R$ #,##0.00'
    
    # Ajustar larguras
    order_widths = [8, 25, 25, 25, 15, 15, 18, 18]
    for col_idx, width in enumerate(order_widths, 1):
        ws_pedidos.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ABA 7: Prestadores
    ws_prestadores = wb.create_sheet("Prestadores")
    
    # Cabeçalhos
    provider_headers = ['ID', 'Nome', 'Email', 'Telefone', 'Serviços Oferecidos', 'Avaliação', 'Data de Cadastro']
    for col, header in enumerate(provider_headers, 1):
        cell = ws_prestadores.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos prestadores
    providers = UserProfile.objects.filter(user_type='professional').select_related('user').prefetch_related('user__custom_services')
    
    for row_idx, provider in enumerate(providers, 2):
        services_count = provider.user.custom_services.count() if hasattr(provider.user, 'custom_services') else 0
        
        data = [
            provider.user.id,
            provider.user.get_full_name() or provider.user.username,
            provider.user.email,
            provider.phone if hasattr(provider, 'phone') else 'N/A',
            services_count,
            f'{provider.rating:.1f}' if hasattr(provider, 'rating') and provider.rating else 'N/A',
            provider.user.date_joined.strftime('%d/%m/%Y')
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_prestadores.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 5, 6, 7]:  # ID, Serviços, Avaliação, Data
                cell.alignment = center_align
    
    # Ajustar larguras
    provider_widths = [8, 25, 30, 15, 18, 12, 15]
    for col_idx, width in enumerate(provider_widths, 1):
        ws_prestadores.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ABA 8: Clientes
    ws_clientes = wb.create_sheet("Clientes")
    
    # Cabeçalhos
    customer_headers = ['ID', 'Nome', 'Email', 'Telefone', 'Total de Pedidos', 'Total Gasto', 'Data de Cadastro']
    for col, header in enumerate(customer_headers, 1):
        cell = ws_clientes.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados dos clientes
    customers = UserProfile.objects.filter(user_type='customer').select_related('user').annotate(
        order_count=Count('user__order_set'),
        total_spent=Sum('user__order_set__total_price')
    )
    
    for row_idx, customer in enumerate(customers, 2):
        data = [
            customer.user.id,
            customer.user.get_full_name() or customer.user.username,
            customer.user.email,
            customer.phone if hasattr(customer, 'phone') else 'N/A',
            customer.order_count,
            float(customer.total_spent) if customer.total_spent else 0,
            customer.user.date_joined.strftime('%d/%m/%Y')
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws_clientes.cell(row_idx, col_idx, value)
            cell.border = border
            if col_idx in [1, 5, 6, 7]:  # ID, Pedidos, Gasto, Data
                cell.alignment = center_align
            if col_idx == 6:  # Total Gasto
                cell.number_format = 'R$ #,##0.00'
    
    # Ajustar larguras
    customer_widths = [8, 25, 30, 15, 15, 15, 15]
    for col_idx, width in enumerate(customer_widths, 1):
        ws_clientes.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Criar resposta
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_completo_admin_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
def update_request_status(request):
    """Update service request status (admin only)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
    
    # Check if user is admin
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Perfil não encontrado'}, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        request_id = data.get('request_id')
        new_status = data.get('status')
        
        if not request_id or not new_status:
            return JsonResponse({'success': False, 'error': 'Dados incompletos'}, status=400)
        
        # Update the request status
        service_request = ServiceRequestModal.objects.get(id=request_id)
        service_request.status = new_status
        service_request.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status atualizado para {new_status}',
            'new_status': new_status
        })
        
    except ServiceRequestModal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Solicitação não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def update_service(request):
    """Update custom service (admin only)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
    
    # Check if user is admin
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Perfil não encontrado'}, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        service_id = data.get('service_id')
        name = data.get('name')
        category = data.get('category')
        price = data.get('price')
        description = data.get('description')
        
        if not service_id or not name or not price:
            return JsonResponse({'success': False, 'error': 'Dados incompletos'}, status=400)
        
        # Update the service
        service = CustomService.objects.get(id=service_id)
        service.name = name
        service.category = category
        service.estimated_price = price
        service.description = description
        service.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Serviço atualizado com sucesso'
        })
        
    except CustomService.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Serviço não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_customer_requests(request, customer_id):
    """Get all service requests from a customer (admin only)"""
    # Check if user is admin
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'admin':
            return JsonResponse({'success': False, 'error': 'Acesso negado'}, status=403)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Perfil não encontrado'}, status=403)
    
    try:
        # Get customer
        customer = User.objects.get(id=customer_id)
        
        # Get all service requests from this customer
        requests = ServiceRequestModal.objects.filter(user=customer).order_by('-created_at')
        
        # Build requests list
        requests_data = []
        for req in requests:
            try:
                # Get service name
                service_name = req.service_name if req.service_name else 'Serviço Personalizado'
                if req.service:
                    service_name = req.service.name
                
                # Get description
                description = req.service_description if req.service_description else (req.description if hasattr(req, 'description') else 'Sem descrição')
                
                requests_data.append({
                    'id': req.id,
                    'service': service_name,
                    'status': req.status,
                    'created_at': req.created_at.strftime('%d/%m/%Y %H:%M'),
                    'description': description
                })
            except Exception as e:
                # Skip this request if there's an error
                print(f"Error processing request {req.id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'requests': requests_data,
            'total': len(requests_data)
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente não encontrado'}, status=404)
    except Exception as e:
        import traceback
        print(f"Error in get_customer_requests: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def admin_settings(request):
    """Admin settings page"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado - perfil de usuário não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Handle form submission
    if request.method == 'POST':
        # Process settings updates here
        messages.success(request, 'Configurações atualizadas com sucesso!')
        return redirect('admin_settings')
    
    # Get current settings
    context = {
        'system_settings': {
            'maintenance_mode': False,
            'user_registration': True,
            'email_verification': True,
        },
        'notification_settings': {
            'email_notifications': True,
            'push_notifications': True,
            'security_alerts': True,
        },
        'appearance_settings': {
            'default_theme': 'dark',
            'default_language': 'pt-br',
        }
    }
    
    return render(request, 'services/admin_settings.html', context)


def about(request):
    """About page"""
    return render(request, 'services/about.html')


def help_support(request):
    """Help and Support page"""
    return render(request, 'services/help_support.html')

def contact(request):
    """Redirect contact page to support system"""
    messages.info(request, 'Agora usamos um sistema de suporte mais completo! Crie um ticket para entrar em contato.')
    return redirect('create_support_ticket')


def faq(request):
    """FAQ page"""
    return render(request, 'services/faq.html')


def information(request):
    """Information hub page"""
    return render(request, 'services/information.html')


def terms(request):
    """Terms of service page"""
    return render(request, 'services/terms.html')


def robots_txt(request):
    """
    Serve o arquivo robots.txt
    """
    return render(request, 'robots.txt', content_type='text/plain')

def sitemap_xml(request):
    """
    Serve o arquivo sitemap.xml
    """
    return render(request, 'sitemap.xml', content_type='application/xml')


@login_required
def admin_provider_services(request, provider_id):
    """Admin view to manage a specific provider's services"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado - perfil de usuário não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Get the provider
    try:
        provider = User.objects.get(id=provider_id)
        provider_profile = UserProfile.objects.get(user=provider)
    except User.DoesNotExist:
        messages.error(request, 'Prestador não encontrado')
        return redirect('admin_providers_list')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil do prestador não encontrado')
        return redirect('admin_providers_list')
    
    # Get provider's custom services
    custom_services = CustomService.objects.filter(provider=provider)
    
    # Filter active services
    active_services = custom_services.filter(is_active=True)
    
    return render(request, 'services/admin_provider_services.html', {
        'provider': provider,
        'provider_profile': provider_profile,
        'custom_services': custom_services,
        'active_services': active_services
    })


@login_required
def admin_providers_list(request):
    """Admin view to list all providers"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado - perfil de usuário não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Get all providers
    providers = User.objects.filter(userprofile__user_type='professional').select_related('userprofile')
    
    return render(request, 'services/admin_providers_list.html', {
        'providers': providers
    })


@login_required
def admin_settings(request):
    """Admin settings page"""
    # Check if user is admin
    if not request.user.is_authenticated:
        messages.error(request, 'Você precisa estar logado para acessar esta página')
        return redirect('login')
        
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'admin':
            messages.error(request, 'Acesso negado - você não tem permissão para acessar esta página')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado - perfil de usuário não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro desconhecido: {str(e)}')
        return redirect('home')
    
    # Handle form submission
    if request.method == 'POST':
        # Process settings updates here
        messages.success(request, 'Configurações atualizadas com sucesso!')
        return redirect('admin_settings')
    
    # Get current settings
    context = {
        'system_settings': {
            'maintenance_mode': False,
            'user_registration': True,
            'email_verification': True,
        },
        'notification_settings': {
            'email_notifications': True,
            'push_notifications': True,
            'security_alerts': True,
        },
        'appearance_settings': {
            'default_theme': 'dark',
            'default_language': 'pt-br',
        }
    }
    
    return render(request, 'services/admin_settings.html', context)


def update_user_location(request):
    """Update user's location coordinates"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        
        # Validate coordinates
        try:
            lat = float(latitude)
            lng = float(longitude)
            
            # Validate ranges
            if lat < -90 or lat > 90:
                return JsonResponse({'error': 'Invalid latitude value'}, status=400)
            
            if lng < -180 or lng > 180:
                return JsonResponse({'error': 'Invalid longitude value'}, status=400)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid coordinates'}, status=400)
        
        # Update user profile
        try:
            user_profile = request.user.userprofile
            user_profile.latitude = lat
            user_profile.longitude = lng
            user_profile.save(update_fields=['latitude', 'longitude'])
            
            return JsonResponse({
                'success': True,
                'message': 'Location updated successfully'
            })
        except Exception as e:
            return JsonResponse({'error': 'Failed to update location'}, status=500)
        
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


def get_nearby_professionals(request):
    """Get nearby professionals based on user's location"""
    try:
        # Try to get user's location
        user_lat = None
        user_lng = None
        user_location_set = False
        
        try:
            user_profile = request.user.userprofile
            if user_profile.latitude and user_profile.longitude:
                user_lat = float(user_profile.latitude)
                user_lng = float(user_profile.longitude)
                user_location_set = True
        except:
            pass
        
        # Get parameters
        service_category = request.GET.get('category', '')
        max_distance = float(request.GET.get('max_distance', 50))  # Default 50km
        
        # Get all professionals
        professionals = UserProfile.objects.filter(user_type='professional')
        
        # Filter by category if provided
        if service_category:
            professionals = professionals.filter(
                user__custom_services__category=service_category
            ).distinct()
        
        # If user location is not set, return all professionals with a message
        if not user_location_set:
            all_professionals = []
            for prof in professionals:
                # Get professional's services
                services = prof.user.custom_services.filter(is_active=True)
                
                all_professionals.append({
                    'id': prof.user.id,
                    'username': prof.user.username,
                    'name': prof.user.get_full_name() or prof.user.username,
                    'rating': float(prof.rating),
                    'review_count': prof.review_count,
                    'latitude': float(prof.latitude) if prof.latitude else None,
                    'longitude': float(prof.longitude) if prof.longitude else None,
                    'distance': None,  # No distance calculation without user location
                    'services': [
                        {
                            'id': service.id,
                            'name': service.name,
                            'category': service.get_category_display(),
                            'price': float(service.estimated_price) if service.estimated_price else 0.0
                        }
                        for service in services
                    ]
                })
            
            return JsonResponse({
                'success': True,
                'professionals': all_professionals,
                'count': len(all_professionals),
                'message': 'Localização do usuário não definida. Mostrando todos os profissionais.'
            })
        
        # Calculate distances and filter by proximity
        nearby_professionals = []
        for prof in professionals:
            if prof.latitude and prof.longitude:
                prof_lat = float(prof.latitude)
                prof_lng = float(prof.longitude)
                
                # Calculate distance using haversine formula
                distance = calculate_distance(user_lat, user_lng, prof_lat, prof_lng)
                
                if distance <= max_distance:
                    # Get professional's services
                    services = prof.user.custom_services.filter(is_active=True)
                    
                    nearby_professionals.append({
                        'id': prof.user.id,
                        'username': prof.user.username,
                        'name': prof.user.get_full_name() or prof.user.username,
                        'rating': float(prof.rating),
                        'review_count': prof.review_count,
                        'latitude': float(prof.latitude),
                        'longitude': float(prof.longitude),
                        'distance': round(distance, 2),
                        'services': [
                            {
                                'id': service.id,
                                'name': service.name,
                                'category': service.get_category_display(),
                                'price': float(service.estimated_price) if service.estimated_price else 0.0
                            }
                            for service in services
                        ]
                    })
        
        # Sort by distance
        nearby_professionals.sort(key=lambda x: x['distance'] if x['distance'] is not None else float('inf'))
        
        return JsonResponse({
            'success': True,
            'professionals': nearby_professionals,
            'count': len(nearby_professionals)
        })
        
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two points using haversine formula"""
    from math import radians, cos, sin, asin, sqrt
    
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    
    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of earth in kilometers
    return c * r


@login_required
def map_view(request):
    """Display map with nearby professionals"""
    return render(request, 'services/map_view.html')


@login_required
def submit_review(request, order_id):
    """Submit a review for a completed order"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the order and ensure it belongs to the current user and is completed
        order = Order.objects.get(id=order_id, customer=request.user, status='completed')
        
        # Check if review already exists
        if hasattr(order, 'review'):
            return JsonResponse({'error': 'Review already submitted for this order'}, status=400)
        
        # Get review data
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '').strip()
        
        # Validate rating
        try:
            rating = int(rating)
            if rating < 1 or rating > 5:
                raise ValueError()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid rating value. Must be between 1 and 5.'}, status=400)
        
        # Create the review
        review = Review.objects.create(
            order=order,
            customer=request.user,
            professional=order.professional,
            rating=rating,
            comment=comment,
            is_verified=True  # In a real app, this would be verified through some process
        )
        
        # Update professional's rating and review count
        try:
            professional_profile = UserProfile.objects.get(user=order.professional)
            # Calculate new average rating
            total_reviews = professional_profile.review_count
            current_rating = float(professional_profile.rating)
            
            # New average = (current_rating * total_reviews + new_rating) / (total_reviews + 1)
            new_rating = (current_rating * total_reviews + rating) / (total_reviews + 1)
            
            professional_profile.rating = new_rating
            professional_profile.review_count = total_reviews + 1
            professional_profile.save(update_fields=['rating', 'review_count'])
        except UserProfile.DoesNotExist:
            # If professional profile doesn't exist, create one
            UserProfile.objects.create(
                user=order.professional,
                user_type='professional',
                rating=rating,
                review_count=1
            )
        except Exception:
            # If there's any error updating the profile, continue without failing
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Review submitted successfully',
            'rating': review.rating,
            'comment': review.comment,
            'created_at': review.created_at.isoformat()
        })
        
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found or not eligible for review'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


def get_professional_reviews(request, professional_id):
    """Get all reviews for a professional"""
    try:
        professional = User.objects.get(id=professional_id)
        reviews = Review.objects.filter(professional=professional).select_related('customer', 'order').order_by('-created_at')
        
        # Prepare reviews data
        reviews_data = []
        for review in reviews:
            reviews_data.append({
                'id': review.id,
                'rating': review.rating,
                'comment': review.comment,
                'created_at': review.created_at.isoformat(),
                'customer': {
                    'username': review.customer.username,
                    'first_name': review.customer.first_name,
                    'last_name': review.customer.last_name
                },
                'order': {
                    'id': review.order.id,
                    'service_name': review.order.service.name if review.order.service else 'Custom Service'
                }
            })
        
        # Get professional profile for rating info
        try:
            profile = UserProfile.objects.get(user=professional)
            professional_rating = float(profile.rating)
            review_count = profile.review_count
        except UserProfile.DoesNotExist:
            professional_rating = 0
            review_count = 0
        
        return JsonResponse({
            'success': True,
            'reviews': reviews_data,
            'professional_rating': professional_rating,
            'review_count': review_count
        })
    except User.DoesNotExist:
        return JsonResponse({'error': 'Professional not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)



def privacy(request):
    """Privacy policy page"""
    return render(request, 'services/privacy.html')


def schedule_service(request):
    """Service scheduling page"""
    # Check if user is authenticated
    if not request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Você precisa estar logado para agendar um serviço'})
        messages.error(request, 'Você precisa estar logado para agendar um serviço')
        return redirect('login')
    
    # Get service details from query parameters or session
    service_id = request.GET.get('service_id')
    service_name = request.GET.get('service_name', 'Serviço Personalizado')
    service_description = request.GET.get('service_description', 'Descrição do serviço')
    service_price = request.GET.get('service_price', '0.00')
    service_category = request.GET.get('service_category', 'Outros')
    
    # Check if continuing from a pending order
    order_id = request.GET.get('order_id')
    existing_order = None
    if order_id:
        try:
            existing_order = Order.objects.get(id=order_id, customer=request.user, status='pending')
            # Pre-fill form data from existing order
            service_name = existing_order.service.name if existing_order.service else 'Serviço Personalizado'
            service_description = existing_order.service.description if existing_order.service else 'Serviço solicitado diretamente ao profissional'
            service_price = str(existing_order.total_price)
        except Order.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Pedido não encontrado ou não está pendente'})
            messages.error(request, 'Pedido não encontrado ou não está pendente')
            return redirect('service_history_new')
    
    # If it's a POST request, process the form
    if request.method == 'POST':
        # Process the scheduling form
        scheduled_datetime = request.POST.get('scheduled-datetime')
        time_preference = request.POST.get('time-preference')
        urgency = request.POST.get('urgency')
        zip_code = request.POST.get('zip-code')
        address = request.POST.get('address')
        number = request.POST.get('number')
        complement = request.POST.get('complement')
        city = request.POST.get('city')
        state = request.POST.get('state')
        problem_description = request.POST.get('problem-description')
        additional_notes = request.POST.get('additional-notes')
        accessibility_needs = request.POST.get('accessibility-needs')
        insurance_required = request.POST.get('insurance-required')
        terms_accepted = request.POST.get('terms-accepted')
        
        # Validate required fields
        if not all([scheduled_datetime, address, number, city, state, terms_accepted]):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Por favor, preencha todos os campos obrigatórios'})
            messages.error(request, 'Por favor, preencha todos os campos obrigatórios')
            return render(request, 'services/schedule_service.html', {
                'service_id': service_id,
                'service_name': service_name,
                'service_description': service_description,
                'service_price': service_price,
                'service_category': service_category,
                'existing_order': existing_order
            })
        
        # In a real application, you would:
        # 1. Save the scheduling information to the database
        # 2. Associate it with the logged-in user
        # 3. Send confirmation email
        # 4. Notify the service provider
        
        # For now, we'll just add a success message
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True, 
                'message': 'Agendamento realizado com sucesso! Você receberá uma confirmação por e-mail.',
                'redirect_url': reverse('service_history_new')
            })
        
        messages.success(request, 'Agendamento realizado com sucesso! Você receberá uma confirmação por e-mail.')
        
        # Redirect to service history page
        return redirect('service_history_new')
    
    # For GET requests, show the scheduling form
    return render(request, 'services/schedule_service.html', {
        'service_id': service_id,
        'service_name': service_name,
        'service_description': service_description,
        'service_price': service_price,
        'service_category': service_category,
        'existing_order': existing_order
    })


def custom_404(request, exception):
    """Custom 404 error page"""
    return render(request, '404.html', status=404)

def custom_500(request):
    """Custom 500 error page"""
    return render(request, '500.html', status=500)

def template_not_found(request, exception=None):
    """Custom template not found error page"""
    return render(request, 'template_not_found.html', status=404)

def test_template_not_found(request):
    """Test view to trigger template not found error"""
    from django.template import TemplateDoesNotExist
    raise TemplateDoesNotExist("test_template.html")


def report_error(request):
    """Handle error reports from users"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        # Extract form data
        error_type = data.get('error_type', 'Unknown')
        error_message = data.get('error_message', 'No message provided')
        user_description = data.get('user_description', '')
        user_email = data.get('user_email', '')
        error_url = data.get('error_url', '')
        user_agent = data.get('user_agent', '')
        timestamp = data.get('timestamp', '')
        
        # Create a notification for admin users
        from django.contrib.auth.models import User
        from .models import Notification
        
        # Get all admin users
        admin_users = User.objects.filter(is_superuser=True)
        
        if admin_users.exists():
            # Create a notification for each admin
            for admin in admin_users:
                Notification.objects.create(
                    user=admin,
                    sender=None,  # System notification
                    notification_type='system',
                    title=f'Error Report: {error_type}',
                    message=f"""
User Report: {user_description}
Error Type: {error_type}
Error Message: {error_message}
User Email: {user_email}
Error URL: {error_url}
User Agent: {user_agent}
Timestamp: {timestamp}
                    """.strip(),
                    is_read=False
                )
            
            return JsonResponse({'success': True, 'message': 'Error report submitted successfully'})
        else:
            # If no admin users exist, log to console
            import logging
            logger = logging.getLogger('services')
            logger.error(f"""
User Report: {user_description}
Error Type: {error_type}
Error Message: {error_message}
User Email: {user_email}
Error URL: {error_url}
User Agent: {user_agent}
Timestamp: {timestamp}
            """.strip())
            
            return JsonResponse({'success': True, 'message': 'Error report submitted successfully'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing report: {str(e)}'})


def custom_403(request, exception):
    """Custom 403 error page"""
    return render(request, '403.html', status=403)


def custom_400(request, exception):
    """Custom 400 error page"""
    return render(request, '400.html', status=400)


def test_500_error(request):
    """Test view to trigger 500 error"""
    # This will cause a division by zero error
    return 1 / 0


def test_404_error(request):
    """Test view to trigger 404 error"""
    from django.http import Http404
    raise Http404("Página de teste 404")


def comprehensive_error_handler(request, error_type, error_message, status_code):
    """Comprehensive error handler for all types of errors"""
    # Map error types to template names
    template_map = {
        400: '400.html',
        403: '403.html',
        404: '404.html',
        500: '500.html',
    }
    
    # Get the appropriate template or default to 500
    template_name = template_map.get(status_code, '500.html')
    
    # Render the error page
    return render(request, template_name, {
        'error_type': error_type,
        'error_message': error_message,
    }, status=status_code)


def test_template(request):
    """Test view to check if template loading works"""
    try:
        from django.template.loader import get_template
        template = get_template('registration/login.html')
        return HttpResponse(f"Template found: {template}")
    except Exception as e:
        return HttpResponse(f"Error loading template: {e}")




def test_service_display(request, custom_service_id):
    """Test view to debug service display issues"""
    try:
        custom_service = CustomService.objects.select_related('provider').get(id=custom_service_id)
        print(f"DEBUG: Found service {custom_service.name} with ID {custom_service.id}")
        print(f"DEBUG: Provider username: {custom_service.provider.username}")
        print(f"DEBUG: Provider full name: {custom_service.provider.get_full_name()}")
        print(f"DEBUG: Service price: {custom_service.estimated_price}")
        return render(request, 'services/test_service.html', {
            'custom_service': custom_service
        })
    except CustomService.DoesNotExist:
        return render(request, 'services/test_service.html', {
            'error': 'Service not found'
        })
    except Exception as e:
        return render(request, 'services/test_service.html', {
            'error': f'Error: {str(e)}'
        })


@login_required
def service_detail(request, order_id):
    """View a service detail"""
    try:
        order = Order.objects.get(
            id=order_id, customer=request.user
        )
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado')
        return redirect('service_history_new')
    
    return render(request, 'services/service_detail.html', {
        'order': order
    })


@login_required
def cancel_order(request, order_id):
    """Cancel an order"""
    print(f"Cancel order called with order_id: {order_id}")  # Debug print
    try:
        order = Order.objects.get(id=order_id, customer=request.user)
        
        # Only allow cancellation of pending orders
        if order.status != 'pending':
            messages.error(request, 'Este pedido não pode ser cancelado.')
            return redirect('service_history_new')
        
        # Update order status
        order.status = 'cancelled'
        order.save()
        
        messages.success(request, 'Pedido cancelado com sucesso!')
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado.')
    
    return redirect('service_history_new')


@login_required
def submit_service_request(request):
    """Handle service request submission via AJAX for existing services"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        # Parse JSON data
        import json
        data = json.loads(request.body)
        
        # Extract form data
        service_id = data.get('service_id')
        custom_service_id = data.get('custom_service_id')
        scheduled_date = data.get('scheduled_date')
        address = data.get('address')
        notes = data.get('notes')
        
        # Validate required fields
        if not scheduled_date or not address:
            return JsonResponse({'success': False, 'message': 'Por favor, preencha todos os campos obrigatórios.'})
        
        # Get the service (either standard or custom)
        service = None
        custom_service = None
        provider = None
        
        if service_id:
            try:
                service = Service.objects.get(id=service_id)
            except Service.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Serviço não encontrado.'})
        elif custom_service_id:
            try:
                custom_service = CustomService.objects.get(id=custom_service_id)
                provider = custom_service.provider
            except CustomService.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Serviço personalizado não encontrado.'})
        else:
            return JsonResponse({'success': False, 'message': 'Nenhum serviço especificado.'})
        
        # Create service request (using the existing Order model)
        order = Order.objects.create(
            customer=request.user,
            service=service,
            professional=provider,
            scheduled_date=scheduled_date,
            address=address,
            notes=notes,
            total_price=custom_service.estimated_price if custom_service else service.base_price,
            # Service details
            service_name=custom_service.name if custom_service else service.name,
            service_description=custom_service.description if custom_service else service.description,
            service_category=custom_service.category if custom_service else service.category,
            status='pending'
        )
        
        # Create notification for provider if it's a custom service
        if provider:
            # Create notification for provider
            Notification.objects.create(
                user=provider,
                sender=request.user,
                notification_type='service_request',
                title='Nova Solicitação de Serviço',
                message=f'Nova solicitação de serviço: {custom_service.name if custom_service else service.name} por {request.user.username}',
                related_object_id=order.id,
                related_object_type='Order'
            )
        
        # Create notification for admin
        admin_user = User.objects.filter(is_superuser=True).first()
        if admin_user:
            Notification.objects.create(
                user=admin_user,
                sender=request.user,
                notification_type='service_request',
                title='Nova Solicitação de Serviço',
                message=f'Nova solicitação de serviço: {custom_service.name if custom_service else service.name} por {request.user.username}',
                related_object_id=order.id,
                related_object_type='Order'
            )
        
        return JsonResponse({
            'success': True, 
            'message': 'Solicitação enviada com sucesso!',
            'order_id': order.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao processar solicitação: {str(e)}'})


@login_required
def provider_service_requests(request):
    """Show service requests for providers"""
    # Check if user is a provider
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Perfil não encontrado'}, status=404)
    
    # Get service requests assigned to this provider
    service_requests = ServiceRequest.objects.filter(
        provider=request.user,
        status__in=['pending', 'accepted']
    ).order_by('-created_at')
    
    # Format data for JSON response
    requests_data = []
    for req in service_requests:
        requests_data.append({
            'id': req.id,
            'title': req.title,
            'description': req.description,
            'category': req.get_category_display(),
            'scheduled_date': req.scheduled_date.isoformat(),
            'address': req.address,
            'notes': req.notes,
            'status': req.get_status_display(),
            'customer_name': req.customer.get_full_name() or req.customer.username,
            'created_at': req.created_at.isoformat()
        })
    
    return JsonResponse({
        'success': True,
        'service_requests': requests_data
    })


@login_required
def update_service_request_status(request, request_id):
    """Update service request status (accept/reject)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        # Check if user is a provider
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
        
        # Get the service request
        service_request = ServiceRequest.objects.get(id=request_id)
        
        # Check if this provider is assigned to the request
        if service_request.provider != request.user:
            return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)
        
        # Get action from request
        import json
        data = json.loads(request.body)
        action = data.get('action')  # 'accept' or 'reject'
        rejection_reason = data.get('rejection_reason', '')
        
        # Update status
        if action == 'accept':
            service_request.status = 'accepted'
            message = 'Solicitação aceita com sucesso!'
        elif action == 'reject':
            service_request.status = 'rejected'
            service_request.rejection_reason = rejection_reason
            message = 'Solicitação rejeitada.'
        else:
            return JsonResponse({'success': False, 'message': 'Ação inválida'})
        
        service_request.save()
        
        # Create notification for customer
        Notification.objects.create(
            user=service_request.customer,
            sender=request.user,
            notification_type=f'service_{action}',
            title=f'Solicitação de Serviço {action.capitalize()}',
            message=f'Sua solicitação "{service_request.title}" foi {action}a pelo prestador {request.user.username}.',
            related_object_id=service_request.id,
            related_object_type='ServiceRequest'
        )
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'status': service_request.get_status_display()
        })
        
    except ServiceRequest.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Solicitação não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao processar solicitação: {str(e)}'})


@login_required
def notification_demo(request):
    return render(request, 'services/notification_demo.html')


@login_required
def provider_service_requests_page(request):
    """Page to show service requests for providers"""
    # Check if user is a provider
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil não encontrado')
        return redirect('home')
    
    return render(request, 'services/provider_service_requests.html')


def notifications_view(request):
    """Display user notifications"""
    # For now, we'll just render a simple template
    # This can be expanded later to show actual notifications
    return render(request, 'services/notifications.html', {
        'notifications': []
    })


def simple_test(request, custom_service_id):
    """Very simple test view"""
    try:
        from services.models import CustomService
        custom_service = CustomService.objects.get(id=custom_service_id)
        return render(request, 'services/simple_test.html', {
            'custom_service': custom_service
        })
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")





# Service Request Views

@login_required
def request_service(request, service_id):
    """Request a service - disabled"""
    messages.info(request, 'Funcionalidade de solicitação de serviço temporariamente desabilitada.')
    return redirect('search_new')
    
    if request.method == 'POST':
        try:
            # Get the service
            service = Service.objects.get(id=service_id)
            
            # Process the form data
            cep = request.POST.get('cep', '')
            address = request.POST.get('address', '')
            number = request.POST.get('number', '')
            complement = request.POST.get('complement', '')
            city = request.POST.get('city', '')
            state = request.POST.get('state', '')
            date = request.POST.get('date', '')
            time = request.POST.get('time', '')
            urgency = request.POST.get('urgency', 'normal')
            description = request.POST.get('description', '')
            notes = request.POST.get('notes', '')
            payment_method_id = request.POST.get('payment_method', '')
            
            # Combine address components
            full_address = f"{address}, {number}"
            if complement:
                full_address += f", {complement}"
            full_address += f", {city} - {state}, {cep}"
            
            # Combine date and time
            scheduled_datetime = f"{date} {time}"
            
            # Calculate total price based on urgency
            total_price = service.base_price
            if urgency == 'urgent':
                total_price *= Decimal('1.3')  # 30% increase
            elif urgency == 'emergency':
                total_price *= Decimal('1.5')  # 50% increase
            
            # Create the order
            order = Order.objects.create(
                customer=request.user,
                service=service,
                scheduled_date=scheduled_datetime,
                address=full_address,
                notes=notes,
                total_price=total_price,
                # Service details
                service_name=service.name,
                service_description=service.description,
                service_category=service.category
            )
            
            # If a payment method was selected, associate it with the order
            if payment_method_id:
                try:
                    payment_method = PaymentMethod.objects.get(id=payment_method_id, user=request.user)
                    order.payment_method = payment_method
                    order.save()
                except PaymentMethod.DoesNotExist:
                    pass  # Handle error case
            
            # Save the payment method to user profile if it's a new one
            payment_type = request.POST.get('payment_type', '')
            card_number_last4 = request.POST.get('card_number_last4', '')
            cardholder_name = request.POST.get('cardholder_name', '')
            expiry_date = request.POST.get('expiry_date', '')
            set_as_default = request.POST.get('set_as_default', False)
            
            if payment_type and card_number_last4:
                # Create new payment method
                payment_method = PaymentMethod.objects.create(
                    user=request.user,
                    payment_type=payment_type,
                    card_number_last4=card_number_last4,
                    cardholder_name=cardholder_name,
                    expiry_date=expiry_date if expiry_date else None,
                    is_default=set_as_default
                )
                
                # Associate with order
                order.payment_method = payment_method
                order.save()
                
                # If set as default, update other payment methods
                if set_as_default:
                    PaymentMethod.objects.filter(user=request.user, is_default=True).exclude(id=payment_method.id).update(is_default=False)
            
            messages.success(request, 'Solicitação de serviço enviada com sucesso!')
            
            # Return JSON response for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Solicitação de serviço enviada com sucesso!',
                    'redirect_url': reverse('order_confirmation', args=[order.id])
                })
            
            # Redirect to confirmation page
            return redirect('order_confirmation', order_id=order.id)
            
        except Service.DoesNotExist:
            messages.error(request, 'Serviço não encontrado')
            return redirect('home')
        except Exception as e:
            messages.error(request, f'Erro ao processar solicitação: {str(e)}')
            return redirect('home')


    
@login_required
def order_payment(request, order_id):
    """Handle order payment - multi-step process"""
    try:
        order = Order.objects.get(id=order_id, user=request.user)
        payment_methods = get_payment_methods()
        if request.method == 'POST':
            # Process the payment details and update the order status
            payment_data = {
                'amount': order.total_price,
                'method': request.POST.get('payment_method', '')
            }
            update_order_status(order, payment_data)
            return render(request, 'services/order_confirmation.html', {
                'order': order,
                'redirect_url': reverse('order_confirmation', args=[order.id])
            })
            
            # Redirect to confirmation page
            return redirect('order_confirmation', order_id=order.id)
            
        return render(request, 'services/order_payment.html', {
            'order': order,
            'payment_methods': payment_methods
        })
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro ao processar pagamento: {str(e)}')
        return redirect('home')


@login_required
def request_custom_service_view(request):
    """Request a custom service - handles the multi-step modal process"""
    try:
        if request.method == 'POST':
            # Process the form data
            service_name = request.POST.get('service_name', '')
            cep = request.POST.get('cep', '')
            address = request.POST.get('address', '')
            number = request.POST.get('number', '')
            complement = request.POST.get('complement', '')
            city = request.POST.get('city', '')
            state = request.POST.get('state', '')
            date = request.POST.get('date', '')
            time = request.POST.get('time', '')
            urgency = request.POST.get('urgency', 'normal')
            description = request.POST.get('description', '')
            notes = request.POST.get('notes', '')
            payment_method_id = request.POST.get('payment_method', '')
            
            # Combine address components
            full_address = f"{address}, {number}"
            if complement:
                full_address += f", {complement}"
            full_address += f", {city} - {state}, {cep}"
            
            # Combine date and time
            scheduled_datetime = f"{date} {time}"
            
            # For custom services, we'll use a default price
            total_price = Decimal('100.00')  # Default price
            if urgency == 'urgent':
                total_price *= Decimal('1.3')  # 30% increase
            elif urgency == 'emergency':
                total_price *= Decimal('1.5')  # 50% increase
            
            # Create a custom service
            custom_service = CustomService.objects.create(
                name=service_name,
                description=description,
                category='other',  # Default category
                estimated_price=total_price,
                provider=None  # Will be assigned later
            )
            
            # Create the order
            order = Order.objects.create(
                customer=request.user,
                service=None,  # No standard service
                scheduled_date=scheduled_datetime,
                address=full_address,
                notes=notes,
                total_price=total_price,
                # Service details
                service_name=custom_service.name,
                service_description=custom_service.description,
                service_category=custom_service.category
            )
            
            # If a payment method was selected, associate it with the order
            if payment_method_id:
                try:
                    payment_method = PaymentMethod.objects.get(id=payment_method_id, user=request.user)
                    order.payment_method = payment_method
                    order.save()
                except PaymentMethod.DoesNotExist:
                    pass  # Handle error case
            
            # Save the payment method to user profile if it's a new one
            payment_type = request.POST.get('payment_type', '')
            card_number_last4 = request.POST.get('card_number_last4', '')
            cardholder_name = request.POST.get('cardholder_name', '')
            expiry_date = request.POST.get('expiry_date', '')
            set_as_default = request.POST.get('set_as_default', False)
            
            if payment_type and card_number_last4:
                # Create new payment method
                payment_method = PaymentMethod.objects.create(
                    user=request.user,
                    payment_type=payment_type,
                    card_number_last4=card_number_last4,
                    cardholder_name=cardholder_name,
                    expiry_date=expiry_date if expiry_date else None,
                    is_default=set_as_default
                )
                
                # Associate with order
                order.payment_method = payment_method
                order.save()
                
                # If set as default, update other payment methods
                if set_as_default:
                    PaymentMethod.objects.filter(user=request.user, is_default=True).exclude(id=payment_method.id).update(is_default=False)
            
            messages.success(request, 'Solicitação de serviço personalizado enviada com sucesso!')
            
            # Return JSON response for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Solicitação de serviço personalizado enviada com sucesso!',
                    'redirect_url': reverse('order_confirmation', args=[order.id])
                })
            
            # Redirect to confirmation page
            return redirect('order_confirmation', order_id=order.id)
    except Exception as e:
        messages.error(request, f'Erro ao processar solicitação: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': f'Erro ao processar solicitação: {str(e)}'
        }) if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('home')


@login_required
def schedule_service(request):
    """Schedule a service"""
    if request.method == 'POST':
        # Process scheduling form
        service_id = request.POST.get('service_id')
        scheduled_date = request.POST.get('scheduled_date')
        
        try:
            service = Service.objects.get(id=service_id)
            
            # Create order
            order = Order.objects.create(
                customer=request.user,
                service=service,
                scheduled_date=scheduled_date,
                total_price=service.base_price
            )
            
            messages.success(request, 'Serviço agendado com sucesso!')
            return redirect('order_confirmation', order_id=order.id)
        except Service.DoesNotExist:
            messages.error(request, 'Serviço não encontrado')
            return redirect('home')
        except Exception as e:
            messages.error(request, f'Erro ao agendar serviço: {str(e)}')
            return redirect('home')
    
    return redirect('home')

@login_required
def order_confirmation(request, order_id):
    """Display order confirmation"""
    try:
        order = Order.objects.get(id=order_id, customer=request.user)
        return render(request, 'services/order_confirmation.html', {
            'order': order
        })
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro ao carregar confirmação: {str(e)}')
        return redirect('home')

@login_required
def order_payment(request, order_id):
    """Process order payment"""
    try:
        order = Order.objects.get(id=order_id, customer=request.user)
        
        if request.method == 'POST':
            # Process payment
            payment_method_id = request.POST.get('payment_method')
            
            if payment_method_id:
                try:
                    payment_method = PaymentMethod.objects.get(id=payment_method_id, user=request.user)
                    order.payment_method = payment_method
                    order.save()
                    
                    messages.success(request, 'Pagamento processado com sucesso!')
                    return redirect('order_confirmation', order_id=order.id)
                except PaymentMethod.DoesNotExist:
                    messages.error(request, 'Método de pagamento não encontrado')
                    return redirect('home')
            else:
                messages.error(request, 'Por favor, selecione um método de pagamento')
                return redirect('home')
        
        # Get user's payment methods
        payment_methods = PaymentMethod.objects.filter(user=request.user)
        
        return render(request, 'services/order_payment.html', {
            'order': order,
            'payment_methods': payment_methods
        })
    except Order.DoesNotExist:
        messages.error(request, 'Pedido não encontrado')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Erro ao processar pagamento: {str(e)}')
        return redirect('home')

@login_required
def profile_view(request):
    """User profile page with complete profile information, order history, and settings"""
    try:
        # Get or create user profile
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Get user's orders
        user_orders = Order.objects.filter(customer=request.user).order_by('-created_at')
        
        # Get user's payment methods
        payment_methods = PaymentMethod.objects.filter(user=request.user)
        
        # Handle POST requests for profile updates
        if request.method == 'POST':
            # Update profile information
            user_profile.phone = request.POST.get('phone', user_profile.phone)
            user_profile.address = request.POST.get('address', user_profile.address)
            user_profile.number = request.POST.get('number', user_profile.number)
            user_profile.complement = request.POST.get('complement', user_profile.complement)
            user_profile.city = request.POST.get('city', user_profile.city)
            user_profile.state = request.POST.get('state', user_profile.state)
            user_profile.zip_code = request.POST.get('zip_code', user_profile.zip_code)
            user_profile.birth_date = request.POST.get('birth_date', user_profile.birth_date)
            
            # Handle avatar upload
            if 'avatar' in request.FILES:
                user_profile.avatar = request.FILES['avatar']
            
            user_profile.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('profile')
        
        context = {
            'user_profile': user_profile,
            'user_orders': user_orders,
            'payment_methods': payment_methods,
        }
        return render(request, 'services/profile.html', context)
    except Exception as e:
        messages.error(request, f'Erro ao carregar o perfil: {str(e)}')
        return redirect('home')




def chat_ai_response(request):
    """Process AI chat messages"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            user_message = data.get('message', '').lower().strip()
            
            # AI Knowledge Base (same as in JavaScript but server-side)
            responses = {
                "como criar uma conta": {
                    "answer": "Para criar uma conta no Job Finder:\n\n1. Clique em 'Cadastrar' no menu superior\n2. Preencha seus dados pessoais\n3. Escolha se é cliente ou profissional\n4. Confirme seu e-mail\n5. Faça login e comece a usar!\n\nPrecisa de mais alguma coisa?",
                    "followUp": ["Como solicitar serviço", "Como ser profissional", "Problemas com login"]
                },
                "como solicitar um serviço": {
                    "answer": "Para solicitar um serviço:\n\n1. Acesse 'Buscar Profissionais' no menu\n2. Escolha a categoria do serviço\n3. Selecione um profissional\n4. Clique em 'Solicitar Serviço'\n5. Preencha os detalhes (data, endereço, etc.)\n6. Confirme e realize o pagamento\n\nO profissional entrará em contato em breve!",
                    "followUp": ["Métodos de pagamento", "Como cancelar serviço", "Problemas com agendamento"]
                },
                "problemas com pagamento": {
                    "answer": "Para problemas com pagamento:\n\n• Cartão recusado: Verifique dados e limite\n• PIX não processado: Aguarde até 2 horas\n• Reembolso: Solicitação em até 7 dias úteis\n• Dúvidas sobre cobrança: Entre em contato\n\nTodos os pagamentos são seguros via Stripe. Posso ajudar com algo específico?",
                    "followUp": ["Solicitar reembolso", "Alterar forma de pagamento", "Falar com humano"]
                },
                "falar com humano": {
                    "answer": "Entendo que você gostaria de falar com um atendente humano.\n\nNossos canais de atendimento humano:\n\n📞 Telefone: (61) 98196-1144\n📧 E-mail: suporte@jobfinder.com.br\n💬 WhatsApp: Disponível 24h\n\nHorário: 24 horas por dia, 7 dias por semana\n\nPosso tentar ajudar com mais alguma coisa antes?",
                    "followUp": ["Abrir WhatsApp", "Enviar e-mail", "Voltar ao menu"]
                }
            }
            
            # Generate response based on keywords
            response = None
            
            # Check for exact matches
            for key, resp in responses.items():
                if key in user_message:
                    response = resp
                    break
            
            # Check for keywords if no exact match
            if not response:
                if any(word in user_message for word in ['conta', 'cadastr', 'registr']):
                    response = responses["como criar uma conta"]
                elif any(word in user_message for word in ['serviço', 'solicitar', 'pedir']):
                    response = responses["como solicitar um serviço"]
                elif any(word in user_message for word in ['pagamento', 'pagar', 'cartão', 'pix']):
                    response = responses["problemas com pagamento"]
                elif any(word in user_message for word in ['humano', 'pessoa', 'atendente']):
                    response = responses["falar com humano"]
                else:
                    # Fallback response
                    response = {
                        "answer": "Desculpe, não entendi completamente sua pergunta. Pode reformular?",
                        "followUp": ["Como criar conta", "Solicitar serviço", "Falar com humano"]
                    }
            
            return JsonResponse({
                'success': True,
                'response': response
            })
            
        except Exception as e:
            logger.error(f"Error in chat AI response: {e}")
            return JsonResponse({
                'success': False,
                'error': 'Erro interno do servidor'
            })
    
    return JsonResponse({'success': False, 'error': 'Método não permitido'})


def ai_suggestions(request):
    """Provide AI-powered suggestions based on user behavior"""
    try:
        # Import AI modules
        from .ml_analytics import WebsiteOptimizer
        from .personalization import PersonalizationEngine
        from .content_generator import ContentGenerator
        
        # Initialize AI modules
        optimizer = WebsiteOptimizer()
        personalization = PersonalizationEngine()
        content_generator = ContentGenerator()
        
        # Get user ID (if authenticated) or session key
        user_id = str(request.user.id) if request.user.is_authenticated else request.session.session_key
        
        # Get personalized recommendations
        recommendations = personalization.recommend_content(user_id, current_page=request.path)
        
        # Get trending services
        trending_services = optimizer.get_trending_services()
        
        # Generate personalized content
        personalized_content = None
        if request.user.is_authenticated:
            user_profile = getattr(request.user, 'userprofile', None)
            if user_profile:
                personalized_content = content_generator.generate_personalized_content(
                    user_profile.user_type, 
                    user_profile.city if user_profile.city else "São Paulo"
                )
        
        return JsonResponse({
            'success': True,
            'recommendations': recommendations,
            'trending_services': list(trending_services.values_list('name', flat=True)) if trending_services else [],
            'personalized_content': personalized_content
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def user_achievements(request):
    """Display user achievements and progress"""
    from .models import UserAchievement
    from django.db.models import Q, F
    
    try:
        # Get user's achievements
        user_achievements = UserAchievement.objects.filter(user=request.user).select_related('achievement')
        
        # Calculate total points
        total_points = sum(ua.achievement.points for ua in user_achievements if ua.is_earned)
        
        # Get next potential achievements
        from django.db.models import Q
        next_achievements = UserAchievement.objects.filter(
            user=request.user
        ).filter(
            Q(progress__lt=F('target')) | Q(progress=0)
        ).select_related('achievement')[:5]
        
        return render(request, 'services/user_achievements.html', {
            'user_achievements': user_achievements,
            'total_points': total_points,
            'next_achievements': next_achievements,
        })
    except Exception as e:
        messages.error(request, f'Erro ao carregar conquistas: {str(e)}')
        return redirect('profile')




def robots_txt(request):
    """
    Serve o arquivo robots.txt
    """
    return render(request, 'robots.txt', content_type='text/plain')

def sitemap_xml(request):
    """
    Serve o arquivo sitemap.xml
    """
    return render(request, 'sitemap.xml', content_type='application/xml')

@login_required
def solicitar_servico_pagina(request):
    """Página de solicitação de serviço - GET para mostrar, POST para processar"""
    
    if request.method == 'GET':
        # Mostrar página de solicitação
        service_id = request.GET.get('id', '')
        service_name = request.GET.get('nome', 'Serviço Personalizado')
        service_description = request.GET.get('descricao', 'Descreva o serviço que você precisa')
        service_price = request.GET.get('preco', '0')
        
        context = {
            'service_id': service_id,
            'service_name': service_name,
            'service_description': service_description,
            'service_price': service_price,
        }
        
        return render(request, 'services/solicitar_servico_page.html', context)
    
    elif request.method == 'POST':
        # Processar solicitação
        try:
            # Obter dados do formulário
            servico_id = request.POST.get('servico_id', '0')
            service_name = request.POST.get('service_name', 'Serviço Personalizado')
            service_description = request.POST.get('service_description', 'Serviço solicitado')
            service_price = request.POST.get('service_price', '0')
            
            # Dados pessoais
            nome_completo = request.POST.get('nome', '').strip()
            cpf = request.POST.get('cpf', '').strip()
            telefone = request.POST.get('telefone', '').strip()
            email = request.POST.get('email', '').strip()
            
            # Dados de endereço
            cep = request.POST.get('cep', '').strip()
            endereco = request.POST.get('endereco', '').strip()
            numero = request.POST.get('numero', '').strip()
            complemento = request.POST.get('complemento', '').strip()
            bairro = request.POST.get('bairro', '').strip()
            cidade = request.POST.get('cidade', '').strip()
            estado = request.POST.get('estado', '').strip()
            
            # Dados de agendamento
            data_preferida = request.POST.get('data_preferida', '').strip()
            horario_preferido = request.POST.get('horario_preferido', '').strip()
            periodo_preferido = request.POST.get('periodo_preferido', '').strip()
            observacoes_agendamento = request.POST.get('observacoes_agendamento', '').strip()
            
            # Dados de pagamento
            forma_pagamento = request.POST.get('forma_pagamento', '').strip()
            observacoes_pagamento = request.POST.get('observacoes_pagamento', '').strip()
            
            # Detalhes específicos de pagamento
            tipo_cartao = request.POST.get('tipo_cartao', '').strip()
            parcelas_cartao = request.POST.get('parcelas_cartao', '1').strip()
            precisa_troco = request.POST.get('precisa_troco', 'nao').strip()
            valor_cliente = request.POST.get('valor_cliente', '').strip()
            
            # Processar valores monetários
            client_money_amount = None
            if valor_cliente:
                try:
                    client_money_amount = float(valor_cliente.replace('R$', '').replace('.', '').replace(',', '.').strip())
                except ValueError:
                    pass
            
            # Observações gerais
            observacoes = request.POST.get('observacoes', '').strip()
            
            # Validar campos obrigatórios básicos
            if not all([nome_completo, telefone, email]):
                messages.error(request, 'Por favor, preencha todos os campos obrigatórios (Nome, Telefone e Email).')
                return redirect(f"{request.path}?id={servico_id}&nome={service_name}&descricao={service_description}&preco={service_price}")
            
            # Validar campos de endereço se preenchidos
            if any([cep, endereco, numero, cidade, estado]) and not all([cep, endereco, numero, cidade, estado]):
                messages.error(request, 'Por favor, preencha todos os campos de endereço obrigatórios.')
                return redirect(f"{request.path}?id={servico_id}&nome={service_name}&descricao={service_description}&preco={service_price}")
            
            # Validar email
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError
            try:
                validate_email(email)
            except ValidationError:
                messages.error(request, 'Por favor, insira um email válido.')
                return redirect(f"{request.path}?id={servico_id}&nome={service_name}&descricao={service_description}&preco={service_price}")
            
            # Buscar o serviço se ID foi fornecido
            servico = None
            estimated_price = 0
            
            if servico_id and servico_id != '0':
                try:
                    from .models import CustomService
                    servico = CustomService.objects.get(id=servico_id)
                    estimated_price = servico.estimated_price
                except CustomService.DoesNotExist:
                    pass
            
            # Obter IP do usuário
            def get_client_ip(request):
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                return ip
            
            # Criar a solicitação
            from .models import ServiceRequestModal
            
            # Converter data e horário se fornecidos
            preferred_date_obj = None
            preferred_time_obj = None
            
            if data_preferida:
                from datetime import datetime
                try:
                    preferred_date_obj = datetime.strptime(data_preferida, '%Y-%m-%d').date()
                except ValueError:
                    pass
            
            if horario_preferido:
                try:
                    preferred_time_obj = datetime.strptime(horario_preferido, '%H:%M').time()
                except ValueError:
                    pass
            
            service_request = ServiceRequestModal.objects.create(
                user=request.user,
                service=servico,
                service_name=service_name,
                service_description=service_description,
                estimated_price=estimated_price or float(service_price.replace(',', '.')) if service_price else 0,
                
                # Dados de contato
                contact_name=nome_completo,
                contact_phone=telefone,
                contact_email=email,
                contact_cpf=cpf,
                
                # Dados de endereço
                address_cep=cep,
                address_street=endereco,
                address_number=numero,
                address_complement=complemento,
                address_neighborhood=bairro,
                address_city=cidade,
                address_state=estado,
                
                # Dados de agendamento
                preferred_date=preferred_date_obj,
                preferred_time=preferred_time_obj,
                preferred_period=periodo_preferido,
                schedule_notes=observacoes_agendamento,
                
                # Dados de pagamento
                payment_method=forma_pagamento,
                payment_notes=observacoes_pagamento,
                card_type=tipo_cartao,
                card_installments=int(parcelas_cartao) if parcelas_cartao.isdigit() else 1,
                needs_change=(precisa_troco == 'sim'),
                client_money_amount=client_money_amount,
                pix_identifier=f"JF{service_request.id if 'service_request' in locals() else ''}",
                
                # Observações gerais
                notes=observacoes,
                
                # Dados técnicos
                status='pending',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
            )
            
            # Log da solicitação
            print(f"Nova solicitação criada: ID {service_request.id} - {nome_completo} - {service_name}")
            
            # Enviar email de notificação (opcional)
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                # Email para o cliente
                assunto_cliente = f'Solicitação Recebida - {service_name}'
                mensagem_cliente = f"""
Olá {nome_completo},

Sua solicitação foi recebida com sucesso!

Detalhes:
- Serviço: {service_name}
- Descrição: {service_description}
- Preço Estimado: R$ {estimated_price}
- Observações: {observacoes}

Entraremos em contato em breve através do telefone {telefone}.

Obrigado!
Equipe Job Finder
                """.strip()
                
                send_mail(
                    assunto_cliente,
                    mensagem_cliente,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=True,
                )
                
                # Email para o prestador (se houver)
                if servico and servico.provider and servico.provider.email:
                    assunto_prestador = f'Nova Solicitação - {service_name}'
                    mensagem_prestador = f"""
Olá {servico.provider.get_full_name() or servico.provider.username},

Você recebeu uma nova solicitação para o serviço: {service_name}

Dados do Cliente:
- Nome: {nome_completo}
- Telefone: {telefone}
- Email: {email}
- Observações: {observacoes}

Acesse o painel para mais detalhes.

Equipe Job Finder
                    """.strip()
                    
                    send_mail(
                        assunto_prestador,
                        mensagem_prestador,
                        settings.DEFAULT_FROM_EMAIL,
                        [servico.provider.email],
                        fail_silently=True,
                    )
                    
            except Exception as e:
                print(f"Erro ao enviar email: {e}")
            
            # Mensagem de sucesso
            messages.success(request, f'Solicitação enviada com sucesso! Entraremos em contato no telefone {telefone} em breve.')
            
            # Redirecionar para página de sucesso ou busca
            return redirect('search_new')
            
        except Exception as e:
            print(f"Erro ao processar solicitação: {e}")
            messages.error(request, 'Erro ao processar solicitação. Tente novamente.')
            return redirect(request.path)
    
    else:
        messages.error(request, 'Método não permitido.')
        return redirect('search_new')


def redirect_to_solicitar_completo(request, service_id=None, custom_service_id=None, order_id=None, request_id=None):
    """Redireciona todas as URLs antigas para a página completa de solicitação"""
    return redirect('solicitar_servico_completo')


@login_required
def solicitar_servico_completo(request):
    """Página completa de solicitação de serviço com múltiplas etapas"""
    
    if request.method == 'GET':
        # Mostrar página de solicitação completa
        service_id = request.GET.get('id', '')
        service_name = request.GET.get('nome', 'Serviço Personalizado')
        service_description = request.GET.get('descricao', 'Descreva o serviço que você precisa')
        service_price = request.GET.get('preco', '0')
        
        context = {
            'service_id': service_id,
            'service_name': service_name,
            'service_description': service_description,
            'service_price': service_price,
        }
        
        return render(request, 'services/solicitar_servico_completo.html', context)
    
    elif request.method == 'POST':
        # Processar solicitação completa (usar a mesma lógica da view anterior)
        return solicitar_servico_pagina(request)
    
    else:
        messages.error(request, 'Método não permitido.')
        return redirect('search_new')

@login_required
def processar_solicitacao(request):
    """Processar solicitação de serviço via formulário"""
    from .models import CustomService, ServiceRequestModal
    from datetime import datetime
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('search_new')
    
    try:
        # Obter dados do formulário
        servico_id = request.POST.get('servico_id', '')
        servico_nome = request.POST.get('servico_nome', 'Serviço Personalizado')
        servico_descricao = request.POST.get('servico_descricao', '')
        nome_completo = request.POST.get('nome_completo')
        telefone = request.POST.get('telefone')
        email = request.POST.get('email')
        cep = request.POST.get('cep')
        endereco = request.POST.get('endereco')
        numero = request.POST.get('numero')
        complemento = request.POST.get('complemento', '')
        bairro = request.POST.get('bairro', '')
        cidade = request.POST.get('cidade')
        estado = request.POST.get('estado', 'DF')
        data_preferida = request.POST.get('data_preferida')
        horario_preferido = request.POST.get('horario_preferido', '')
        observacoes = request.POST.get('observacoes', '')
        
        # Validar campos obrigatórios
        if not all([nome_completo, telefone, email, cep, endereco, numero, cidade, data_preferida]):
            messages.error(request, 'Todos os campos obrigatórios devem ser preenchidos.')
            return redirect('solicitar_servico_completo')
        
        # Buscar o serviço e prestador se ID foi fornecido
        servico = None
        prestador = None
        preco_estimado = 0
        if servico_id:
            try:
                servico = CustomService.objects.get(id=servico_id)
                prestador = servico.provider
                preco_estimado = servico.estimated_price
            except CustomService.DoesNotExist:
                pass
        
        # Converter data
        try:
            data_obj = datetime.strptime(data_preferida, "%Y-%m-%d").date()
        except:
            data_obj = datetime.now().date()
        
        # Converter horário se fornecido
        horario_obj = None
        if horario_preferido:
            try:
                horario_obj = datetime.strptime(horario_preferido, "%H:%M").time()
            except:
                pass
        
        # Criar ServiceRequestModal
        solicitacao = ServiceRequestModal.objects.create(
            user=request.user,
            provider=prestador,
            service=servico,
            service_name=servico_nome,
            service_description=servico_descricao or f'Solicitação de {servico_nome}',
            estimated_price=preco_estimado,
            contact_name=nome_completo,
            contact_phone=telefone,
            contact_email=email,
            address_cep=cep,
            address_street=endereco,
            address_number=numero,
            address_complement=complemento,
            address_neighborhood=bairro,
            address_city=cidade,
            address_state=estado,
            preferred_date=data_obj,
            preferred_time=horario_obj,
            schedule_notes=observacoes,
            payment_method='dinheiro',  # Padrão
            status='pending'
        )
        
        messages.success(request, 'Solicitação enviada com sucesso! Entraremos em contato em breve.')
        return render(request, 'services/solicitacao_sucesso.html')
            
    except Exception as e:
        print(f"Erro ao processar solicitação: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, 'Erro ao processar solicitação. Tente novamente.')
        return redirect('solicitar_servico_completo')
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'})
    
    try:
        # Obter dados do formulário
        servico_id = request.POST.get('servico_id')
        nome_completo = request.POST.get('nome_completo')
        telefone = request.POST.get('telefone')
        email = request.POST.get('email')
        cep = request.POST.get('cep')
        endereco = request.POST.get('endereco')
        numero = request.POST.get('numero')
        complemento = request.POST.get('complemento', '')
        cidade = request.POST.get('cidade')
        data_preferida = request.POST.get('data_preferida')
        horario_preferido = request.POST.get('horario_preferido', '')
        observacoes = request.POST.get('observacoes', '')
        
        # Validar campos obrigatórios
        if not all([nome_completo, telefone, email, cep, endereco, numero, cidade, data_preferida]):
            return JsonResponse({'success': False, 'message': 'Todos os campos obrigatórios devem ser preenchidos'})
        
        # Buscar o serviço se ID foi fornecido
        servico = None
        if servico_id and servico_id != '0':
            try:
                servico = CustomService.objects.get(id=servico_id)
            except CustomService.DoesNotExist:
                pass
        
        # Criar a solicitação usando o modelo Order (que já existe)
        try:
            # Criar um pedido simples
            order = Order.objects.create(
                user=request.user,
                service=servico,
                total_amount=servico.estimated_price if servico else 0,
                status='pending',
                notes=f"""
Solicitação via modal:
Nome: {nome_completo}
Telefone: {telefone}
Email: {email}
Endereço: {endereco}, {numero} - {cidade}
CEP: {cep}
Data Preferida: {data_preferida}
Horário: {horario_preferido}
Observações: {observacoes}
                """.strip()
            )
            
            # Adicionar mensagem de sucesso
            messages.success(request, 'Solicitação enviada com sucesso! Entraremos em contato em breve.')
            
            return JsonResponse({
                'success': True, 
                'message': 'Solicitação enviada com sucesso! Entraremos em contato em breve.',
                'order_id': order.id
            })
            
        except Exception as e:
            print(f"Erro ao criar pedido: {e}")
            # Fallback: apenas retornar sucesso
            return JsonResponse({
                'success': True, 
                'message': 'Solicitação recebida com sucesso! Entraremos em contato em breve.'
            })
            
    except Exception as e:
        print(f"Erro ao processar solicitação: {e}")
        return JsonResponse({
            'success': False, 
            'message': 'Erro interno do servidor. Tente novamente.'
        })

@login_required
def solicitacoes_prestador(request):
    """Visualizar solicitações recebidas pelo prestador"""
    from django.core.paginator import Paginator
    from django.db.models import Q, Count
    from .models import ServiceRequestModal, CustomService
    
    # Verificar se o usuário é um prestador
    try:
        if not hasattr(request.user, 'userprofile') or request.user.userprofile.user_type != 'professional':
            messages.error(request, 'Acesso negado. Apenas prestadores podem acessar esta página.')
            return redirect('home')
    except:
        messages.error(request, 'Perfil não encontrado. Complete seu cadastro primeiro.')
        return redirect('profile_new')
    
    # Buscar serviços do prestador
    servicos_prestador = CustomService.objects.filter(provider=request.user)
    
    if not servicos_prestador.exists():
        messages.info(request, 'Você ainda não possui serviços cadastrados.')
        return redirect('provider_dashboard')
    
    # Buscar solicitações para os serviços do prestador
    solicitacoes = ServiceRequestModal.objects.filter(
        service__in=servicos_prestador
    ).select_related('user', 'service').order_by('-created_at')
    
    # Filtrar por status se especificado
    status_filter = request.GET.get('status')
    if status_filter:
        solicitacoes = solicitacoes.filter(status=status_filter)
    
    # Calcular estatísticas
    stats = {
        'pending': ServiceRequestModal.objects.filter(provider=request.user, status='pending').count(),
        'contacted': ServiceRequestModal.objects.filter(provider=request.user, status='contacted').count(),
        'scheduled': ServiceRequestModal.objects.filter(provider=request.user, status='scheduled').count(),
        'completed': ServiceRequestModal.objects.filter(provider=request.user, status='completed').count(),
        'cancelled': ServiceRequestModal.objects.filter(provider=request.user, status='cancelled').count(),
    }
    
    # Paginação
    paginator = Paginator(solicitacoes, 10)  # 10 solicitações por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'services/solicitacoes_prestador.html', {
        'solicitacoes': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'stats': stats,
        'status_filter': status_filter,
    })

@login_required
def alterar_status_solicitacao(request, solicitacao_id):
    """Alterar status de uma solicitação (apenas para prestadores)"""
    if request.method != 'POST':
        messages.error(request, 'Método não permitido')
        return redirect('painel_prestador')
    
    try:
        from .models import ServiceRequestModal
        
        # Verificar se o usuário é um prestador
        if not hasattr(request.user, 'userprofile') or request.user.userprofile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('painel_prestador')
        
        # Buscar a solicitação (pode ser por provider direto ou por service.provider)
        try:
            solicitacao = ServiceRequestModal.objects.get(
                id=solicitacao_id,
                provider=request.user  # Garantir que é do prestador logado
            )
        except ServiceRequestModal.DoesNotExist:
            # Tentar buscar por service.provider
            solicitacao = ServiceRequestModal.objects.get(
                id=solicitacao_id,
                service__provider=request.user
            )
        
        # Obter novo status (suporta tanto AJAX quanto formulário normal)
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            novo_status = data.get('status') or data.get('novo_status')
        else:
            novo_status = request.POST.get('status') or request.POST.get('novo_status')
        
        # Validar status
        status_validos = ['pending', 'contacted', 'scheduled', 'completed', 'cancelled']
        if novo_status not in status_validos:
            error_msg = 'Status inválido'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg})
            else:
                messages.error(request, error_msg)
                return redirect('painel_prestador')
        
        # Validar transições de status
        status_atual = solicitacao.status
        transicoes_validas = {
            'pending': ['contacted', 'cancelled'],
            'contacted': ['scheduled', 'cancelled'],
            'scheduled': ['completed', 'cancelled'],
            'completed': [],  # Status final
            'cancelled': []   # Status final
        }
        
        if novo_status not in transicoes_validas.get(status_atual, []):
            error_msg = f'Não é possível alterar de "{solicitacao.get_status_display()}" para "{dict(ServiceRequestModal.STATUS_CHOICES)[novo_status]}"'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg})
            else:
                messages.error(request, error_msg)
                return redirect('painel_prestador')
        
        # Atualizar status
        status_anterior = solicitacao.get_status_display()
        solicitacao.status = novo_status
        solicitacao.save()
        
        # Log da alteração
        print(f"Status alterado: Solicitação #{solicitacao.id} - {status_anterior} → {solicitacao.get_status_display()}")
        
        # Enviar email de notificação ao cliente
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            status_messages = {
                'contacted': 'Seu prestador entrou em contato! Aguarde o retorno.',
                'scheduled': 'Seu serviço foi agendado! O prestador entrará em contato com mais detalhes.',
                'completed': 'Seu serviço foi concluído! Obrigado por usar nossos serviços.',
                'cancelled': 'Infelizmente seu serviço foi cancelado. Entre em contato conosco se precisar de ajuda.'
            }
            
            if novo_status in status_messages:
                assunto = f'Atualização do seu serviço - {solicitacao.service_name}'
                mensagem = f"""
Olá {solicitacao.contact_name},

{status_messages[novo_status]}

Detalhes do serviço:
- Serviço: {solicitacao.service_name}
- Status: {solicitacao.get_status_display()}
- Prestador: {solicitacao.service.provider.get_full_name() or solicitacao.service.provider.username}

Se tiver dúvidas, entre em contato conosco.

Obrigado!
Equipe Job Finder
                """.strip()
                
                send_mail(
                    assunto,
                    mensagem,
                    settings.DEFAULT_FROM_EMAIL,
                    [solicitacao.contact_email],
                    fail_silently=True,
                )
        except Exception as e:
            print(f"Erro ao enviar email de notificação: {e}")
        
        messages.success(request, f'Status alterado para "{solicitacao.get_status_display()}" com sucesso.')
        
        # Retornar JSON para AJAX ou redirect para formulário normal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Status alterado para "{solicitacao.get_status_display()}" com sucesso.',
                'novo_status': novo_status,
                'novo_status_display': solicitacao.get_status_display()
            })
        else:
            return redirect('painel_prestador')
        
    except ServiceRequestModal.DoesNotExist:
        error_msg = 'Solicitação não encontrada ou você não tem permissão para alterá-la.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': error_msg})
        else:
            messages.error(request, error_msg)
            return redirect('painel_prestador')
    except Exception as e:
        print(f"Erro ao alterar status da solicitação: {e}")
        import traceback
        traceback.print_exc()
        error_msg = 'Erro interno do servidor.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': error_msg})
        else:
            messages.error(request, error_msg)
            return redirect('painel_prestador')

@login_required
def dashboard_prestador_solicitacoes(request):
    """Dashboard resumido de solicitações para prestadores"""
    from .models import ServiceRequestModal, CustomService
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    # Verificar se é prestador
    try:
        if not hasattr(request.user, 'userprofile') or request.user.userprofile.user_type != 'professional':
            return JsonResponse({'error': 'Acesso negado'})
    except:
        return JsonResponse({'error': 'Perfil não encontrado'})
    
    # Buscar serviços do prestador
    servicos_prestador = CustomService.objects.filter(provider=request.user)
    
    # Estatísticas gerais
    hoje = timezone.now().date()
    semana_passada = hoje - timedelta(days=7)
    
    stats = {
        'total_solicitacoes': ServiceRequestModal.objects.filter(provider=request.user).count(),
        'pendentes': ServiceRequestModal.objects.filter(provider=request.user, status='pending').count(),
        'esta_semana': ServiceRequestModal.objects.filter(
            provider=request.user, 
            created_at__date__gte=semana_passada
        ).count(),
        'concluidos': ServiceRequestModal.objects.filter(provider=request.user, status='completed').count(),
    }
    
    # Solicitações recentes (últimas 5)
    solicitacoes_recentes = ServiceRequestModal.objects.filter(
        provider=request.user
    ).select_related('user', 'service').order_by('-created_at')[:5]
    
    return render(request, 'services/dashboard_prestador_solicitacoes.html', {
        'stats': stats,
        'solicitacoes_recentes': solicitacoes_recentes,
    })



@login_required
def minhas_solicitacoes(request):
    """Visualizar solicitações do usuário"""
    from django.core.paginator import Paginator
    from .models import ServiceRequestModal
    
    solicitacoes = ServiceRequestModal.objects.filter(user=request.user).order_by('-created_at')
    
    # Paginação
    paginator = Paginator(solicitacoes, 12)  # 12 solicitações por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'services/minhas_solicitacoes.html', {
        'solicitacoes': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    })

@login_required
def cancelar_solicitacao(request, solicitacao_id):
    """Cancelar uma solicitação"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'})
    
    try:
        from .models import ServiceRequestModal
        
        solicitacao = ServiceRequestModal.objects.get(
            id=solicitacao_id, 
            user=request.user,
            status='pending'  # Só pode cancelar se estiver pendente
        )
        
        solicitacao.status = 'cancelled'
        solicitacao.save()
        
        messages.success(request, 'Solicitação cancelada com sucesso.')
        
        return JsonResponse({
            'success': True,
            'message': 'Solicitação cancelada com sucesso.'
        })
        
    except ServiceRequestModal.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Solicitação não encontrada ou não pode ser cancelada.'
        })
    except Exception as e:
        print(f"Erro ao cancelar solicitação: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Erro interno do servidor.'
        })

@login_required
def solicitar_servico(request):
    """View para processar solicitações de serviço via modal AJAX"""
    print(f"🚀 solicitar_servico chamada - Método: {request.method}")
    
    if request.method == 'POST':
        try:
            # Importar o modelo
            from .models import ServiceRequestModal, CustomService
            
            # Capturar dados do formulário
            service_id = request.POST.get('service_id', '')
            service_name = request.POST.get('service_name', 'Serviço Personalizado')
            service_description = request.POST.get('service_description', '')
            estimated_price = request.POST.get('estimated_price', '0.00')
            contact_name = request.POST.get('contact_name', '')
            contact_phone = request.POST.get('contact_phone', '')
            contact_email = request.POST.get('contact_email', '')
            notes = request.POST.get('notes', '')
            
            print(f"📊 Dados recebidos:")
            print(f"   service_id: {service_id}")
            print(f"   service_name: {service_name}")
            print(f"   service_description: {service_description}")
            print(f"   estimated_price: {estimated_price}")
            print(f"   contact_name: {contact_name}")
            print(f"   contact_phone: {contact_phone}")
            print(f"   contact_email: {contact_email}")
            print(f"   notes: {notes}")
            
            # Validar campos obrigatórios
            if not contact_name or not contact_phone or not contact_email:
                return JsonResponse({
                    'success': False,
                    'message': 'Por favor, preencha todos os campos obrigatórios.'
                })
            
            # Buscar serviço se ID foi fornecido
            service = None
            if service_id:
                try:
                    service = CustomService.objects.get(id=service_id)
                    service_name = service.name
                    service_description = service.description
                    estimated_price = str(service.estimated_price)
                except CustomService.DoesNotExist:
                    pass
            
            # Converter preço para Decimal
            try:
                price_decimal = Decimal(str(estimated_price).replace(',', '.'))
            except:
                price_decimal = Decimal('0.00')
            
            # Criar solicitação
            solicitacao = ServiceRequestModal.objects.create(
                user=request.user,
                service=service,
                service_name=service_name,
                service_description=service_description,
                estimated_price=price_decimal,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email,
                notes=notes,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            # Log para debug
            print(f"✅ Solicitação criada: ID {solicitacao.id}")
            print(f"   Usuário: {request.user.username}")
            print(f"   Serviço: {service_name}")
            print(f"   Preço: R$ {price_decimal}")
            
            # Enviar email de notificação (opcional)
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                # Email para o cliente
                send_mail(
                    subject=f'Solicitação de Serviço Recebida - {service_name}',
                    message=f'''
Olá {contact_name},

Sua solicitação de serviço foi recebida com sucesso!

Detalhes:
- Serviço: {service_name}
- Descrição: {service_description}
- Preço Estimado: R$ {price_decimal}

Entraremos em contato em breve.

Atenciosamente,
Equipe Job Finder
                    ''',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[contact_email],
                    fail_silently=True
                )
                
                # Email para o prestador (se houver)
                if service and service.provider:
                    send_mail(
                        subject=f'Nova Solicitação de Serviço - {service_name}',
                        message=f'''
Olá {service.provider.get_full_name()},

Você recebeu uma nova solicitação de serviço!

Cliente: {contact_name}
Telefone: {contact_phone}
Email: {contact_email}
Serviço: {service_name}
Observações: {notes}

Acesse seu painel para responder.

Atenciosamente,
Equipe Job Finder
                        ''',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[service.provider.email],
                        fail_silently=True
                    )
                
            except Exception as e:
                print(f"⚠️ Erro ao enviar email: {e}")
            
            # Resposta de sucesso
            return JsonResponse({
                'success': True,
                'message': f'Solicitação enviada com sucesso! Entraremos em contato em breve.',
                'solicitacao_id': solicitacao.id
            })
            
        except Exception as e:
            print(f"❌ Erro ao processar solicitação: {e}")
            return JsonResponse({
                'success': False,
                'message': f'Erro interno do servidor. Tente novamente.'
            })
    
    # Se não for POST, retornar erro
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido.'
    })


def minhas_solicitacoes(request):
    """View para listar solicitações do usuário"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    from .models import ServiceRequestModal
    
    solicitacoes = ServiceRequestModal.objects.filter(
        user=request.user
    ).order_by('-created_at')
    
    return render(request, 'services/minhas_solicitacoes.html', {
        'solicitacoes': solicitacoes
    })


@login_required
def cancelar_solicitacao(request, solicitacao_id):
    """View para cancelar uma solicitação"""
    from .models import ServiceRequestModal
    
    try:
        solicitacao = ServiceRequestModal.objects.get(
            id=solicitacao_id,
            user=request.user
        )
        
        if solicitacao.status == 'pending':
            solicitacao.status = 'cancelled'
            solicitacao.save()
            messages.success(request, 'Solicitação cancelada com sucesso.')
        else:
            messages.error(request, 'Esta solicitação não pode ser cancelada.')
            
    except ServiceRequestModal.DoesNotExist:
        messages.error(request, 'Solicitação não encontrada.')
    
    return redirect('minhas_solicitacoes')


@login_required
def solicitacoes_prestador(request):
    """Redireciona para o painel do prestador"""
    return redirect('painel_prestador')


@login_required
def dashboard_prestador_solicitacoes(request):
    """Dashboard para prestadores gerenciarem solicitações"""
    from .models import ServiceRequestModal
    from django.db.models import Count
    
    # Verificar se é prestador
    try:
        if request.user.userprofile.user_type != 'professional':
            messages.error(request, 'Acesso negado.')
            return redirect('home')
    except:
        messages.error(request, 'Perfil não encontrado.')
        return redirect('home')
    
    # Estatísticas
    solicitacoes = ServiceRequestModal.objects.filter(service__provider=request.user)
    
    stats = {
        'total': solicitacoes.count(),
        'pendentes': solicitacoes.filter(status='pending').count(),
        'contatadas': solicitacoes.filter(status='contacted').count(),
        'agendadas': solicitacoes.filter(status='scheduled').count(),
        'concluidas': solicitacoes.filter(status='completed').count(),
    }
    
    # Solicitações recentes
    solicitacoes_recentes = solicitacoes.order_by('-created_at')[:10]
    
    return render(request, 'services/dashboard_prestador_solicitacoes.html', {
        'stats': stats,
        'solicitacoes_recentes': solicitacoes_recentes
    })

# ==========================================
# NOVO FLUXO DE SOLICITAÇÃO EM 4 ETAPAS
# ==========================================

from django.utils import timezone
from datetime import timedelta
import json

def get_or_create_session(request, service_id):
    """Função auxiliar para gerenciar sessão do fluxo"""
    from .models import ServiceRequestSession
    
    session_key = request.session.session_key
    if not session_key:
        request.session.create()
        session_key = request.session.session_key
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            service_id=service_id,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        # Criar nova sessão com expiração de 2 horas
        expires_at = timezone.now() + timedelta(hours=2)
        session_obj = ServiceRequestSession.objects.create(
            session_key=session_key,
            user=request.user if request.user.is_authenticated else None,
            service_id=service_id,
            expires_at=expires_at
        )
    
    return session_obj


@login_required
def solicitar_step1(request, service_id):
    """Etapa 1: Dados básicos da solicitação"""
    from .models import CustomService
    
    try:
        service = CustomService.objects.get(id=service_id)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado.')
        return redirect('search_new')
    
    # Criar ou recuperar sessão
    session_obj = get_or_create_session(request, service_id)
    
    # Dados do formulário (se existirem na sessão)
    form_data = session_obj.step1_data or {}
    
    # Pré-preencher com dados do usuário se não existirem
    if not form_data and request.user.is_authenticated:
        form_data = {
            'contact_name': request.user.get_full_name(),
            'contact_email': request.user.email,
        }
    
    context = {
        'service': service,
        'current_step': 1,
        'form_data': form_data,
        'session_obj': session_obj,
    }
    
    return render(request, 'services/solicitar_step1.html', context)


@login_required
def solicitar_step1_post(request):
    """Processar dados da Etapa 1 e avançar para Etapa 2"""
    if request.method != 'POST':
        return redirect('search_new')
    
    # Recuperar sessão
    session_key = request.session.session_key
    if not session_key:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    # Validar dados obrigatórios
    required_fields = ['contact_name', 'contact_email', 'contact_phone', 'service_description']
    errors = []
    
    for field in required_fields:
        if not request.POST.get(field, '').strip():
            errors.append(f'Campo {field} é obrigatório.')
    
    # Validar email
    import re
    email = request.POST.get('contact_email', '').strip()
    if email and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        errors.append('Formato de email inválido.')
    
    # Validar telefone
    phone = request.POST.get('contact_phone', '').strip()
    if phone and not re.match(r'^\(\d{2}\)\s\d{4,5}-\d{4}$', phone):
        errors.append('Formato de telefone inválido. Use: (11) 99999-9999')
    
    if errors:
        for error in errors:
            messages.error(request, error)
        return redirect('solicitar_servico_step1', service_id=session_obj.service_id)
    
    # Salvar dados na sessão
    step1_data = {
        'contact_name': request.POST.get('contact_name', '').strip(),
        'contact_email': email,
        'contact_phone': phone,
        'contact_cpf': request.POST.get('contact_cpf', '').strip(),
        'service_description': request.POST.get('service_description', '').strip(),
        'notes': request.POST.get('notes', '').strip(),
        'urgency': request.POST.get('urgency', 'medium'),
    }
    
    session_obj.step1_data = step1_data
    session_obj.current_step = 2
    session_obj.save()
    
    messages.success(request, 'Dados salvos! Prossiga para o agendamento.')
    return redirect('solicitar_step2')


@login_required
def solicitar_step2(request):
    """Etapa 2: Agendamento do serviço"""
    from .models import CustomService
    
    # Recuperar sessão
    session_key = request.session.session_key
    if not session_key:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    # Verificar se etapa 1 foi concluída
    if not session_obj.step1_data or session_obj.current_step < 2:
        messages.warning(request, 'Complete a etapa anterior primeiro.')
        return redirect('solicitar_servico_step1', service_id=session_obj.service_id)
    
    # Buscar serviço
    try:
        service = CustomService.objects.get(id=session_obj.service_id)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado.')
        return redirect('search_new')
    
    if request.method == 'POST':
        # Processar dados da etapa 2
        required_fields = ['preferred_date', 'address_cep', 'address_street', 
                          'address_number', 'address_neighborhood', 'address_city', 'address_state']
        errors = []
        
        for field in required_fields:
            if not request.POST.get(field, '').strip():
                errors.append(f'Campo {field} é obrigatório.')
        
        # Validar data
        from datetime import datetime, date
        try:
            preferred_date = datetime.strptime(request.POST.get('preferred_date'), '%Y-%m-%d').date()
            if preferred_date < date.today():
                errors.append('A data não pode ser no passado.')
        except (ValueError, TypeError):
            errors.append('Data inválida.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            # Salvar dados na sessão
            step2_data = {
                'preferred_date': request.POST.get('preferred_date'),
                'preferred_time': request.POST.get('preferred_time', ''),
                'preferred_period': request.POST.get('preferred_period', ''),
                'address_cep': request.POST.get('address_cep', '').strip(),
                'address_street': request.POST.get('address_street', '').strip(),
                'address_number': request.POST.get('address_number', '').strip(),
                'address_complement': request.POST.get('address_complement', '').strip(),
                'address_neighborhood': request.POST.get('address_neighborhood', '').strip(),
                'address_city': request.POST.get('address_city', '').strip(),
                'address_state': request.POST.get('address_state', '').strip(),
                'reference_point': request.POST.get('reference_point', '').strip(),
                'schedule_notes': request.POST.get('schedule_notes', '').strip(),
            }
            
            session_obj.step2_data = step2_data
            session_obj.current_step = 3
            session_obj.save()
            
            messages.success(request, 'Agendamento salvo! Escolha a forma de pagamento.')
            return redirect('solicitar_step3')
    
    # Dados do formulário
    form_data = session_obj.step2_data or {}
    
    context = {
        'service': service,
        'current_step': 2,
        'form_data': form_data,
        'today': date.today().strftime('%Y-%m-%d'),
    }
    
    return render(request, 'services/solicitar_step2.html', context)


@login_required
def solicitar_step3(request):
    """Etapa 3: Forma de pagamento"""
    from .models import CustomService
    
    # Recuperar sessão
    session_key = request.session.session_key
    if not session_key:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    # Verificar se etapas anteriores foram concluídas
    if not session_obj.step2_data or session_obj.current_step < 3:
        messages.warning(request, 'Complete as etapas anteriores primeiro.')
        return redirect('solicitar_step2')
    
    # Buscar serviço
    try:
        service = CustomService.objects.get(id=session_obj.service_id)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado.')
        return redirect('search_new')
    
    if request.method == 'POST':
        # Processar dados da etapa 3
        payment_method = request.POST.get('payment_method', '').strip()
        
        if not payment_method:
            messages.error(request, 'Selecione uma forma de pagamento.')
        else:
            # Salvar dados na sessão
            step3_data = {
                'payment_method': payment_method,
                'payment_notes': request.POST.get('payment_notes', '').strip(),
            }
            
            # Dados específicos por método de pagamento
            if payment_method == 'dinheiro':
                step3_data.update({
                    'needs_change': request.POST.get('needs_change') == 'on',
                    'client_money_amount': request.POST.get('client_money_amount', ''),
                })
            elif payment_method == 'cartao':
                step3_data.update({
                    'card_type': request.POST.get('card_type', ''),
                    'card_installments': int(request.POST.get('card_installments', 1)),
                })
            elif payment_method == 'pix':
                step3_data.update({
                    'pix_identifier': request.POST.get('pix_identifier', '').strip(),
                })
            
            session_obj.step3_data = step3_data
            session_obj.current_step = 4
            session_obj.save()
            
            messages.success(request, 'Pagamento configurado! Revise e confirme sua solicitação.')
            return redirect('solicitar_step4')
    
    # Dados do formulário
    form_data = session_obj.step3_data or {}
    
    context = {
        'service': service,
        'current_step': 3,
        'form_data': form_data,
    }
    
    return render(request, 'services/solicitar_step3.html', context)


@login_required
def solicitar_step4(request):
    """Etapa 4: Confirmação final"""
    from .models import CustomService
    
    # Recuperar sessão
    session_key = request.session.session_key
    if not session_key:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    # Verificar se todas as etapas anteriores foram concluídas
    if not all([session_obj.step1_data, session_obj.step2_data, session_obj.step3_data]) or session_obj.current_step < 4:
        messages.warning(request, 'Complete todas as etapas anteriores primeiro.')
        return redirect('solicitar_step3')
    
    # Buscar serviço
    try:
        service = CustomService.objects.get(id=session_obj.service_id)
    except CustomService.DoesNotExist:
        messages.error(request, 'Serviço não encontrado.')
        return redirect('search_new')
    
    if request.method == 'POST':
        # Processar confirmação final
        required_checkboxes = ['terms_service', 'data_accuracy', 'payment_agreement', 'contact_permission']
        
        for checkbox in required_checkboxes:
            if not request.POST.get(checkbox):
                messages.error(request, 'Você deve aceitar todos os termos para continuar.')
                break
        else:
            # Todos os termos aceitos, criar solicitação
            return redirect('solicitar_confirm')
    
    # Calcular troco se necessário
    troco_calculado = 0
    if (session_obj.step3_data.get('payment_method') == 'dinheiro' and 
        session_obj.step3_data.get('needs_change') and 
        session_obj.step3_data.get('client_money_amount')):
        try:
            client_amount = float(session_obj.step3_data.get('client_money_amount', 0))
            service_price = float(service.price)
            troco_calculado = client_amount - service_price
        except (ValueError, TypeError):
            troco_calculado = 0
    
    context = {
        'service': service,
        'current_step': 4,
        'step1_data': session_obj.step1_data,
        'step2_data': session_obj.step2_data,
        'step3_data': session_obj.step3_data,
        'troco_calculado': troco_calculado,
    }
    
    return render(request, 'services/solicitar_step4.html', context)


def send_service_request_notifications(service_request):
    """Envia notificações por email para cliente e prestador"""
    from django.core.mail import send_mail
    from django.conf import settings
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        # Email para o cliente
        client_subject = f"Solicitação de Serviço Confirmada - {service_request.service_name}"
        client_message = f"""
Olá {service_request.contact_name},

Sua solicitação de serviço foi recebida com sucesso!

Serviço: {service_request.service_name}
Prestador: {service_request.provider.get_full_name() if service_request.provider else 'N/A'}
Data: {service_request.preferred_date}
Horário: {service_request.preferred_time if service_request.preferred_time else service_request.preferred_period}
Endereço: {service_request.address_street}, {service_request.address_number} - {service_request.address_neighborhood}, {service_request.address_city}/{service_request.address_state}

O prestador foi notificado e entrará em contato em breve.

Você pode acompanhar o status da sua solicitação acessando "Meus Pedidos" no site.

Atenciosamente,
Equipe Job Finder
"""
        
        send_mail(
            client_subject,
            client_message,
            settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@jobfinder.com',
            [service_request.contact_email],
            fail_silently=True,
        )
        logger.info(f"Email enviado para cliente: {service_request.contact_email}")
        
    except Exception as e:
        logger.error(f"Erro ao enviar email para cliente: {str(e)}")
    
    try:
        # Email para o prestador
        if service_request.provider and service_request.provider.email:
            provider_subject = f"Nova Solicitação de Serviço - {service_request.service_name}"
            provider_message = f"""
Olá {service_request.provider.get_full_name()},

Você recebeu uma nova solicitação de serviço!

Cliente: {service_request.contact_name}
Telefone: {service_request.contact_phone}
Email: {service_request.contact_email}

Serviço: {service_request.service_name}
Descrição: {service_request.service_description}

Data solicitada: {service_request.preferred_date}
Horário: {service_request.preferred_time if service_request.preferred_time else service_request.preferred_period}

Endereço:
{service_request.address_street}, {service_request.address_number}
{service_request.address_complement}
{service_request.address_neighborhood}
{service_request.address_city}/{service_request.address_state}
CEP: {service_request.address_cep}

Forma de pagamento: {service_request.get_payment_method_display()}

Observações: {service_request.notes if service_request.notes else 'Nenhuma'}

Acesse seu painel para confirmar ou gerenciar esta solicitação.

Atenciosamente,
Equipe Job Finder
"""
            
            send_mail(
                provider_subject,
                provider_message,
                settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@jobfinder.com',
                [service_request.provider.email],
                fail_silently=True,
            )
            logger.info(f"Email enviado para prestador: {service_request.provider.email}")
            
    except Exception as e:
        logger.error(f"Erro ao enviar email para prestador: {str(e)}")


@login_required
def solicitar_confirm(request):
    """Processar confirmação final e criar solicitação"""
    if request.method != 'POST':
        return redirect('search_new')
    
    from .models import ServiceRequestModal, CustomService
    from datetime import datetime
    
    # Recuperar sessão
    session_key = request.session.session_key
    if not session_key:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    try:
        session_obj = ServiceRequestSession.objects.get(
            session_key=session_key,
            expires_at__gt=timezone.now()
        )
    except ServiceRequestSession.DoesNotExist:
        messages.error(request, 'Sessão expirada. Reinicie o processo.')
        return redirect('search_new')
    
    # Verificar se todas as etapas foram concluídas
    if not all([session_obj.step1_data, session_obj.step2_data, session_obj.step3_data]):
        messages.error(request, 'Processo incompleto. Reinicie a solicitação.')
        return redirect('search_new')
    
    try:
        # Buscar serviço
        service = CustomService.objects.get(id=session_obj.service_id)
        
        # CRÍTICO: Identificar o prestador
        provider = service.provider
        
        # Criar solicitação
        solicitacao = ServiceRequestModal.objects.create(
            user=request.user,
            provider=provider,
            service=service,
            service_name=service.name,
            service_description=session_obj.step1_data['service_description'],
            estimated_price=service.estimated_price,
            
            # Dados do cliente (Etapa 1)
            contact_name=session_obj.step1_data['contact_name'],
            contact_email=session_obj.step1_data['contact_email'],
            contact_phone=session_obj.step1_data['contact_phone'],
            contact_cpf=session_obj.step1_data.get('contact_cpf', ''),
            notes=session_obj.step1_data.get('notes', ''),
            
            # Dados de agendamento (Etapa 2)
            preferred_date=datetime.strptime(session_obj.step2_data['preferred_date'], '%Y-%m-%d').date(),
            preferred_time=datetime.strptime(session_obj.step2_data['preferred_time'], '%H:%M').time() if session_obj.step2_data.get('preferred_time') else None,
            preferred_period=session_obj.step2_data.get('preferred_period', ''),
            address_cep=session_obj.step2_data['address_cep'],
            address_street=session_obj.step2_data['address_street'],
            address_number=session_obj.step2_data['address_number'],
            address_complement=session_obj.step2_data.get('address_complement', ''),
            address_neighborhood=session_obj.step2_data['address_neighborhood'],
            address_city=session_obj.step2_data['address_city'],
            address_state=session_obj.step2_data['address_state'],
            schedule_notes=session_obj.step2_data.get('schedule_notes', ''),
            
            # Dados de pagamento (Etapa 3)
            payment_method=session_obj.step3_data['payment_method'],
            payment_notes=session_obj.step3_data.get('payment_notes', ''),
            card_type=session_obj.step3_data.get('card_type', ''),
            card_installments=session_obj.step3_data.get('card_installments', 1),
            needs_change=session_obj.step3_data.get('needs_change', False),
            client_money_amount=float(session_obj.step3_data.get('client_money_amount', 0)) if session_obj.step3_data.get('client_money_amount') else None,
            pix_identifier=session_obj.step3_data.get('pix_identifier', ''),
            
            # Metadados
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        # Log de sucesso
        logger = logging.getLogger(__name__)
        logger.info(f"ServiceRequestModal {solicitacao.id} criado com sucesso: "
                   f"cliente={request.user.id}, prestador={provider.id}, serviço={service.name}")
        
        # Enviar notificações
        try:
            send_service_request_notifications(solicitacao)
            solicitacao.client_notified = True
            solicitacao.provider_notified = True
            solicitacao.save()
            logger.info(f"Notificações enviadas para ServiceRequestModal {solicitacao.id}")
        except Exception as e:
            logger.error(f"Erro ao enviar notificações para ServiceRequestModal {solicitacao.id}: {str(e)}")
        
        # Limpar sessão
        session_obj.delete()
        
        messages.success(request, f'Solicitação #{solicitacao.id} criada com sucesso! O prestador será notificado.')
        return redirect('acompanhar_solicitacao', request_id=solicitacao.id)
        
    except CustomService.DoesNotExist:
        logger.error(f"Serviço não encontrado: {session_obj.service_id}")
        messages.error(request, 'Serviço não encontrado.')
        return redirect('search_new')
    except Exception as e:
        logger.error(f"Erro ao criar ServiceRequestModal: {str(e)}", exc_info=True)
        messages.error(request, f'Erro ao criar solicitação: {str(e)}')
        return redirect('search_new')


@login_required
def acompanhar_solicitacao(request, request_id):
    """Página de acompanhamento da solicitação - exibe detalhes completos de uma solicitação específica"""
    from .models import ServiceRequestModal
    from django.shortcuts import get_object_or_404
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info(f"acompanhar_solicitacao acessado por user={request.user.id}, request_id={request_id}")
    
    # Buscar solicitação garantindo que pertence ao usuário logado (isolamento de dados)
    request_obj = get_object_or_404(
        ServiceRequestModal,
        id=request_id,
        user=request.user
    )
    
    logger.info(f"Solicitação {request_id} encontrada: status={request_obj.status}, service={request_obj.service_name}")
    
    context = {
        'request_obj': request_obj,
    }
    
    return render(request, 'services/acompanhar_solicitacao.html', context)


@login_required
def meus_pedidos(request):
    """Página Meus Pedidos - exibe todas as solicitações do cliente"""
    from .models import ServiceRequestModal
    from django.db.models import Count, Q
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info(f"meus_pedidos acessado por user={request.user.id}")
    
    try:
        # Validar filtro de status
        status_filter = request.GET.get('status')
        valid_statuses = ['pending', 'contacted', 'scheduled', 'completed', 'cancelled']
        
        if status_filter and status_filter not in valid_statuses:
            messages.warning(request, f'Status inválido: {status_filter}')
            status_filter = None
        
        # Filtrar solicitações do usuário logado (cliente)
        solicitacoes = ServiceRequestModal.objects.filter(
            user=request.user
        ).select_related('provider', 'service').order_by('-created_at')
        
        # Aplicar filtro de status se válido
        if status_filter:
            solicitacoes = solicitacoes.filter(status=status_filter)
            logger.info(f"Filtro aplicado: status={status_filter}")
        
        # Calcular estatísticas com uma única query otimizada
        stats = ServiceRequestModal.objects.filter(user=request.user).aggregate(
            total=Count('id'),
            pendentes=Count('id', filter=Q(status='pending')),
            agendadas=Count('id', filter=Q(status='scheduled')),
            concluidas=Count('id', filter=Q(status='completed'))
        )
        
        logger.info(f"Solicitações encontradas: total={stats['total']}, pendentes={stats['pendentes']}")
        
        context = {
            'solicitacoes': solicitacoes,
            'total': stats['total'],
            'pendentes': stats['pendentes'],
            'agendadas': stats['agendadas'],
            'concluidas': stats['concluidas'],
            'status_filter': status_filter,
        }
        
        return render(request, 'services/meus_pedidos.html', context)
        
    except Exception as e:
        logger.error(f'Erro em meus_pedidos: {str(e)}', exc_info=True)
        messages.error(request, 'Ocorreu um erro ao carregar suas solicitações. Por favor, tente novamente.')
        return render(request, 'services/meus_pedidos.html', {
            'solicitacoes': [],
            'total': 0,
            'pendentes': 0,
            'agendadas': 0,
            'concluidas': 0,
            'status_filter': None,
        })


@login_required
def painel_prestador(request):
    """Painel do Prestador - exibe solicitações recebidas"""
    from .models import ServiceRequestModal, UserProfile
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info(f"painel_prestador acessado por user={request.user.id}")
    
    # Verificar se usuário é prestador
    try:
        user_profile = request.user.userprofile
        if user_profile.user_type != 'professional':
            messages.warning(request, 'Você precisa ser um prestador para acessar este painel.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil não encontrado.')
        return redirect('home')
    
    # Filtrar solicitações recebidas pelo prestador
    solicitacoes = ServiceRequestModal.objects.filter(
        provider=request.user
    ).select_related('user', 'service').order_by('-created_at')
    
    # Filtros por status
    status_filter = request.GET.get('status', 'pending')
    if status_filter and status_filter != 'all':
        solicitacoes = solicitacoes.filter(status=status_filter)
        logger.info(f"Filtro aplicado: status={status_filter}")
    
    # Contadores
    total = ServiceRequestModal.objects.filter(provider=request.user).count()
    pendentes = ServiceRequestModal.objects.filter(provider=request.user, status='pending').count()
    agendadas = ServiceRequestModal.objects.filter(provider=request.user, status='scheduled').count()
    concluidas = ServiceRequestModal.objects.filter(provider=request.user, status='completed').count()
    
    logger.info(f"Solicitações recebidas: total={total}, pendentes={pendentes}")
    
    context = {
        'solicitacoes': solicitacoes,
        'total': total,
        'pendentes': pendentes,
        'agendadas': agendadas,
        'concluidas': concluidas,
        'status_filter': status_filter,
    }
    
    return render(request, 'services/painel_prestador.html', context)


@login_required
def confirmar_solicitacao(request, request_id):
    """Prestador confirma uma solicitação"""
    from .models import ServiceRequestModal
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        solicitacao = ServiceRequestModal.objects.get(
            id=request_id,
            provider=request.user
        )
        
        if solicitacao.status == 'pending':
            solicitacao.status = 'scheduled'
            solicitacao.save()
            
            logger.info(f"Solicitação {request_id} confirmada por prestador {request.user.id}")
            messages.success(request, 'Solicitação confirmada com sucesso!')
            
            # Enviar notificação ao cliente
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                send_mail(
                    'Solicitação Confirmada',
                    f'Olá {solicitacao.contact_name},\n\nSua solicitação de serviço foi confirmada pelo prestador!\n\nServiço: {solicitacao.service_name}\nData: {solicitacao.preferred_date}\n\nO prestador entrará em contato em breve.',
                    settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@jobfinder.com',
                    [solicitacao.contact_email],
                    fail_silently=True,
                )
            except Exception as e:
                logger.error(f"Erro ao enviar email de confirmação: {str(e)}")
        else:
            messages.warning(request, 'Esta solicitação já foi processada.')
            
    except ServiceRequestModal.DoesNotExist:
        messages.error(request, 'Solicitação não encontrada.')
    
    return redirect('painel_prestador')


@login_required
def rejeitar_solicitacao(request, request_id):
    """Prestador rejeita uma solicitação"""
    from .models import ServiceRequestModal
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        messages.error(request, 'Método não permitido.')
        return redirect('painel_prestador')
    
    try:
        solicitacao = ServiceRequestModal.objects.get(
            id=request_id,
            provider=request.user
        )
        
        if solicitacao.status == 'pending':
            # Obter motivo e observações do formulário
            motivo = request.POST.get('motivo_rejeicao', '')
            observacoes = request.POST.get('observacoes', '')
            
            # Mapear motivos para texto legível
            motivos_dict = {
                'agenda_cheia': 'Agenda cheia',
                'fora_area': 'Fora da área de atendimento',
                'nao_realizo': 'Não realizo este tipo de serviço',
                'valor_baixo': 'Valor não compatível',
                'outro': 'Outro motivo'
            }
            motivo_texto = motivos_dict.get(motivo, motivo)
            
            # Atualizar status e adicionar motivo nas notas
            solicitacao.status = 'cancelled'
            solicitacao.rejection_reason = f"Motivo: {motivo_texto}"
            if observacoes:
                solicitacao.rejection_reason += f"\nObservações: {observacoes}"
            solicitacao.save()
            
            logger.info(f"Solicitação {request_id} rejeitada por prestador {request.user.id}. Motivo: {motivo}")
            messages.success(request, 'Solicitação recusada com sucesso.')
            
            # Enviar notificação ao cliente
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                mensagem_email = f"""Olá {solicitacao.contact_name},

Infelizmente o prestador não pode aceitar sua solicitação de serviço no momento.

Serviço: {solicitacao.service_name}
Motivo: {motivo_texto}
"""
                if observacoes:
                    mensagem_email += f"\nObservações do prestador: {observacoes}\n"
                
                mensagem_email += "\nVocê pode buscar outros prestadores disponíveis em nossa plataforma."
                
                send_mail(
                    'Solicitação Não Aceita',
                    mensagem_email,
                    settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@jobfinder.com',
                    [solicitacao.contact_email],
                    fail_silently=True,
                )
            except Exception as e:
                logger.error(f"Erro ao enviar email de rejeição: {str(e)}")
        else:
            messages.warning(request, 'Esta solicitação já foi processada.')
            
    except ServiceRequestModal.DoesNotExist:
        messages.error(request, 'Solicitação não encontrada.')
    
    return redirect('painel_prestador')


# APIs auxiliares
@login_required
def validate_step_api(request):
    """API para validar dados de uma etapa"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    # Implementar validações específicas
    return JsonResponse({'valid': True})


@login_required
def check_availability_api(request):
    """API para verificar disponibilidade de horários"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    # Implementar verificação de disponibilidade
    return JsonResponse({'available': True})


@login_required
def teste_solicitacao(request):
    """Página de teste para acessar o fluxo de solicitação"""
    from .models import CustomService
    
    # Buscar alguns serviços para teste
    services = CustomService.objects.all()[:10]
    
    context = {
        'services': services,
    }
    
    return render(request, 'services/teste_solicitacao.html', context)


# ============================================
# PROVIDER FEATURES: Agendamento, Avaliações, Financeiro, Relatórios
# ============================================

@login_required
def provider_agendamento(request):
    """Página de agendamento do prestador"""
    # Verificar se é prestador
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Buscar agendamentos do prestador
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    hoje = timezone.now().date()
    proximos_7_dias = hoje + timedelta(days=7)
    
    # Buscar solicitações agendadas
    try:
        from .models import ServiceRequestModal
        
        # CORRIGIDO: Filtrar diretamente por provider
        agendamentos_hoje = ServiceRequestModal.objects.filter(
            provider=request.user,
            preferred_date=hoje,
            status__in=['scheduled', 'contacted']
        ).order_by('preferred_time')
        
        agendamentos_proximos = ServiceRequestModal.objects.filter(
            provider=request.user,
            preferred_date__range=[hoje + timedelta(days=1), proximos_7_dias],
            status__in=['scheduled', 'contacted']
        ).order_by('preferred_date', 'preferred_time')
        
        agendamentos_pendentes = ServiceRequestModal.objects.filter(
            provider=request.user,
            status='pending'
        ).order_by('-created_at')[:10]
        
    except Exception as e:
        agendamentos_hoje = []
        agendamentos_proximos = []
        agendamentos_pendentes = []
        messages.error(request, f'Erro ao carregar agendamentos: {str(e)}')
    
    context = {
        'agendamentos_hoje': agendamentos_hoje,
        'agendamentos_proximos': agendamentos_proximos,
        'agendamentos_pendentes': agendamentos_pendentes,
        'hoje': hoje,
    }
    
    return render(request, 'services/provider_agendamento.html', context)


@login_required
def provider_avaliacoes(request):
    """Página de avaliações do prestador"""
    # Verificar se é prestador
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Buscar avaliações do prestador
    try:
        # Buscar pedidos concluídos com avaliações (usando o relacionamento review)
        pedidos_avaliados = Order.objects.filter(
            professional=request.user,
            status='completed'
        ).exclude(review__isnull=True).select_related('review', 'customer').order_by('-updated_at')
        
        # Calcular estatísticas
        total_avaliacoes = pedidos_avaliados.count()
        
        if total_avaliacoes > 0:
            soma_ratings = sum([p.review.rating for p in pedidos_avaliados if hasattr(p, 'review')])
            media_rating = soma_ratings / total_avaliacoes if total_avaliacoes > 0 else 0
            
            # Contar por estrelas (usando review__rating)
            ratings_5 = pedidos_avaliados.filter(review__rating=5).count()
            ratings_4 = pedidos_avaliados.filter(review__rating=4).count()
            ratings_3 = pedidos_avaliados.filter(review__rating=3).count()
            ratings_2 = pedidos_avaliados.filter(review__rating=2).count()
            ratings_1 = pedidos_avaliados.filter(review__rating=1).count()
        else:
            media_rating = 0
            ratings_5 = ratings_4 = ratings_3 = ratings_2 = ratings_1 = 0
        
    except Exception as e:
        pedidos_avaliados = []
        total_avaliacoes = 0
        media_rating = 0
        ratings_5 = ratings_4 = ratings_3 = ratings_2 = ratings_1 = 0
        messages.error(request, f'Erro ao carregar avaliações: {str(e)}')
    
    # Calcular porcentagens
    if total_avaliacoes > 0:
        percent_5 = round((ratings_5 / total_avaliacoes) * 100, 1)
        percent_4 = round((ratings_4 / total_avaliacoes) * 100, 1)
        percent_3 = round((ratings_3 / total_avaliacoes) * 100, 1)
        percent_2 = round((ratings_2 / total_avaliacoes) * 100, 1)
        percent_1 = round((ratings_1 / total_avaliacoes) * 100, 1)
    else:
        percent_5 = percent_4 = percent_3 = percent_2 = percent_1 = 0
    
    context = {
        'avaliacoes': pedidos_avaliados[:20],  # Últimas 20 avaliações
        'total_avaliacoes': total_avaliacoes,
        'media_rating': round(media_rating, 1),
        'ratings_5': ratings_5,
        'ratings_4': ratings_4,
        'ratings_3': ratings_3,
        'ratings_2': ratings_2,
        'ratings_1': ratings_1,
        'percent_5': percent_5,
        'percent_4': percent_4,
        'percent_3': percent_3,
        'percent_2': percent_2,
        'percent_1': percent_1,
    }
    
    return render(request, 'services/provider_avaliacoes.html', context)


@login_required
def provider_financeiro(request):
    """Página financeira do prestador"""
    # Verificar se é prestador
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Buscar dados financeiros
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Sum, Count, Q
    
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    inicio_mes_passado = (inicio_mes - timedelta(days=1)).replace(day=1)
    
    try:
        # Pedidos do prestador
        todos_pedidos = Order.objects.filter(professional=request.user)
        
        # Receita total
        receita_total = 0
        for pedido in todos_pedidos.filter(status='completed'):
            try:
                receita_total += float(pedido.total_price)
            except:
                pass
        
        # Receita deste mês
        receita_mes_atual = 0
        pedidos_mes_atual = todos_pedidos.filter(
            status='completed',
            updated_at__date__gte=inicio_mes
        )
        for pedido in pedidos_mes_atual:
            try:
                receita_mes_atual += float(pedido.total_price)
            except:
                pass
        
        # Receita mês passado
        receita_mes_passado = 0
        pedidos_mes_passado = todos_pedidos.filter(
            status='completed',
            updated_at__date__gte=inicio_mes_passado,
            updated_at__date__lt=inicio_mes
        )
        for pedido in pedidos_mes_passado:
            try:
                receita_mes_passado += float(pedido.total_price)
            except:
                pass
        
        # Calcular crescimento
        if receita_mes_passado > 0:
            crescimento_percentual = ((receita_mes_atual - receita_mes_passado) / receita_mes_passado) * 100
        else:
            crescimento_percentual = 100 if receita_mes_atual > 0 else 0
        
        # Pedidos por status
        pedidos_concluidos = todos_pedidos.filter(status='completed').count()
        pedidos_pendentes = todos_pedidos.filter(status='pending').count()
        pedidos_em_andamento = todos_pedidos.filter(status='in_progress').count()
        
        # Últimas transações
        ultimas_transacoes = todos_pedidos.filter(status='completed').order_by('-updated_at')[:10]
        
    except Exception as e:
        receita_total = 0
        receita_mes_atual = 0
        receita_mes_passado = 0
        crescimento_percentual = 0
        pedidos_concluidos = 0
        pedidos_pendentes = 0
        pedidos_em_andamento = 0
        ultimas_transacoes = []
        messages.error(request, f'Erro ao carregar dados financeiros: {str(e)}')
    
    context = {
        'receita_total': receita_total,
        'receita_mes_atual': receita_mes_atual,
        'receita_mes_passado': receita_mes_passado,
        'crescimento_percentual': round(crescimento_percentual, 1),
        'pedidos_concluidos': pedidos_concluidos,
        'pedidos_pendentes': pedidos_pendentes,
        'pedidos_em_andamento': pedidos_em_andamento,
        'ultimas_transacoes': ultimas_transacoes,
        'mes_atual': inicio_mes.strftime('%B %Y'),
    }
    
    return render(request, 'services/provider_financeiro.html', context)


@login_required
def provider_relatorios(request):
    """Página de relatórios do prestador"""
    # Verificar se é prestador
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.user_type != 'professional':
            messages.error(request, 'Acesso negado')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Acesso negado')
        return redirect('home')
    
    # Buscar dados para relatórios
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Count, Avg
    
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    ultimos_6_meses = hoje - timedelta(days=180)
    
    try:
        # Pedidos do prestador
        todos_pedidos = Order.objects.filter(professional=request.user)
        
        # Relatório de desempenho mensal (últimos 6 meses)
        desempenho_mensal = []
        for i in range(6):
            mes_inicio = (inicio_mes - timedelta(days=i*30)).replace(day=1)
            mes_fim = (mes_inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            pedidos_mes = todos_pedidos.filter(
                created_at__date__gte=mes_inicio,
                created_at__date__lte=mes_fim
            )
            
            receita_mes = 0
            for pedido in pedidos_mes.filter(status='completed'):
                try:
                    receita_mes += float(pedido.total_price)
                except:
                    pass
            
            desempenho_mensal.append({
                'mes': mes_inicio.strftime('%b/%y'),
                'pedidos': pedidos_mes.count(),
                'receita': receita_mes,
                'concluidos': pedidos_mes.filter(status='completed').count(),
            })
        
        desempenho_mensal.reverse()
        
        # Serviços mais solicitados
        servicos_populares = CustomService.objects.filter(
            provider=request.user
        ).annotate(
            num_solicitacoes=Count('servicerequestmodal')
        ).order_by('-num_solicitacoes')[:5]
        
        # Taxa de conversão
        total_solicitacoes = 0
        total_concluidos = 0
        try:
            from .models import ServiceRequestModal
            servicos_prestador = CustomService.objects.filter(provider=request.user)
            total_solicitacoes = ServiceRequestModal.objects.filter(service__in=servicos_prestador).count()
            total_concluidos = ServiceRequestModal.objects.filter(
                service__in=servicos_prestador,
                status='completed'
            ).count()
        except:
            pass
        
        taxa_conversao = (total_concluidos / total_solicitacoes * 100) if total_solicitacoes > 0 else 0
        
        # Horários mais solicitados
        horarios_populares = {
            'manha': 0,
            'tarde': 0,
            'noite': 0,
        }
        try:
            from .models import ServiceRequestModal
            solicitacoes = ServiceRequestModal.objects.filter(service__in=servicos_prestador)
            for sol in solicitacoes:
                periodo = sol.preferred_period
                if periodo in horarios_populares:
                    horarios_populares[periodo] += 1
        except:
            pass
        
    except Exception as e:
        desempenho_mensal = []
        servicos_populares = []
        taxa_conversao = 0
        horarios_populares = {'manha': 0, 'tarde': 0, 'noite': 0}
        messages.error(request, f'Erro ao gerar relatórios: {str(e)}')
    
    context = {
        'desempenho_mensal': desempenho_mensal,
        'servicos_populares': servicos_populares,
        'taxa_conversao': round(taxa_conversao, 1),
        'horarios_populares': horarios_populares,
    }
    
    return render(request, 'services/provider_relatorios.html', context)


# API Deprecation Management Views (Requirements: 12.2, 12.3, 12.4)
from .api.deprecation_views import (
    DeprecationDashboardView,
    ScheduleDeprecationView,
    CancelDeprecationView,
    DeprecationStatusView
)

# Expose deprecation views as functions for URL routing
api_deprecation_dashboard = DeprecationDashboardView.as_view()
schedule_deprecation_api = ScheduleDeprecationView.as_view()
cancel_deprecation_api = CancelDeprecationView.as_view()
deprecation_status_api = DeprecationStatusView.as_view()
